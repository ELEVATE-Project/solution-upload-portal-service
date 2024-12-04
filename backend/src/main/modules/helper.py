import os
import time
from configparser import ConfigParser, ExtendedInterpolation
import xlrd
import uuid
import csv
from bson.objectid import ObjectId
import json
from datetime import datetime
import requests
from difflib import get_close_matches
from requests import post, get, delete
import sys
import time
import shutil
from xlutils.copy import copy
import shutil
import re
import pandas as pd
from xlrd import open_workbook
from xlutils.copy import copy as xl_copy
import logging.handlers
import time
from logging.handlers import TimedRotatingFileHandler
import xlsxwriter
import argparse
import sys
from os import path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from common_config import *
import threading
import wget
import gdown

# Global variable declaration
criteriaLookUp = dict()
millisecond = None
programNameInp = None
environment = None
observationId = None
solutionName = None
pointBasedValue = None
entityType = None
allow_multiple_submissions = None
scopeEntityType = ""
programName = None
userEntity = None
roles = ""
mainRole = ""
dictCritLookUp = {}
isProgramnamePresent = None
solutionLanguage = None
keyWords = None
entityTypeId = None
solutionDescription = None
creator = None
dikshaLoginId = None
criteriaName = None
solutionId = None
API_log = None
listOfFoundRoles = []
entityToUpload = None
programID = None
programExternalId = None
programDescription = None
criteriaLookUp = dict()
themesSheetList = []
themeRubricFileObj = dict()
criteriaLevelsReport = False
ecm_sections = dict()
criteriaLevelsCount = 0
numberOfResponses = 0
criteriaIdNameDict = dict()
criteriaLevels = list()
matchedShikshalokamLoginId = None
scopeEntities = []
scopeRoles = []
countImps = 0
ecmToSection = dict()
entitiesPGM = []
entitiesPGMID = []
solutionRolesArr = []
startDateOfResource = None
endDateOfResource = None
startDateOfProgram = None
endDateOfProgram = None
rolesPGM =None
solutionRolesArray = []
solutionStartDate = ""
solutionEndDate = ""
projectCreator = ""
orgIds = []
OrgName = []
ccRootOrgName = None
ccRootOrgId  = None
certificatetemplateid = None
question_sequence_arr = []
regex = "\"?([-a-zA-Z0-9.`?{}]+@\w+\.\w+)\"?"
downloaded_file = None
addObservationSolution = None
class Helpers:
    def __init__(self):
        self.millisecond = None
        self.scopeEntityType = ""

    def programCreation(accessToken, parentFolder, externalId, pName, pDescription, keywords, entities, roles, orgIds,creatorKeyCloakId, creatorName,entitiesPGM,mainRole,rolesPGM):
        messageArr = []
        messageArr.append("++++++++++++ Program Creation ++++++++++++")
        # program creation url 
        programCreationurl =  internal_kong_ip + programcreationurl
        messageArr.append("Pogram Creation URL : " + programCreationurl)
        # program creation payload
        payload = json.dumps({
            "externalId": externalId,
            "name": pName,
            "description": pDescription,
            "resourceType": [
                "program"
            ],
            "language": [
                "English"
            ],
            "keywords": keywords,
            "concepts": [],
            "createdFor": orgIds,
            "rootOrganisations": orgIds,
            "imageCompression": {
                "quality": 10
            },
            "creator": creatorName,
            "owner": creatorKeyCloakId,
            "author": creatorKeyCloakId,
            "scope": {
                "entityType": scopeEntityType,
                "entities": entities,
                "roles": roles
            }})
        messageArr.append("Body : " + str(payload))
        headers = {'X-authenticated-user-token': accessToken,
                   'internal-access-token': internal_access_token,
                   'Content-Type': 'application/json',
                   'Authorization':authorization}

        # program creation 
        responsePgmCreate = requests.request("POST", programCreationurl, headers=headers, data=(payload))
        messageArr.append("Program Creation Status Code : " + str(responsePgmCreate.status_code))
        messageArr.append("Program Creation Response : " + str(responsePgmCreate.text))
        messageArr.append("Program body : " + str(payload))

        # save logs 
        # createAPILog(parentFolder, messageArr)
        # # check status 
        # fileheader = [pName, ('Program Sheet Validation'), ('Passed')]
        # createAPILog(parentFolder, messageArr)
        # apicheckslog(parentFolder, fileheader)
        print(responsePgmCreate.text,responsePgmCreate)
        if responsePgmCreate.status_code == 200:
            responsePgmCreateResp = responsePgmCreate.json()
        else:
            # terminate execution
            print("Program creation API failed. Please check logs.")


    def programmappingpdpmsheetcreation(MainFilePath,accessToken, program_file,programexternalId,parentFolder):
        pdpmsheet = MainFilePath+ "/pdpmmapping/"
        if not os.path.exists(pdpmsheet):
            os.mkdir(pdpmsheet)

        wbproject = xlrd.open_workbook(program_file, on_demand=True)
        projectSheetNames = wbproject.sheet_names()

        mappingsheet = wbproject.sheet_by_name('Program Details')
        keysProject = [mappingsheet.cell(1, col_index_env).value for col_index_env in
                       range(mappingsheet.ncols)]

        pdpmcolo1 = ["user","role","entity","entityOperation","keycloak-userId","acl_school","acl_cluster","programOperation",
                    "platform_role","programs","_arrayFields"]
        with open(pdpmsheet + 'mapping.csv', 'w',encoding='utf-8') as file:
             writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
             writer.writerows([pdpmcolo1])

        wbPgm = xlrd.open_workbook(program_file, on_demand=True)
        global programNameInp
        sheetNames = wbPgm.sheet_names()
        for sheetEnv in sheetNames:
            if sheetEnv == "Instructions":
                pass
            elif sheetEnv.strip().lower() == 'program details':
                print("--->Checking Program details sheet...")
                detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                      for
                                      col_index_env in range(detailsEnvSheet.ncols)}
                    programNameInp = dictDetailsEnv['Title of the Program'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Title of the Program'] else Helpers.terminatingMessage("\"Title of the Program\" must not be Empty in \"Program details\" sheet")

                extIdPGM = dictDetailsEnv['Program ID'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Program ID'] else Helpers.terminatingMessage("\"Program ID\" must not be Empty in \"Program details\" sheet")

                programdesigner = dictDetailsEnv['Diksha username/user id/email id/phone no. of Program Designer'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Program ID'] else Helpers.terminatingMessage("\"Diksha username/user id/email id/phone no. of Program Designer\" must not be Empty in \"Program details\" sheet")
                userDetails = Helpers.fetchUserDetails(accessToken, programdesigner)

                creatorKeyCloakId = userDetails[0]
                creatorName = userDetails[1]
                if "PROGRAM_DESIGNER" in userDetails[3]:
                    creatorKeyCloakId = userDetails[0]
                    creatorName = userDetails[1]
                else :
                    print("user does't have program designer role")

                pdpmcolo1 = [creatorName, " ", " ", " ", creatorKeyCloakId, " ", " ","ADD","PROGRAM_DESIGNER", extIdPGM, "programs"]
                with open(pdpmsheet + 'mapping.csv', 'a',encoding='utf-8') as file:
                    writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                    writer.writerows([pdpmcolo1])
                    fileheader = [creatorName,"program designer mapped successfully","Passed"]
                    # apicheckslog(parentFolder,fileheader)


            elif sheetEnv.strip().lower() == 'program manager details':
                print("--->Program Manager Details...")
                detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                      for
                                      col_index_env in range(detailsEnvSheet.ncols)}

                    if str(dictDetailsEnv['Is a SSO user?']).strip() == "YES":
                        programmanagername2 = dictDetailsEnv['Sunbird user id ( profile ID)'] if dictDetailsEnv['Sunbird user id ( profile ID)'] else Helpers.terminatingMessage("\"Sunbird user id ( profile ID)\" must not be Empty in \"Program details\" sheet")
                    else:
                        try :
                            programmanagername2 = dictDetailsEnv['Login ID on Sunbird'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Login ID on Sunbird'] else Helpers.terminatingMessage("\"Login ID on Sunbird\" must not be Empty in \"Program details\" sheet")
                            userDetails = Helpers.fetchUserDetails(accessToken, programmanagername2)
                        except :
                            programmanagername2 = dictDetailsEnv['Sunbird user id ( profile ID)'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Sunbird user id ( profile ID)'] else Helpers.terminatingMessage("\"Sunbird user id ( profile ID)\" must not be Empty in \"Program details\" sheet")
                            userDetails = Helpers.fetchUserDetails(accessToken, programmanagername2)

                    userDetails = Helpers.fetchUserDetails(accessToken, programmanagername2)
                    creatorKeyCloakId = userDetails[0]
                    creatorName = userDetails[1]
                    if "PROGRAM_MANAGER" in userDetails[3]:
                        creatorKeyCloakId = userDetails[0]
                        creatorName = userDetails[1]
                    else:
                        print("user does't have program manager role")

                    pdpmcolo1 = [creatorName, " ", " ", " ", creatorKeyCloakId, " ", " ","ADD","PROGRAM_MANAGER", extIdPGM, "programs"]

                    with open(pdpmsheet + 'mapping.csv', 'a',encoding='utf-8') as file:
                        writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                        writer.writerows([pdpmcolo1])
                    # messageArr.append("Response : " + str(pdpmcolo1))
                    # Helpers.createAPILog(parentFolder, messageArr)

                    fileheader = [creatorName,"program manager mapped succesfully","Passed"]
                    # apicheckslog(parentFolder,fileheader)


    # this function is used for call the api and map the pdpm roles which we created
    def Programmappingapicall(MainFilePath,accessToken, program_file,parentFolder):
        urlpdpmapi = internal_kong_ip + pdpmurl
        headerpdpmApi = {
            'Authorization':authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            'internal-access-token': internal_access_token
        }
        payload = {}
        filesProject = {
            'userRoles': open(MainFilePath + '/pdpmmapping/mapping.csv', 'rb')
        }

        responseProgrammappingApi = requests.post(url=urlpdpmapi, headers=headerpdpmApi,
                                                 data=payload,
                                                 files=filesProject)
        messageArr = ["program mapping sheet.",
                      "File path : " + MainFilePath + '/pdpmmapping/mapping.csv']
        messageArr.append("Upload status code : " + str(responseProgrammappingApi.status_code))
        Helpers.createAPILog(parentFolder, messageArr)

        if responseProgrammappingApi.status_code == 200:
            print('--->program manager and designer mapping is Success')
            with open(MainFilePath + '/pdpmmapping/mappinginternal.csv', 'w+',encoding='utf-8') as projectRes:
                projectRes.write(responseProgrammappingApi.text)
                messageArr.append("Response : " + str(responseProgrammappingApi.text))
                Helpers.createAPILog(parentFolder, messageArr)
        else:
            messageArr.append("Response : " + str(responseProgrammappingApi.text))
            Helpers.createAPILog(parentFolder, messageArr)
            fileheader = ["PDPM mapping","PDPM mapping is failed","Failed","check PDPM sheet"]
            Helpers.apicheckslog(parentFolder,fileheader)
            sys.exit()


    
    def createFileStructForProgram(programFile):
        #  print("programFile:-------------",programFile)
        if not os.path.isdir('programFiles'):
            os.mkdir('programFiles')
        if "/" in str(programFile):
            fileNameSplit = str(programFile).split('/')[-1:]
            # print("newfileNameSplit ;",fileNameSplit)
        else :

            fileNameSplit = os.path.basename(programFile)
        # print("updatedfileNameSplit : ",fileNameSplit)

        if isinstance(fileNameSplit, list):
            fileNameSplit = fileNameSplit[0]
        # fileNameSplit = str(programFile)
        if fileNameSplit.endswith(".xlsx"):
            # print("latest fileNameSplit",fileNameSplit)
            ts = str(time.time()).replace(".", "_")
            
            folderName = fileNameSplit.replace(".xlsx", "-" + str(ts))
            # print("folderName :",folderName)
            os.mkdir('programFiles/' + str(folderName))
            path = os.path.join('programFiles', str(folderName))
            # print("done",path)
        else:
            print("something")
        returnPathStr = os.path.join('programFiles', str(folderName))
        # print("returnPathStr",returnPathStr)

        return returnPathStr
    
    def fetchScopeRole(solutionName_for_folder_path, accessToken, roleNameList):
        urlFetchRolesListApi = internal_kong_ip + listofrolesapi
        headerFetchRolesListApi = {
            'Content-Type':  content_type,
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
        }
        responseFetchRolesListApi = requests.post(url=urlFetchRolesListApi, headers=headerFetchRolesListApi)
        rolesLookup = dict()
        rolesReturn = list()
        messageArr = ["Roles list fetch API called.", "URL  : " + str(urlFetchRolesListApi),
                      "Status Code : " + str(responseFetchRolesListApi.status_code)]
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        if responseFetchRolesListApi.status_code == 200:
            responseFetchRolesListApi = responseFetchRolesListApi.json()
            for listRoles in responseFetchRolesListApi['result']:
                eachDict = dict()
                eachDict['id'] = listRoles['_id'].lstrip().rstrip()
                eachDict['code'] = listRoles['code'].lstrip().rstrip()
                rolesLookup[listRoles['code']] = eachDict['id']
                rolesReturn.append(listRoles['code'].lstrip().rstrip())
        else:
            print("---> error in subroles API.")

        userRolesFromInp = roleNameList
        listOfFoundRoles = list()
        if len(userRolesFromInp) == 0:
            print("Roles fields must not be empty.")
        for ur in userRolesFromInp:
            rolesFlag = True
            try:
                roleDetails = rolesLookup[ur.lstrip().rstrip()]
                rolesFlag = True
            except:
                rolesFlag = False

            if rolesFlag:
                print("Role Found... : " + ur)
                listOfFoundRoles.append(ur)
            else:
                if "all" in userRolesFromInp:
                    listOfFoundRoles = ["all"]
                else:
                    print("Role error...")
                    print("Role : " + ur)
                    messageArr = ["Roles Error", "URL  : ", "Role : " + ur]
                    Helpers.createAPILog(solutionName_for_folder_path, messageArr)

        messageArr = ["Accepted Roles : " + str(listOfFoundRoles)]
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        if len(listOfFoundRoles) == 0:
            messageArr = ["No roles matched our DB "]
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print("No Roles matched our DB.")
        return listOfFoundRoles


    
    def getProgramInfo(accessTokenUser, solutionName_for_folder_path, programNameInp):
        global programID, programExternalId, programDescription, isProgramnamePresent, programName
        programName = programNameInp
        programUrl = internal_kong_ip + fetchprograminfoapiurl
        # print(programUrl)
        payload = json.dumps({
            "query": {
                "name": programNameInp,
                "isAPrivateProgram": False,
                "status": "active"
                },
                "mongoIdKeys": []
                })
        
        # print(programNameInp,"programNameInp")

        headersProgramSearch = {'Authorization': authorization,
                                'Content-Type': 'application/json', 'X-authenticated-user-token': accessTokenUser,
                                'internal-access-token': internal_access_token}
        responseProgramSearch = requests.post(url=programUrl, headers=headersProgramSearch,data=payload)
        # print(responseProgramSearch.text)
        messageArr = []

        if responseProgramSearch.status_code == 200:
            print('--->Program fetch API Success')
            messageArr.append("--->Program fetch API Success")
            responseProgramSearch = responseProgramSearch.json()
            countOfPrograms = len(responseProgramSearch['result'])
            messageArr.append("--->Program Count : " + str(countOfPrograms))
            if countOfPrograms == 0:
                messageArr.append("No program found with the name : " + str(programName))
                messageArr.append("******************** Preparing for program Upload **********************")
                print("No program found with the name : " + str(programName))
                print("******************** Preparing for program Upload **********************")
                
                return False
            else:
                getProgramDetails = []
                for eachPgm in responseProgramSearch['result']:
                    if eachPgm['isAPrivateProgram'] == False:
                        programID = eachPgm['_id']
                        programExternalId = eachPgm['externalId']
                        programDescription = eachPgm['description']
                        isAPrivateProgram = eachPgm['isAPrivateProgram']
                        getProgramDetails.append([programID, programExternalId, programDescription, isAPrivateProgram])
                        if len(getProgramDetails) == 0:
                            print("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                            messageArr.append("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                            
                            fileheader = ["program find api is running","found"+str(len(
                                getProgramDetails))+"programs in backend","Failed","found"+str(len(
                                getProgramDetails))+"programs ,check logs"]
                           
                        elif len(getProgramDetails) > 1:
                            print("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                            messageArr.append("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                           

                        else:
                            programID = getProgramDetails[0][0]
                            programExternalId = getProgramDetails[0][1]
                            programDescription = getProgramDetails[0][2]
                            isAPrivateProgram = getProgramDetails[0][3]
                            isProgramnamePresent = True
                            messageArr.append("programID : " + str(programID))
                            messageArr.append("programExternalId : " + str(programExternalId))
                            messageArr.append("programDescription : " + str(programDescription))
                            messageArr.append("isAPrivateProgram : " + str(isAPrivateProgram))
                        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        else:
            print("Program search API failed...")
            messageArr.append("Program search API failed...")
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            Helpers.terminatingMessage("Response Code : " + str(responseProgramSearch.status_code))
        return True
    
    def fetchEntityId(solutionName_for_folder_path, accessToken, entitiesNameList, scopeEntityType):
        # print(scopeEntityType,"scopeEntityType--------------")
        urlFetchEntityListApi = host+searchforlocation
        headerFetchEntityListApi = {
            'Content-Type': content_type,
            'Authorization': authorizationforhost,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
        }
        payload = {
            "request": {
                "filters": {
                    "type": scopeEntityType
                },
                "limit": 1000
            }
        }
        responseFetchEntityListApi = requests.post(url=urlFetchEntityListApi, headers=headerFetchEntityListApi,data=json.dumps(payload))
        # print(responseFetchEntityListApi,"responseFetchEntityListApi")

        messageArr = ["Entities List Fetch API executed.", "URL  : " + str(urlFetchEntityListApi),
                      "Status : " + str(responseFetchEntityListApi.status_code)]
       
        if responseFetchEntityListApi.status_code == 200:
            responseFetchEntityListApi = responseFetchEntityListApi.json()
            entitiesLookup = dict()
            entityToUpload = list()
            for listEntities in responseFetchEntityListApi['result']['response']:
                entitiesLookup[listEntities['name'].lower().lstrip().rstrip()] = listEntities['id'].lstrip().rstrip()
            entitiesFlag = False
            for eachUserEntity in entitiesNameList:
                try:
                    entityId = entitiesLookup[eachUserEntity.lower().lstrip().rstrip()]
                    entitiesFlag = True
                except:
                    entitiesFlag = False
                if entitiesFlag:
                    entityToUpload.append(entityId)
                else:
                    print("Entity Not found in DB...")
                    print("Entity name : " + str(eachUserEntity))
                    messageArr = ["Entity Not found : ", "URL  : " + str(eachUserEntity)]
                    

            messageArr = ["Entities to upload : " + str(entityToUpload)]
            
            if len(entityToUpload) == 0:
                print("--->Scope Entity error.")
            return entityToUpload
        else:
            messageArr = ["Error in Location search",str(responseFetchEntityListApi.status_code)]
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            Helpers.terminatingMessage("---> Error in location search.")

    
    def programsFileCheck(filePathAddPgm, accessToken, parentFolder, MainFilePath):
        program_file = filePathAddPgm
        # open excel file 
        wbPgm = xlrd.open_workbook(filePathAddPgm, on_demand=True)
        global programNameInp
        sheetNames = wbPgm.sheet_names()
        # list of sheets in the program sheet 
        pgmSheets = ["Instructions", "Program Details", "Resource Details","Program Manager Details"]

        # checking the sheets in the program sheet 
        if (len(sheetNames) == len(pgmSheets)) and ((set(sheetNames) == set(pgmSheets))):
            print("--->Program Template detected.<---")
            # iterate through the sheets 
            for sheetEnv in sheetNames:

                if sheetEnv == "Instructions":
                    # skip Instructions sheet 
                    pass
                elif sheetEnv.strip().lower() == 'program details':
                    print("--->Checking Program details sheet...")
                    detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                    keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                               range(detailsEnvSheet.ncols)]
                    for row_index_env in range(2, detailsEnvSheet.nrows):
                        dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                          for
                                          col_index_env in range(detailsEnvSheet.ncols)}
                        programNameInp = dictDetailsEnv['Title of the Program'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Title of the Program'] else Helpers.terminatingMessage("\"Title of the Program\" must not be Empty in \"Program details\" sheet")
                        extIdPGM = dictDetailsEnv['Program ID'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Program ID'] else Helpers.terminatingMessage("\"Program ID\" must not be Empty in \"Program details\" sheet")
                        returnvalues = []
                        global entitiesPGM
                        entitiesPGM = dictDetailsEnv['Targeted state at program level'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Targeted state at program level'] else Helpers.terminatingMessage("\"Targeted state at program level\" must not be Empty in \"Program details\" sheet")
                        districtentitiesPGM = dictDetailsEnv['Targeted district at program level'].encode('utf-8').decode('utf-8')
                        global startDateOfProgram, endDateOfProgram
                        startDateOfProgram = dictDetailsEnv['Start date of program']
                        endDateOfProgram = dictDetailsEnv['End date of program']
                        # taking the start date of program from program template and converting YYYY-MM-DD 00:00:00 format

                        startDateArr = str(startDateOfProgram).split("-")
                        startDateOfProgram = startDateArr[2] + "-" + startDateArr[1] + "-" + startDateArr[0] + " 00:00:00"

                        # taking the end date of program from program template and converting YYYY-MM-DD 00:00:00 format

                        endDateArr = str(endDateOfProgram).split("-")
                        endDateOfProgram = endDateArr[2] + "-" + endDateArr[1] + "-" + endDateArr[0] + " 23:59:59"

                        global scopeEntityType
                        scopeEntityType = "state"


                        if districtentitiesPGM:
                            entitiesPGM = districtentitiesPGM
                            EntityType = "district"
                        else:
                            entitiesPGM = entitiesPGM
                            EntityType = "state"

                        scopeEntityType = EntityType

                        global entitiesPGMID
                        # print(entitiesPGMID,"entitiesPGMID")
                        entitiesPGMID = Helpers.fetchEntityId(parentFolder, accessToken,
                                                      entitiesPGM.lstrip().rstrip().split(","), scopeEntityType)
                        print(entitiesPGMID)
                        global orgIds



                        if not Helpers.getProgramInfo(accessToken, parentFolder, programNameInp.encode('utf-8').decode('utf-8')):
                            # print("reached till here")
                            extIdPGM = dictDetailsEnv['Program ID'].encode('utf-8').decode('utf-8')
                            if str(dictDetailsEnv['Program ID']).strip() == "Do not fill this field":
                                print("change the program id")
                            descriptionPGM = dictDetailsEnv['Description of the Program'].encode('utf-8').decode('utf-8')
                            keywordsPGM = dictDetailsEnv['Keywords'].encode('utf-8').decode('utf-8')
                            entitiesPGM = dictDetailsEnv['Targeted state at program level'].encode('utf-8').decode('utf-8') 
                            districtentitiesPGM = dictDetailsEnv['Targeted district at program level'].encode('utf-8').decode('utf-8')
                            # selecting entity type based on the users input 
                            if districtentitiesPGM:
                                entitiesPGM = districtentitiesPGM
                                EntityType = "district"
                            else:
                                entitiesPGM = entitiesPGM
                                EntityType = "state"

                            scopeEntityType = EntityType

                            mainRole = dictDetailsEnv['Targeted role at program level'] 
                            # print(mainRole,"mainRole")
                            global rolesPGM
                            rolesPGM = dictDetailsEnv['Targeted subrole at program level']
                            # print(rolesPGM,rolesPGM)

                            if "teacher" in mainRole.strip().lower():
                                rolesPGM = str(rolesPGM).strip() + ",TEACHER"
                            userDetails = Helpers.fetchUserDetails(accessToken, dictDetailsEnv['Diksha username/user id/email id/phone no. of Program Designer'])
                            OrgName=userDetails[4]
                            # print(OrgName,"OrgName")
                            orgIds=Helpers.fetchOrgId(accessToken, parentFolder, OrgName)
                            print(orgIds,"orgIds")
                            creatorKeyCloakId = userDetails[0]
                            creatorName = userDetails[2]

                            messageArr = []

                            scopeEntityType = EntityType
                            # fetch entity details 
                            entitiesPGMID = Helpers.fetchEntityId(parentFolder, accessToken,entitiesPGM.lstrip().rstrip().split(","), scopeEntityType)
                            # print(entitiesPGMID,"entitiesPGMID")

                            # sys.exit()
                            # fetch sub-role details 
                            rolesPGMID = Helpers.fetchScopeRole(parentFolder, accessToken, rolesPGM.lstrip().rstrip().split(","))
                            # print(rolesPGMID,"rolesPGMID")

                            # sys.exit()

                            # call function to create program 
                            Helpers.programCreation(accessToken, parentFolder, extIdPGM, programNameInp, descriptionPGM,keywordsPGM.lstrip().rstrip().split(","), entitiesPGMID, rolesPGMID, orgIds,creatorKeyCloakId, creatorName,entitiesPGM,mainRole,rolesPGM)
                            # sys.exit()
                            Helpers.programmappingpdpmsheetcreation(MainFilePath, accessToken, program_file, extIdPGM,parentFolder)

                            # map PM / PD to the program 
                            Helpers.Programmappingapicall(MainFilePath, accessToken, program_file,parentFolder)

                            # check if program is created or not 
                            if Helpers.getProgramInfo(accessToken, parentFolder, programNameInp):
                                print("Program Created SuccessFully.")
                            else :
                                print("Program creation failed! Please check logs.")

                elif sheetEnv.strip().lower() == 'resource details':
                    # checking Resource details sheet 
                    print("--->Checking Resource Details sheet...")
                    detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                    keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                               range(detailsEnvSheet.ncols)]
                    # iterate through each row in Resource Details sheet and validate 
                    for row_index_env in range(2, detailsEnvSheet.nrows):
                        dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                          for
                                          col_index_env in range(detailsEnvSheet.ncols)}
                        resourceNamePGM = dictDetailsEnv['Name of resources in program'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Name of resources in program'] else Helpers.terminatingMessage("\"Name of resources in program\" must not be Empty in \"Resource Details\" sheet")
                        resourceTypePGM = dictDetailsEnv['Type of resources'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Type of resources'] else Helpers.terminatingMessage("\"Type of resources\" must not be Empty in \"Resource Details\" sheet")
                        resourceLinkOrExtPGM = dictDetailsEnv['Resource Link']
                        resourceStatusOrExtPGM = dictDetailsEnv['Resource Status'] if dictDetailsEnv['Resource Status'] else Helpers.terminatingMessage("\"Resource Status\" must not be Empty in \"Resource Details\" sheet")
                        # setting start and end dates globally. 
                        global startDateOfResource, endDateOfResource
                        startDateOfResource = dictDetailsEnv['Start date of resource']
                        endDateOfResource = dictDetailsEnv['End date of resource']
                        # checking resource types and calling relevant functions 
                        # if resourceTypePGM.lstrip().rstrip().lower() == "course":
                        #     coursemapping = courseMapToProgram(accessToken, resourceLinkOrExtPGM, parentFolder)
                        #     if startDateOfResource:
                        #         startDateArr = str(startDateOfResource).split("-")
                        #         bodySolutionUpdate = {"startDate": startDateArr[2] + "-" + startDateArr[1] + "-" + startDateArr[0] + " 00:00:00"}
                        #         solutionUpdate(parentFolder, accessToken, coursemapping, bodySolutionUpdate)
                        #     if endDateOfResource:
                        #         endDateArr = str(endDateOfResource).split("-")
                        #         bodySolutionUpdate = {
                        #             "endDate": endDateArr[2] + "-" + endDateArr[1] + "-" + endDateArr[0] + " 23:59:59"}
                        #         solutionUpdate(parentFolder, accessToken, coursemapping, bodySolutionUpdate)
                        


# Function create File structure for Solutions
    def createFileStruct(MainFilePath, addSolutionFile):
        if not os.path.isdir(MainFilePath + '/SolutionFiles'):
            os.mkdir(MainFilePath + '/SolutionFiles')
        if "\\" in str(addSolutionFile):
            fileNameSplit = str(addSolutionFile).split('\\')[-1:]
        elif "/" in str(addSolutionFile):
            fileNameSplit = str(addSolutionFile).split('/')[-1:]
        else:
            fileNameSplit = str(addSolutionFile)
        if ".xlsx" in str(fileNameSplit[0]):
            ts = str(time.time()).replace(".", "_")
            folderName = fileNameSplit[0].replace(".xlsx", "-" + str(ts))
            os.mkdir(MainFilePath + '/SolutionFiles/' + str(folderName))
            path = os.path.join(MainFilePath + '/SolutionFiles', str(folderName))
            path = os.path.join(path, str('apiHitLogs'))
            os.mkdir(path)
        else:
            print("File Error.offff")
        returnPathStr = os.path.join(MainFilePath + '/SolutionFiles', str(folderName))

        if not os.path.isdir(returnPathStr + "/user_input_file"):
            os.mkdir(returnPathStr + "/user_input_file")

        shutil.copy(addSolutionFile, os.path.join(returnPathStr + "user_input_file.xlsx"))
        # shutil.copy(programFile, os.path.join(returnPathStr + "user_input_file"))
        return returnPathStr
    

    # Generate access token for the APIs. 
    def generateAccessToken(solutionName_for_folder_path):
    # production search user api - start
        headerKeyClockUser = {'Content-Type': keyclockapicontent_type}
    
        responseKeyClockUser = requests.post(url=host + (keyclockapiurl), headers=headerKeyClockUser,
                                         data=(keyclockapibody))
        print(responseKeyClockUser)
    
        if responseKeyClockUser.status_code == 200:
            responseKeyClockUser = responseKeyClockUser.json()
            accessTokenUser = responseKeyClockUser['access_token']
            print("--->Access Token Generated!")
        return accessTokenUser


    def checkEmailValidation(email):
        if (re.search(regex, email)):
            return True
        else:
            return False
    def fetchUserDetails(accessToken, dikshaId):
        global OrgName,creatorId
        url =  host + userinfoapiurl
        headers = {'Content-Type': 'application/json',
               'Authorization': authorizationforhost,
               'X-authenticated-user-token': accessToken}
        isEmail = Helpers.checkEmailValidation(dikshaId.lstrip().rstrip())
        
        if isEmail:
            body = "{\n  \"request\": {\n    \"filters\": {\n    \t\"email\": \"" + dikshaId.lstrip().rstrip() + "\"\n    },\n      \"fields\" :[],\n    \"limit\": 1000,\n    \"sort_by\": {\"createdDate\": \"desc\"}\n  }\n}"
        else:
            body = "{\n  \"request\": {\n    \"filters\": {\n    \t\"userName\": \"" + dikshaId.lstrip().rstrip() + "\"\n    },\n      \"fields\" :[],\n    \"limit\": 1000,\n    \"sort_by\": {\"createdDate\": \"desc\"}\n  }\n}"

        responseUserSearch = requests.request("POST", url, headers=headers, data=body)
        response_json = responseUserSearch.json()
        print(responseUserSearch.text)
        # print(json.dumps(response_json, indent=4))
        #sys.exit()
        print(responseUserSearch, "---------------------------------------------------------------")
        if responseUserSearch.status_code == 200:
            responseUserSearch = responseUserSearch.json()
            if responseUserSearch['result']['response']['content']:
                userKeycloak = responseUserSearch['result']['response']['content'][0]['userId']
                creatorId = userKeycloak
                userName = responseUserSearch['result']['response']['content'][0]['userName']
                firstName = responseUserSearch['result']['response']['content'][0]['firstName']
                rootOrgId = responseUserSearch['result']['response']['content'][0]['rootOrgId']
                for index in responseUserSearch['result']['response']['content'][0]['organisations']:
                    if rootOrgId == index['organisationId']:
                        roledetails = index['roles']
                        rootOrgName = index['orgName']
                        OrgName.append(rootOrgName)
                print(roledetails)
            else:
                print("-->Given username/email is not present in KB platform<--.")
        else:
            print(responseUserSearch.text)

        return [userKeycloak, userName, firstName,roledetails,rootOrgName,rootOrgId]
    
    def SolutionFileCheck(filePathAddPgm, accessToken, parentFolder, MainFilePath):
        global creatorId,solutionNameForSuccess
        wbPgm = xlrd.open_workbook(filePathAddPgm, on_demand=True)
        global solutionNameInp
        sheetNames = wbPgm.sheet_names()
        for sheetEnv in sheetNames:
            if sheetEnv.strip().lower() == 'details':
                print("--->Checking resource details sheet...")
                detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                            range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                        for
                                        col_index_env in range(detailsEnvSheet.ncols)}
                    solutionNameInp = dictDetailsEnv['solution_name'].encode('utf-8').decode('utf-8')
                    solutionNameForSuccess = solutionNameInp
                    global entitiesPGM

                    global startDateOfProgram, endDateOfProgram
                    startDateOfProgram = dictDetailsEnv['start_date']
                    endDateOfProgram = dictDetailsEnv['end_date']

                    # taking the start date of program from program template and converting YYYY-MM-DD 00:00:00 format

                    startDateArr = str(startDateOfProgram).split("-")
                    startDateOfProgram = startDateArr[2] + "-" + startDateArr[1] + "-" + startDateArr[0] + " 00:00:00"

                    # taking the end date of program from program template and converting YYYY-MM-DD 00:00:00 format

                    endDateArr = str(endDateOfProgram).split("-")
                    endDateOfProgram = endDateArr[2] + "-" + endDateArr[1] + "-" + endDateArr[0] + " 23:59:59"
                    if not Helpers.getProgramInfo(accessToken, parentFolder, solutionNameInp.encode('utf-8').decode('utf-8')):
                        extIdPGM = dictDetailsEnv['solution_name'].encode('utf-8').decode('utf-8')
                        programName = extIdPGM = dictDetailsEnv['solution_name'].encode('utf-8').decode('utf-8')
                        userDetails = Helpers.fetchUserDetails(accessToken, dictDetailsEnv['creator_username'])
                        OrgName=userDetails[4]
                        print(OrgName,"OrgName")
                        orgIds=Helpers.fetchOrgId(accessToken, parentFolder, OrgName)
                        creatorKeyCloakId = userDetails[0]
                        creatorName = userDetails[2]
                        if Helpers.getProgramInfo(accessToken, parentFolder, extIdPGM):
                            print("Program Created SuccessFully.")
                        else :
                            print("program creation API called")
                            Helpers.programCreation(accessToken, parentFolder, extIdPGM, programName,orgIds,creatorKeyCloakId, creatorName)

    def prepareProjectAndTasksSheets(project_inputFile, projectName_for_folder_path, accessToken):
        millisecond = int(time.time() * 1000)
        PreviousTaskname = None
        PreviousTaskid = None
        projectFilePath = projectName_for_folder_path + '/projectUpload/'
        taskFilePath = projectName_for_folder_path + '/taskUpload/'
        file_exists = os.path.isfile(projectName_for_folder_path + '/projectUpload/projectUpload.csv')
        if not os.path.exists(projectFilePath):
            os.mkdir(projectFilePath)
        if not os.path.exists(taskFilePath):
            os.mkdir(taskFilePath)

        wbproject = xlrd.open_workbook(project_inputFile, on_demand=True)
        projectSheetNames = wbproject.sheet_names()

        projectDetailsSheet = wbproject.sheet_by_name('Project upload')
        keysProject = [projectDetailsSheet.cell(1, col_index_env).value for col_index_env in
                       range(projectDetailsSheet.ncols)]
        projectColnames1 = ["title", "externalId", "categories","recommendedFor", "description", "entityType", "goal"]
        learningResource_count = 0
        for projectHeader in keysProject:
            if str(projectHeader).startswith('learningResources'):

                learningResource_count += 1
        learningResource_count = int(learningResource_count) / 2

        lr_count = 1
        for lr in range(0, int(learningResource_count)):
            projectColnames1.append("learningResources" + str(lr_count) + "-name")
            projectColnames1.append("learningResources" + str(lr_count) + "-link")
            projectColnames1.append("learningResources" + str(lr_count) + "-app")
            projectColnames1.append("learningResources" + str(lr_count) + "-id")
            lr_count += 1
        projectColnames2 = ["rationale", "primaryAudience", "taskCreationForm", "duration", "concepts", "keywords","successIndicators", "risks", "approaches", "_arrayFields"]
        for columns in projectColnames2:
            projectColnames1.append(columns)
        with open(projectFilePath + 'projectUpload.csv', 'w',encoding='utf-8') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
            writer.writerows([projectColnames1])

        for row_index_env in range(2, projectDetailsSheet.nrows):
            dictProjectDetails = {keysProject[col_index_env]: projectDetailsSheet.cell(row_index_env, col_index_env).value
                                  for col_index_env in range(projectDetailsSheet.ncols)}
            title = str(dictProjectDetails["title"]).encode('utf-8').decode('utf-8').strip()
            externalId = str(dictProjectDetails["projectId"]).strip() + "-" + str(millisecond)
            categories_list = ["teachers", "students", "infrastructure", "community", "educationLeader", "schoolProcess"]
            categories = str(dictProjectDetails["categories"]).encode('utf-8').decode('utf-8').split(",")
            categories_final = ""
            projectGoal = "TEMP"
            for cat in categories:
                if categories_final == "":
                    categories_final = categories_final + str(
                        (get_close_matches(cat.strip().lower().replace(" ", ""), categories_list)[0]))
                else:
                    categories_final = categories_final + "," + str(
                        (get_close_matches(cat.strip().lower().replace(" ", ""), categories_list)[0]))
            global projectCreator, projectAuthor

            projectAuthor = str(dictProjectDetails["Diksha_loginId"]).encode('utf-8').decode('utf-8').strip()
            recommendedFor = str(dictProjectDetails["recommendedFor"]).encode('utf-8').decode('utf-8').strip()
            objective = str(dictProjectDetails["objective"]).encode('utf-8').decode('utf-8').strip()
            entityType = None
            project_values = [title, externalId, categories_final,recommendedFor, objective, entityType,projectGoal]
            lr_value_count = 1
            for lr in range(0, int(learningResource_count)):
                lr_name = str(dictProjectDetails["learningResources" + str(lr_value_count) + "-name"]).strip()
                lr_link = str(dictProjectDetails["learningResources" + str(lr_value_count) + "-link"]).strip()
                if lr_name == "" and lr_link == "":
                    project_values.append("")
                    project_values.append("")
                    project_values.append("")
                    project_values.append("")
                    lr_value_count += 1
                else:
                    project_values.append(lr_name)
                    lr_link_id = lr_link.split("/")[-1]
                    project_values.append(lr_link)
                    project_values.append("Diksha")
                    project_values.append(lr_link_id)
                    lr_value_count += 1
            remaining_project_values = ["rationale", "primaryAudience", "taskCreationForm", "duration", "concepts",
                                        "keywords", "successIndicators", "risks", "approaches", "_arrayFields"] 
            for values in remaining_project_values:
                try:
                    project_values.append(str(dictProjectDetails[values]).strip())
                except:
                    if values == "_arrayFields":
                        project_values.append(
                            "categories,recommendedFor,primaryAudience,successIndicators,risks,approaches")
                    else:
                        project_values.append("")

            with open(projectFilePath + 'projectUpload.csv','a',encoding='utf-8') as file:
                writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                writer.writerows([project_values])

        tasksDetailsSheet = wbproject.sheet_by_name('Tasks upload')
        keysTasks = [tasksDetailsSheet.cell(1, col_index_env).value for col_index_env in
                     range(tasksDetailsSheet.ncols)]
        taskColumns1 = ["name", "externalId", "description", "type", "hasAParentTask", "parentTaskOperator",
                        "parentTaskValue",
                        "parentTaskId", "solutionType", "solutionSubType", "solutionId", "isDeletable"]
        taskLearningResource_count = 0

        for tasksHeader in keysTasks:
            if str(tasksHeader).startswith('learningResources'):
                taskLearningResource_count += 1
        taskLearningResource_count = int(taskLearningResource_count) / 2
        taskslr_count = 1
        for lr in range(0, int(taskLearningResource_count)):
            taskColumns1.append("learningResources" + str(taskslr_count) + "-name")
            taskColumns1.append("learningResources" + str(taskslr_count) + "-link")
            taskColumns1.append("learningResources" + str(taskslr_count) + "-app")
            taskColumns1.append("learningResources" + str(taskslr_count) + "-id")
            taskslr_count += 1
        taskColumns1.append("minNoOfSubmissionsRequired")
        taskColumns1.append("sequenceNumber")

        with open(taskFilePath + 'taskUpload.csv', 'w',encoding='utf-8') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
            writer.writerows([taskColumns1])
        sequenceNumber = 0
        for row_index_env in range(2, tasksDetailsSheet.nrows):
            dictTasksDetails = {keysTasks[col_index_env]: tasksDetailsSheet.cell(row_index_env, col_index_env).value
                                for col_index_env in range(tasksDetailsSheet.ncols)}
            taskName = str(dictTasksDetails["TaskTitle"]).encode('utf-8').decode('utf-8').strip()
            subtaskname = str(dictTasksDetails["Subtask"]).encode('utf-8').decode('utf-8').strip()

            if dictTasksDetails['TaskId'] :
               taskId = str(dictTasksDetails["TaskId"]).encode('utf-8').decode('utf-8').strip() + "-" + str(millisecond)
               taskminNoOfSubmissionsRequired = str(dictTasksDetails["Number of submissions for observation"]).strip()
               sequenceNumber = sequenceNumber + 1
               taskSolutionType = ""
               try:
                   taskDescription = str(dictTasksDetails["description"]).strip()
               except:
                   taskDescription = ""
               if dictTasksDetails["observation Name"] != "":
                   taskType = "observation"
               elif dictTasksDetails["learningResources1-name"] != "" and dictTasksDetails["learningResources1-link"] != "":
                   taskType = "content"
               else:
                   taskType = "simple"

               hasAParentTask = "NO"
               parentTaskOperator = ""
               parentTaskValue = ""
               parentTaskId = ""

               if dictTasksDetails["observation Name"] != "":
                   solutionNameOrId = dictTasksDetails["observation Name"].encode('utf-8').decode('utf-8')
                   taskSolutionType = "observation"
                   solutionDetailsInTask = Helpers.checkEntityOfSolution(projectName_for_folder_path, solutionNameOrId, accessToken)
                   solutionSubType = solutionDetailsInTask[0]
                   solutionId = solutionDetailsInTask[1]

                   projectUpload = pd.read_csv(projectFilePath + "projectUpload.csv")
                   # updating the column value/data
                   projectUpload.loc[0, 'entityType'] = solutionDetailsInTask[0]

                   # writing into the file
                   projectUpload.to_csv(projectFilePath + "projectUpload.csv", index=False)
               else:
                   solutionId = ""
                   taskSolutionType = ""
                   solutionSubType = ""

               if str(dictTasksDetails["Mandatory task(Yes or No)"]).strip().strip().lower() == "no":
                   isDeletable = "TRUE"
               else:
                   isDeletable = "FALSE"
               task_values = [taskName, taskId, taskDescription, taskType, hasAParentTask, parentTaskOperator, parentTaskValue,
                              parentTaskId, taskSolutionType, solutionSubType, solutionId, isDeletable]
               task_lr_value_count = 1
               for task_lr in range(0, int(taskLearningResource_count)):
                   task_lr_name = str(dictTasksDetails["learningResources" + str(task_lr_value_count) + "-name"]).strip()
                   task_lr_link = str(dictTasksDetails["learningResources" + str(task_lr_value_count) + "-link"]).strip()
                   if task_lr_name == "" and task_lr_link == "":
                       task_values.append("")
                       task_values.append("")
                       task_values.append("")
                       task_values.append("")
                       task_lr_value_count += 1
                   else:
                       task_values.append(task_lr_name)
                       task_lr_link_id = task_lr_link.split("/")[-1]
                       task_values.append(task_lr_link)
                       task_values.append("Diksha")
                       task_values.append(task_lr_link_id)
                       task_lr_value_count += 1
               task_values.append(taskminNoOfSubmissionsRequired)
               task_values.append(sequenceNumber)

               # To check weather the previous-task and the curent-task Taskname & Taskid is same 
               if str(taskName) == str(PreviousTaskname) and str(taskId) == str(PreviousTaskid):
                    print("true")
               else:
                   print("false")
                   with open(taskFilePath + 'taskUpload.csv','a',encoding='utf-8') as file:
                    writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                    writer.writerows([task_values])
               subtaskname2 = str(dictTasksDetails["Subtask"]).encode('utf-8').decode('utf-8').strip()
               PreviousTaskname = taskName
               PreviousTaskid = taskId

        c = 0
        for row_index_env in range(2, tasksDetailsSheet.nrows):
            dictTasksDetails = {keysTasks[col_index_env]: tasksDetailsSheet.cell(row_index_env, col_index_env).value
                                for col_index_env in range(tasksDetailsSheet.ncols)}
            if dictTasksDetails['TaskId'] and dictTasksDetails["Subtask"]:
                if dictTasksDetails["Subtask"] != "":
                    taskHasAParentTask = "YES"
                    taskparentTaskOperator = "EQUALS"
                    taskparentTaskValue = "started"
                    c = c + 1
                    cn = "Task"+str(c)
                    parentTaskIdofsubtask = str(dictTasksDetails["TaskId"]).strip() + "-" + str(millisecond)
                    taskminNoOfSubmissionsRequired = str(dictTasksDetails["Number of submissions for observation"]).strip()
                    sequenceNumber = sequenceNumber + 1
                    try:
                        proejcttaskDescription = str(dictTasksDetails["description"]).strip()
                    except:
                        proejcttaskDescription = ""
                    if dictTasksDetails["observation Name"] != "":
                        projecttaskType = "observation"
                    elif dictTasksDetails["learningResources1-name"] != "" and dictTasksDetails[
                        "learningResources1-link"] != "":
                        projecttaskType = "content"
                    else:
                        projecttaskType = "simple"


                subtaskId = str(dictTasksDetails["TaskId"]).encode('utf-8').decode('utf-8').strip() + "-" + str(millisecond) + cn

                subtaskName1 = str(dictTasksDetails["Subtask"]).strip()
                if str(dictTasksDetails["Mandatory task(Yes or No)"]).strip().strip().lower() == "no":
                    isDeletable = "TRUE"
                else:
                    isDeletable = "FALSE"
                subtaskvalues = [subtaskName1, subtaskId,proejcttaskDescription,projecttaskType,taskHasAParentTask,taskparentTaskOperator,taskparentTaskValue,
                                 parentTaskIdofsubtask, taskSolutionType, solutionSubType, solutionId, isDeletable]
                task_lr_value_count = 1
                for task_lr in range(0, int(taskLearningResource_count)):
                    task_lr_name = str(dictTasksDetails["learningResources" + str(task_lr_value_count) + "-name"]).strip()
                    task_lr_link = str(dictTasksDetails["learningResources" + str(task_lr_value_count) + "-link"]).strip()
                    if task_lr_name == "" and task_lr_link == "":
                        task_values.append("")
                        task_values.append("")
                        task_values.append("")
                        task_values.append("")
                        task_lr_value_count += 1
                    else:
                        task_values.append(task_lr_name)
                        task_lr_link_id = task_lr_link.split("/")[-1]
                        task_values.append(task_lr_link)
                        task_values.append("Diksha")
                        task_values.append(task_lr_link_id)
                        task_lr_value_count += 1
                task_values.append(taskminNoOfSubmissionsRequired)
                task_values.append(sequenceNumber)

                with open(taskFilePath + 'taskUpload.csv', 'a',encoding='utf-8') as file:
                    writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                    writer.writerows([subtaskvalues])


            if dictTasksDetails["Subtask"] and not dictTasksDetails['TaskTitle']:
                if dictTasksDetails["Subtask"] != "":
                    taskHasAParentTask = "YES"
                    taskparentTaskOperator = "EQUALS"
                    taskparentTaskValue = "started"
                    # c = c + 1
                    # cn = "Task"+str(c)
                    parentTaskId = str(dictTasksDetails["TaskId"]).encode('utf-8').decode('utf-8').strip() + "-" + str(millisecond)
                    try:
                        proejcttaskDescription = str(dictTasksDetails["description"]).strip()
                    except:
                        proejcttaskDescription = ""
                    if dictTasksDetails["observation Name"] != "":
                        projecttaskType = "observation"
                    elif dictTasksDetails["learningResources1-name"] != "" and dictTasksDetails[
                        "learningResources1-link"] != "":
                        projecttaskType = "content"
                    else:
                        projecttaskType = "simple"

                subtaskId = str(dictTasksDetails["TaskId"]).encode('utf-8').decode('utf-8').strip() + "-" + str(millisecond) + cn

                subtaskName1 = str(dictTasksDetails["Subtask"]).encode('utf-8').decode('utf-8').strip()
                subtaskvalues = [subtaskName1, subtaskId,proejcttaskDescription,projecttaskType,taskHasAParentTask,taskparentTaskOperator,taskparentTaskValue,
                                 parentTaskId, taskSolutionType, solutionSubType, solutionId, isDeletable]

                with open(taskFilePath + 'taskUpload.csv','a',encoding='utf-8') as file:
                    writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                    writer.writerows([subtaskvalues])


    def projectUpload(projectFile, projectName_for_folder_path, accessToken):
        urlProjectUploadApi = internal_kong_ip + projectuploadapi
        headerProjectUploadApi = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id':  x_channel_id,
            'internal-access-token': internal_access_token
        }
        project_payload = {}
        filesProject = {
            'projectTemplates': open(projectName_for_folder_path + '/projectUpload/projectUpload.csv', 'rb')
        }

        responseProjectUploadApi = requests.post(url=urlProjectUploadApi, headers=headerProjectUploadApi,data=project_payload,files=filesProject)
        messageArr = ["program mapping is success.","File path : " + projectName_for_folder_path + '/projectUpload/projectUpload.csv']
        messageArr.append("Upload status code : " + str(responseProjectUploadApi.status_code))
        Helpers.createAPILog(projectName_for_folder_path, messageArr)

        if responseProjectUploadApi.status_code == 200:
            print('ProjectUploadApi Success')
            with open(projectName_for_folder_path + '/projectUpload/projectInternal.csv','w+',encoding='utf-8') as projectRes:
                projectRes.write(responseProjectUploadApi.text)
        else:
            print("Project Upload failed.")
            messageArr.append("Response : " + str(responseProjectUploadApi.text))
            Helpers.createAPILog(projectName_for_folder_path, messageArr)
            sys.exit()

    def taskUpload(projectFile, projectName_for_folder_path, accessToken):
        projectInternalfile = open(projectName_for_folder_path + '/projectUpload/projectInternal.csv', mode='r',encoding='utf-8')
        projectInternalfile = csv.DictReader(projectInternalfile)
        for projectInternal in projectInternalfile:
            projectExternalId = projectInternal["externalId"]
            project_id = projectInternal["_SYSTEM_ID"]
            if str(project_id).strip() == "Could not pushed to kafka":
                fetchProjectIdApi = internal_kong_ip + fetchprojectlist
                headerfetchProjectIdApi = {
                    'Authorization': authorization,
                    'X-authenticated-user-token': accessToken,
                    'X-Channel-id': x_channel_id,
                    'internal-access-token': internal_access_token
                }
                fetchProjectIdPayload = {}

                responseProjectListApi = requests.get(url=fetchProjectIdApi, headers=headerfetchProjectIdApi,
                                                      data=fetchProjectIdPayload)
                messageArr = ["Tasks Upload Sheet Prepared.",
                              "File path : " + projectName_for_folder_path + '/taskUpload/taskUpload.csv']
                messageArr.append("URL : " + str(fetchProjectIdApi))
                messageArr.append("Upload status code : " + str(responseProjectListApi.status_code))
                Helpers.createAPILog(projectName_for_folder_path, messageArr)

                if responseProjectListApi.status_code == 200:
                    print('project fetch api Success')
                    responsejson = responseProjectListApi.json()
                    projectList = responsejson['result']['data']
                    for project in projectList:
                        if project['externalId'] == projectExternalId:
                            project_id = project['_id']
                else:
                    messageArr.append("Response : " + str(responseProjectListApi.text))
                    Helpers.createAPILog(projectName_for_folder_path, messageArr)
                    Helpers.terminatingMessage("project fetch api failed.")

            urlTasksUploadApi = internal_kong_ip + taskuploadapi + project_id
            headerTasksUploadApi = {
                'Authorization': authorization,
                'X-authenticated-user-token': accessToken,
                'X-Channel-id': x_channel_id,
                'internal-access-token': internal_access_token
            }
            task_payload = {}
            filesTasks = {
                'projectTemplateTasks': open(projectName_for_folder_path + '/taskUpload/taskUpload.csv',
                                             'rb')
            }

            responseTasksUploadApi = requests.post(url=urlTasksUploadApi, headers=headerTasksUploadApi,
                                                   data=task_payload,
                                                   files=filesTasks)
            messageArr = ["Tasks Upload Sheet Prepared.",
                          "File path : " + projectName_for_folder_path + '/taskUpload/taskUpload.csv']
            messageArr.append("URL : " + str(urlTasksUploadApi))
            messageArr.append("Upload status code : " + str(responseTasksUploadApi.status_code))
            Helpers.createAPILog(projectName_for_folder_path, messageArr)

            if responseTasksUploadApi.status_code == 200:
                print('TaskUploadApi Success')
                with open(projectName_for_folder_path + '/taskUpload/taskInternal.csv','w+',encoding='utf-8') as tasksRes:
                    tasksRes.write(responseTasksUploadApi.text)
            else:
                messageArr.append("Response : " + str(responseTasksUploadApi.text))
                Helpers.createAPILog(projectName_for_folder_path, messageArr)
                Helpers.terminatingMessage("--->Tasks Upload failed.")

    def prepareaddingcertificatetemp(filePathAddProject, projectName_for_folder_path, accessToken, solutionId, programID,baseTemplate_id):
        wbproject = xlrd.open_workbook(filePathAddProject, on_demand=True)
        projectsheetforcertificate = wbproject.sheet_names()
        tasksLevelEvidance = []
        projectMinNooEvide = None
        projectLevelEvidance = []
        taskMinNooEvide =[]


        for prosheet in projectsheetforcertificate:
            if prosheet.strip().lower() == 'Project upload'.lower():
                detailsColCheck = wbproject.sheet_by_name(prosheet)
                keysColCheckDetai = [detailsColCheck.cell(0, col_index_check).value for col_index_check in
                                     range(detailsColCheck.ncols)]

                detailsEnvSheet = wbproject.sheet_by_name(prosheet)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {
                        keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                        for col_index_env in range(detailsEnvSheet.ncols)}

                    projectLevelMinNooEvidence = dictDetailsEnv["Minimum No. of Evidence"]
                    print(projectLevelMinNooEvidence)
                    projectLevelEvidance = dictDetailsEnv["Project Level Evidence"].lower()
                    if projectLevelMinNooEvidence == "":
                        projectLevelMinNooEvidence = 1  # Set default value to 1
                        projectMinNooEvide = int(projectLevelMinNooEvidence)
                    else:
                        projectMinNooEvide = int(projectLevelMinNooEvidence)


        for prosheet in projectsheetforcertificate:
            if prosheet.strip().lower() == 'Tasks upload'.lower():
                detailsColCheck = wbproject.sheet_by_name(prosheet)
                keysColCheckDetai = [detailsColCheck.cell(0, col_index_check).value for col_index_check in
                                     range(detailsColCheck.ncols)]

                detailsEnvSheet = wbproject.sheet_by_name(prosheet)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]

                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {
                        keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                        for col_index_env in range(detailsEnvSheet.ncols)}


                    taskLevelEvidence = dictDetailsEnv["Task Level Evidence"].lower()
                    minNoOfEvidence = dictDetailsEnv["Minimum No. of Evidence"]

                    if taskLevelEvidence == "yes":
                        tasksLevelEvidance.append(dictDetailsEnv["TaskTitle"])
                        if minNoOfEvidence == "":
                            minNoOfEvidence = 1  # Set default value to 1
                            taskMinNooEvide.append(minNoOfEvidence)
                        else:
                            taskMinNooEvide.append(minNoOfEvidence)



        addcetificateFilePath = projectName_for_folder_path + '/addCertificate/'
        if not os.path.exists(addcetificateFilePath):
            os.mkdir(addcetificateFilePath)

        urladdcertificate = internal_kong_ip + addcertificatetemplate
        headeraddcertificateApi = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            'internal-access-token': internal_access_token,
            'Content-Type': 'application/json'
        }

        if str(projectLevelEvidance).strip().lower() == "yes":
            payload = {}
            payload['criteria'] = {}
            payload['criteria']['validationText'] = "Complete validation message"
            payload['criteria']['expression'] = ""
            payload['criteria']['conditions'] = {}
            payload['criteria']['conditions']['C1'] = {}
            payload['criteria']['conditions']['C1']['validationText'] = "Submit your project."
            payload['criteria']['conditions']['C1']['expression'] = "C1"
            payload['criteria']['conditions']['C1']['conditions'] = {}
            payload['criteria']['conditions']['C1']['conditions']['C1'] = {}
            payload['criteria']['conditions']['C1']['conditions']['C1']['scope'] = "project"
            payload['criteria']['conditions']['C1']['conditions']['C1']['key'] = "status"
            payload['criteria']['conditions']['C1']['conditions']['C1']['operator'] = "=="
            payload['criteria']['conditions']['C1']['conditions']['C1']['value'] = "submitted"
            payload['criteria']['conditions']['C2'] = {}
            payload['criteria']['conditions']['C2']['validationText'] = f"Add {int(projectMinNooEvide)} evidence at the project level",
            payload['criteria']['conditions']['C2']['expression'] = "C1"
            payload['criteria']['conditions']['C2']['conditions'] = {}
            payload['criteria']['conditions']['C2']['conditions']['C1'] = {}
            payload['criteria']['conditions']['C2']['conditions']['C1']['scope'] = "project"
            payload['criteria']['conditions']['C2']['conditions']['C1']['key'] = "attachments"
            payload['criteria']['conditions']['C2']['conditions']['C1']['function'] = "count"
            payload['criteria']['conditions']['C2']['conditions']['C1']['filter'] = {}
            payload['criteria']['conditions']['C2']['conditions']['C1']['filter']['key'] = "type"
            payload['criteria']['conditions']['C2']['conditions']['C1']['filter']['value'] = "all"  
            payload['criteria']['conditions']['C2']['conditions']['C1']['operator'] = ">="
            payload['criteria']['conditions']['C2']['conditions']['C1']['value'] = int(projectMinNooEvide)
            payload['issuer'] ={}
            payload['issuer']['name']=""
            payload['status'] = "active"
            payload['solutionId'] = solutionId
            payload['programId'] = programID
            payload['baseTemplateId'] = ""

        else:
            # str(projectLevelEvidance).strip().lower() == "no":
            payload = {}
            payload['criteria'] = {}
            payload['criteria']['validationText'] = "Complete validation message"
            payload['criteria']['expression'] = ""
            payload['criteria']['conditions'] = {}
            payload['criteria']['conditions']['C1'] = {}
            payload['criteria']['conditions']['C1']['validationText'] = "Submit your project."
            payload['criteria']['conditions']['C1']['expression'] = "C1"
            payload['criteria']['conditions']['C1']['conditions'] = {}
            payload['criteria']['conditions']['C1']['conditions']['C1'] = {}
            payload['criteria']['conditions']['C1']['conditions']['C1']['scope'] = "project"
            payload['criteria']['conditions']['C1']['conditions']['C1']['key'] = "status"
            payload['criteria']['conditions']['C1']['conditions']['C1']['operator'] = "=="
            payload['criteria']['conditions']['C1']['conditions']['C1']['value'] = "submitted"
            payload['issuer'] ={}
            payload['issuer']['name']=""
            payload['status'] = "active"
            payload['solutionId'] = solutionId
            payload['programId'] = programID
            payload['baseTemplateId'] = ""


        if prosheet.strip().lower() == 'Certificate details'.lower():
            print("--->Checking Certificate details  sheet...")
            detailsColCheck = wbproject.sheet_by_name(prosheet)
            keysColCheckDetai = [detailsColCheck.cell(0, col_index_check).value for col_index_check in
                    range(detailsColCheck.ncols)]

            detailsEnvSheet = wbproject.sheet_by_name(prosheet)
            keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
            for row_index_env in range(2, detailsEnvSheet.nrows):

                dictDetailsEnv = {
                        keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                        for
                        col_index_env in range(detailsEnvSheet.ncols)}
                certificateissuer = dictDetailsEnv['Certificate issuer'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Certificate issuer'] else Helpers.terminatingMessage("\"Certificate issuer\" must not be Empty in \"Certificate details\" sheet")
                payload["issuer"]["name"] = certificateissuer

                Typeofcertificate = dictDetailsEnv['Type of certificate'] if dictDetailsEnv['Type of certificate'] in ["One Logo - One Signature", "One Logo - Two Signature", "Two Logo - One Signature","Two Logo - Two Signature"] else Helpers.terminatingMessage("\"Type of certificate\" must not be Empty in \"Certificate details\" sheet")

                payload["baseTemplateId"]=baseTemplate_id

        projectInternalfile = open(projectName_for_folder_path + '/projectUpload/projectInternal.csv', mode='r',encoding='utf-8')
        projectInternalfile = csv.DictReader(projectInternalfile)
        for projectInternal in projectInternalfile:
            projectExternalId = projectInternal["externalId"]
            project_id = projectInternal["_SYSTEM_ID"]

        taskinternalfile = open(projectName_for_folder_path + '/taskUpload/taskInternal.csv', mode='r',encoding='utf-8')
        taskinternalfile = csv.DictReader(taskinternalfile)
        projectTemplatefile = open(projectName_for_folder_path + '/solutionDetails/solutionDetails.csv', mode='r',encoding='utf-8')
        projectTemplatefile = csv.DictReader(projectTemplatefile)
        for Projecttemp in projectTemplatefile:
            projectTemplateId = Projecttemp["duplicateTemplate_id"]
        c = 2
        for task in taskinternalfile:
            if task['name'] in tasksLevelEvidance:
                hasAparent = task["hasAParentTask"]
                if task["hasAParentTask"].lower() == "no":

                    task_id = task["_SYSTEM_ID"]

                    c = c + 1
                    cn = "C" + str(c)
                    taskconditions = {
                        cn: {
                            "validationText": f"Add {int(taskMinNooEvide[c-3])} evidence for the task {tasksLevelEvidance[c-3]}",
                            "expression": "C1",
                            "conditions": {
                                "C1": {
                                    "scope": "task",
                                    "key": "attachments",
                                    "function": "count",
                                    "filter": {
                                        "key": "type",
                                        "value": "all"
                                    },
                                    "operator": ">=",
                                    "value": int(taskMinNooEvide[c-3]),
                                    "taskDetails": [
                                        task_id
                                    ]
                                }
                            }
                        }
                    }
                    payload["criteria"]["conditions"].update(taskconditions)
            else:
                pass


        condition = ""
        for a, i in enumerate(payload["criteria"]["conditions"]):
            if a == 0:
                condition = condition + str(i)
            else:
                condition = condition + "&&" + str(i)
        payload["criteria"]["expression"] = condition


        print(json.dumps(payload, indent=1))
        # sys.exit()

        responseaddcertificateUploadApi = requests.request("POST",url=urladdcertificate, headers=headeraddcertificateApi,
                                               data=json.dumps(payload))
        messageArr = ["Add certificate json is prepared",
                      "File path : " + projectName_for_folder_path + '/addCertificate/Addcertificate.text']
        messageArr.append("URL : " + str(responseaddcertificateUploadApi))
        messageArr.append("Upload status code : " + str(responseaddcertificateUploadApi.status_code))
        Helpers.createAPILog(projectName_for_folder_path, messageArr)
        with open(projectName_for_folder_path + '/addCertificate/Addcertificatejson.json',
                  'w+',encoding='utf-8') as tasksRes:
            tasksRes.write(json.dumps(payload))

        if responseaddcertificateUploadApi.status_code == 200:
            responseaddcetificate = responseaddcertificateUploadApi.json()
            certificatetemplateid = responseaddcetificate['result']['id']
            print("-->Certificate template id generated <--", certificatetemplateid)


            with open(projectName_for_folder_path + '/addCertificate/Addcertificate.text',
                      'w+',encoding='utf-8') as tasksRes:
                tasksRes.write(responseaddcertificateUploadApi.text)

        else:
            print("Add certificate mission failed please check logs")
            messageArr.append("Response : " + str(responseaddcertificateUploadApi.text))
            Helpers.createAPILog(projectName_for_folder_path, messageArr)
            sys.exit()

        urluploadcertificatepi =internal_kong_ip + uploadcertificatetosvg + certificatetemplateid

        headeruploadcertificateApi = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            'internal-access-token': internal_access_token
        }
        task_payload = {}
        task_file = []
        certificateaddtotemplate = ('file', ( 'Dowloaded.svg',open(projectName_for_folder_path + '/Dowloadedsvg/Dowloaded.svg', 'rb'), 'image/svg+xml'))
        task_file.append(certificateaddtotemplate)


        responseDownloadsvgApi = requests.request("POST",url=urluploadcertificatepi, headers=headeruploadcertificateApi,
                                               data=task_payload,
                                               files=task_file)
        if responseDownloadsvgApi.status_code == 200:
            responseeditsvg = responseDownloadsvgApi.json()
            svgid = responseeditsvg['result']['data']['templateId']

            urlsolutionupdateapi = internal_kong_ip + updatecertificatesolu + solutionId

            headersolutionupdateApi = {
                'Authorization': authorization,
                'X-authenticated-user-token': accessToken,
                'X-Channel-id': x_channel_id,
                'internal-access-token': internal_access_token,
                'Content-Type': 'application/json'
            }

            certificate_payload = json.dumps({
                'certificateTemplateId':certificatetemplateid
            })
            responseupdatecertificateApi = requests.request("POST", url=urlsolutionupdateapi,
                                                      headers=headersolutionupdateApi,
                                                      data=certificate_payload)


            if responseupdatecertificateApi.status_code == 200:
                print("--->certificate added to the solution<---")

            else:
                print("error in updating solution")
                sys.exit()

            urlprojecttemplateapi = internal_kong_ip + updateprojecttemplate + projectTemplateId
            headerprojectrtemplateupdateApi = {
                'Authorization': authorization,
                'X-authenticated-user-token': accessToken,
                'X-Channel-id': x_channel_id,
                'internal-access-token': internal_access_token,
                'Content-Type': 'application/json'
            }

            certificate_payload = json.dumps({
                'certificateTemplateId': certificatetemplateid
            })
            responseupdatecertificateApi = requests.request("POST", url=urlprojecttemplateapi,
                                                            headers=headerprojectrtemplateupdateApi,
                                                            data=certificate_payload)
            if responseupdatecertificateApi.status_code == 200:
                print("--->Certificate added to project<---")

            else:
                print("error in updating certificate with project")
                sys.exit()
# Th    is function is used to add SVG to the certificate based on type of certificate


    def editsvg(accessToken,filePathAddProject,projectName_for_folder_path,baseTemplate_id):
        wbproject = xlrd.open_workbook(filePathAddProject, on_demand=True)
        projectsheetforcertificate = wbproject.sheet_names()
        for prosheet in projectsheetforcertificate:
            if prosheet.strip().lower() == 'Certificate details'.lower():
                print("--->Checking Certificate details  sheet...")
                detailsColCheck = wbproject.sheet_by_name(prosheet)
                keysColCheckDetai = [detailsColCheck.cell(0, col_index_check).value for col_index_check in
                                     range(detailsColCheck.ncols)]

                detailsEnvSheet = wbproject.sheet_by_name(prosheet)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):

                    dictDetailsEnv = {
                        keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                        for
                        col_index_env in range(detailsEnvSheet.ncols)}
                    certificateissuer = dictDetailsEnv['Certificate issuer'].encode('utf-8').decode('utf-8')
                    Typeofcertificate = dictDetailsEnv['Type of certificate']
                    Certificateisuuer = dictDetailsEnv['Certificate issuer'].encode('utf-8').decode('utf-8')
                    Logo1 = dictDetailsEnv['Logo - 1']
                    authsignaturelogo1 = dictDetailsEnv['Authorised Signature Image - 1']
                    authrigedsignaturename1 = dictDetailsEnv['Authorised Signature Name - 1'].encode('utf-8').decode('utf-8')
                    authrigeddesignation1 = dictDetailsEnv['Authorised Designation - 1'].encode('utf-8').decode('utf-8')
                    authrigedlogo2 = dictDetailsEnv['Authorised Signature Image - 2']
                    authrigedsignaturename2 = dictDetailsEnv['Authorised Signature Name - 2'].encode('utf-8').decode('utf-8')
                    authrigeddesignation2 = dictDetailsEnv['Authorised Designation - 2'].encode('utf-8').decode('utf-8')

                    payload = {}
                    downloadedfiles = []
                    baseTemplateId = ''
                    if Typeofcertificate == 'One Logo - One Signature':
                        print("-->This is One Logo - One Signature<--")

                        stateLogo1 = ('stateLogo1',('logo1.jpg',open(projectName_for_folder_path +'/Logofile/logo1.jpg' ,'rb'),'image/jpeg'))
                        downloadedfiles.append(stateLogo1)
                        payload['stateTitle'] = Certificateisuuer
                        signatureImg1 = ('signatureImg1',('signature1.jpg',open(projectName_for_folder_path +'/Logofile/signature1.jpg','rb'),'image/jpeg'))
                        downloadedfiles.append(signatureImg1)
                        payload['signatureTitleName1'] = authrigedsignaturename1
                        payload['signatureTitleDesignation1'] = authrigeddesignation1
                        baseTemplateId=baseTemplate_id


                    elif Typeofcertificate == 'One Logo - Two Signature':
                        print("-->This is One Logo - Two Signature<--")

                        stateLogo1 = ('stateLogo1', (
                        'logo1.jpg', open(projectName_for_folder_path + '/Logofile/logo1.jpg', 'rb'), 'image/jpeg'))
                        downloadedfiles.append(stateLogo1)
                        payload['stateTitle'] = Certificateisuuer
                        signatureImg1 = ('signatureImg1', (
                        'signature1.jpg', open(projectName_for_folder_path + '/Logofile/signature1.jpg', 'rb'),
                        'image/jpeg'))
                        downloadedfiles.append(signatureImg1)
                        signatureImg2 = ('signatureImg2', ('signature2.jpg', open(projectName_for_folder_path + '/Logofile/signature2.jpg', 'rb'),'image/jpeg'))
                        downloadedfiles.append(signatureImg2)
                        payload['signatureTitleName1'] = authrigedsignaturename1
                        payload['signatureTitleDesignation1'] = authrigeddesignation1
                        payload['signatureTitleName2'] = authrigedsignaturename2
                        payload['signatureTitleDesignation2'] = authrigeddesignation2
                        baseTemplateId=baseTemplate_id

                    elif Typeofcertificate == 'Two Logo - One Signature':
                        print("-->This is Two Logo - One Signature<--")
                        stateLogo1 = ('stateLogo1', (
                            'logo1.jpg', open(projectName_for_folder_path + '/Logofile/logo1.jpg', 'rb'), 'image/jpeg'))
                        downloadedfiles.append(stateLogo1)
                        payload['stateTitle'] = Certificateisuuer
                        signatureImg1 = ('signatureImg1', ('signature1.jpg', open(projectName_for_folder_path + '/Logofile/signature1.jpg', 'rb'),'image/jpeg'))
                        downloadedfiles.append(signatureImg1)
                        stateLogo2 = ('stateLogo2', ('logo2.jpg', open(projectName_for_folder_path + '/Logofile/logo2.jpg', 'rb'), 'image/jpeg'))
                        downloadedfiles.append(stateLogo2)
                        payload['signatureTitleName1'] = authrigedsignaturename1
                        payload['signatureTitleDesignation1'] = authrigeddesignation1
                        baseTemplateId=baseTemplate_id

                    elif Typeofcertificate == 'Two Logo - Two Signature':
                        print("-->This is Two Logo - Two Signature<--")
                        stateLogo1 = ('stateLogo1', ('logo1.jpg', open(projectName_for_folder_path + '/Logofile/logo1.jpg', 'rb'), 'image/jpeg'))
                        downloadedfiles.append(stateLogo1)
                        payload['stateTitle'] = Certificateisuuer
                        signatureImg1 = ('signatureImg1', ('signature1.jpg', open(projectName_for_folder_path + '/Logofile/signature1.jpg', 'rb'),'image/jpeg'))
                        downloadedfiles.append(signatureImg1)
                        stateLogo2 = ('stateLogo2', ('logo2.jpg', open(projectName_for_folder_path + '/Logofile/logo2.jpg', 'rb'), 'image/jpeg'))
                        downloadedfiles.append(stateLogo2)
                        signatureImg2 = ('signatureImg2', ('signature2.jpg', open(projectName_for_folder_path + '/Logofile/signature2.jpg', 'rb'),'image/jpeg'))
                        downloadedfiles.append(signatureImg2)
                        payload['signatureTitleName1'] = authrigedsignaturename1
                        payload['signatureTitleDesignation1'] = authrigeddesignation1
                        payload['signatureTitleName2'] = authrigedsignaturename2
                        payload['signatureTitleDesignation2'] = authrigeddesignation2
                        baseTemplateId=baseTemplate_id

                    urleditnigsvgApi =  internal_kong_ip + editsvgtemp + baseTemplateId
                    headereditingsvgApi = {
                        'Authorization': authorization,
                        'X-authenticated-user-token': accessToken,
                        'X-Channel-id': x_channel_id,
                        'internal-access-token': internal_access_token

                    }
                    responseeditsvg = requests.request("POST",url=urleditnigsvgApi, headers=headereditingsvgApi,data=payload, files=downloadedfiles)

                    if responseeditsvg.status_code == 200:
                        responseeditsvg = responseeditsvg.json()
                        svgid = responseeditsvg['result']['url']
                        filesvg = svgid
                        Logofilepath = projectName_for_folder_path + '/Dowloadedsvg/'
                        if not os.path.exists(Logofilepath):
                            os.mkdir(Logofilepath)
                        dest_file = Logofilepath + 'Dowloaded.svg'
                        Logofile1 = gdown.download(filesvg, dest_file, quiet=False)

                    else:
                        print("-->Error in downloading SVG file please check logs<--")

    


    def fetchCertificateBaseTemplate(filePathAddProject,accessToken,projectName_for_folder_path):
        wbproject = xlrd.open_workbook(filePathAddProject, on_demand=True)
        projectsheetforcertificate = wbproject.sheet_names()
        for prosheet in projectsheetforcertificate:
            if prosheet.strip().lower() == 'Certificate details'.lower():
                detailsColCheck = wbproject.sheet_by_name(prosheet)
                keysColCheckDetai = [detailsColCheck.cell(0, col_index_check).value for col_index_check in
                                     range(detailsColCheck.ncols)]

                detailsEnvSheet = wbproject.sheet_by_name(prosheet)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {
                        keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                        for col_index_env in range(detailsEnvSheet.ncols)}

                    typeOfCertificate = dictDetailsEnv["Type of certificate"]
                    print(typeOfCertificate)

        urldbFind = internal_kong_ip + dbfindapi
        headerdbFindApi = {
            'Authorization':  authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            'internal-access-token': internal_access_token,
            'Content-Type': 'application/json'
        }
        payload = json.dumps({
            "query": {},
            "mongoIdKeys": []
        })

        responsedbFindApi = requests.request("POST", url=urldbFind, headers=headerdbFindApi,
                                             data=payload)
        if responsedbFindApi.status_code == 200:
            responseaddcetificate = responsedbFindApi.json()
            result_list = responseaddcetificate['result']
            baseTemplateLookup = {}
            for i in result_list:
                baseTemplateLookup[i['code']] = i['_id']
            typeOfCertificate=typeOfCertificate.lower()
            typeOfCertificate=typeOfCertificate.replace("-","_")
            typeOfCertificate = typeOfCertificate.replace(" ","")
            baseTemplateCode= certificatetypeof[typeOfCertificate]
            print(baseTemplateCode,"baseTemplateCode")
            print(baseTemplateLookup,"baseTemplateLookup")

            return baseTemplateLookup[baseTemplateCode]

        else:
            print("--->Error in fetching DBfind data please give proper code value<---")
            #messageArr.append("Response : " + str(responseaddcetificate.text))
            #createAPILog(projectName_for_folder_path, messageArr)
            sys.exit()


    def downloadlogosign(filePathAddProject,projectName_for_folder_path):
        wbproject = xlrd.open_workbook(filePathAddProject, on_demand=True)
        projectsheetforcertificate = wbproject.sheet_names()
        for prosheet in projectsheetforcertificate:
            if prosheet.strip().lower() == 'Certificate details'.lower():
                print("--->Checking Certificate details  sheet...")
                detailsColCheck = wbproject.sheet_by_name(prosheet)
                keysColCheckDetai = [detailsColCheck.cell(0, col_index_check).value for col_index_check in range(detailsColCheck.ncols)]

                detailsEnvSheet = wbproject.sheet_by_name(prosheet)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):

                    dictDetailsEnv = {
                        keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                        for
                        col_index_env in range(detailsEnvSheet.ncols)}
                    certificateissuer = dictDetailsEnv['Certificate issuer'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Certificate issuer'] else Helpers.terminatingMessage("\"Certificate issuer\" must not be Empty in \"Certificate details\" sheet")

                    typeOfCertificate = dictDetailsEnv['Type of certificate'] if dictDetailsEnv['Type of certificate'] else Helpers.terminatingMessage("\"Type of certificate\" must not be Empty in \"Certificate details\" sheet")

                    if typeOfCertificate == 'One Logo - One Signature':
                       Logo1 = dictDetailsEnv['Logo - 1']
                       logo_split = str(Logo1).split('/')[5]

                       file_url = 'https://drive.google.com/uc?export=download&id='+logo_split

                       Logofilepath = projectName_for_folder_path + '/Logofile/'
                       if not os.path.exists(Logofilepath):
                           os.mkdir(Logofilepath)
                       dest_file = Logofilepath + '/logo1.jpg'
                       Logofile1 = gdown.download(file_url, dest_file,quiet=False)


                       Authsign1 = dictDetailsEnv['Authorised Signature Image - 1']
                       logo_split = str(Authsign1).split('/')[5]

                       file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split


                       dest_file = Logofilepath + '/signature1.jpg'
                       signature1 = gdown.download(file_url, dest_file, quiet=False)

                    elif typeOfCertificate == 'One Logo - Two Signature':

                        Logo1 = dictDetailsEnv['Logo - 1']
                        logo_split = str(Logo1).split('/')[5]

                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split

                        Logofilepath = projectName_for_folder_path + '/Logofile/'
                        if not os.path.exists(Logofilepath):
                            os.mkdir(Logofilepath)
                        dest_file = Logofilepath + '/logo1.jpg'
                        Logofile1 = gdown.download(file_url, dest_file, quiet=False)


                        Authsign1 = dictDetailsEnv['Authorised Signature Image - 1']
                        logo_split = str(Authsign1).split('/')[5]

                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split

                        dest_file = Logofilepath + '/signature1.jpg'
                        signature1 = gdown.download(file_url, dest_file, quiet=False)

                        Authsign2 = dictDetailsEnv['Authorised Signature Image - 2']
                        logo_split = str(Authsign2).split('/')[5]

                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split


                        dest_file = Logofilepath + '/signature2.jpg'
                        signature2 = gdown.download(file_url, dest_file, quiet=False)

                    elif typeOfCertificate == 'Two Logo - One Signature':

                        Logo1 = dictDetailsEnv['Logo - 1']
                        logo_split = str(Logo1).split('/')[5]

                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split

                        Logofilepath = projectName_for_folder_path + '/Logofile/'
                        if not os.path.exists(Logofilepath):
                            os.mkdir(Logofilepath)
                        dest_file = Logofilepath + '/logo1.jpg'
                        Logofile1 = gdown.download(file_url, dest_file, quiet=False)


                        Logo2 = dictDetailsEnv['Logo - 2']
                        logo_split = str(Logo2).split('/')[5]

                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split


                        dest_file = Logofilepath + '/logo2.jpg'
                        Logofile2 = gdown.download(file_url, dest_file, quiet=False)


                        Authsign1 = dictDetailsEnv['Authorised Signature Image - 1']
                        logo_split = str(Authsign1).split('/')[5]


                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split


                        dest_file = Logofilepath + '/signature1.jpg'
                        signature1 = gdown.download(file_url, dest_file, quiet=False)


                    elif typeOfCertificate == 'Two Logo - Two Signature':

                        Logo1 = dictDetailsEnv['Logo - 1']
                        logo_split = str(Logo1).split('/')[5]


                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split

                        Logofilepath = projectName_for_folder_path + '/Logofile/'
                        if not os.path.exists(Logofilepath):
                            os.mkdir(Logofilepath)
                        dest_file = Logofilepath + '/logo1.jpg'
                        Logofile1 = gdown.download(file_url, dest_file, quiet=False)


                        Logo2 = dictDetailsEnv['Logo - 2']
                        logo_split = str(Logo2).split('/')[5]


                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split


                        dest_file = Logofilepath + '/logo2.jpg'
                        Logofile2 = gdown.download(file_url, dest_file, quiet=False)


                        Authsign1 = dictDetailsEnv['Authorised Signature Image - 1']
                        logo_split = str(Authsign1).split('/')[5]


                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split


                        dest_file = Logofilepath + '/signature1.jpg'
                        signature1 = gdown.download(file_url, dest_file, quiet=False)


                        Authsign2 = dictDetailsEnv['Authorised Signature Image - 2']
                        logo_split = str(Authsign2).split('/')[5]


                        file_url = 'https://drive.google.com/uc?export=download&id=' + logo_split


                        dest_file = Logofilepath + '/signature2.jpg'
                        signature2 = gdown.download(file_url, dest_file, quiet=False)

                    else:
                        print("--->Logos and signature downlading are failed(check if drive link are  Anyone with the link or not)<---")




    def fetchSolutionDetailsFromProgramSheet(solutionName_for_folder_path, programFile, solutionId, accessToken):
        # print("entered fetchSolutionDetailsFromProgramSheet")
        # print("solutionId",solutionId)
        # print(programFile,"programFile")
        global solutionRolesArray, solutionStartDate, solutionEndDate
        urlFetchSolutionApi = internal_kong_ip + fetchsolutiondoc + solutionId

        headerFetchSolutionApi = {
            'Content-Type': 'application/json',
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            'internal-access-token': internal_access_token
        }
        payloadFetchSolutionApi = {}

        responseFetchSolutionApiUrl = requests.post(url=urlFetchSolutionApi, headers=headerFetchSolutionApi,
                                                 data=payloadFetchSolutionApi)
        responseFetchSolutionJson = responseFetchSolutionApiUrl.json()
        messageArr = ["Solution Fetch Link.",
                      "solution name : " + responseFetchSolutionJson["result"]["name"],
                      "solution ExternalId : " + responseFetchSolutionJson["result"]["externalId"]]
        messageArr.append("Upload status code : " + str(responseFetchSolutionApiUrl.status_code))
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)

        if responseFetchSolutionApiUrl.status_code == 200:
            print('Fetch solution Api Success')

            solutionName = responseFetchSolutionJson["result"]["name"]

            xfile = openpyxl.load_workbook(programFile)

            resourceDetailsSheet = xfile['Resource Details']
            # print(resourceDetailsSheet,"1911")
            rowCountRD = resourceDetailsSheet.max_row
            # print(rowCountRD,"rowCountRD")
            columnCountRD = resourceDetailsSheet.max_column
            for row in range(3, rowCountRD + 1):
                # print("here we reacher")
                # print(resourceDetailsSheet,"resourceDetailsSheet")
                solutionNameCell = resourceDetailsSheet[f"A{row}"].value
                # print(f"Row {row} Solution Name 1919: {solutionNameCell}")
                # print(solutionName,"solutionName")
                # print(resourceDetailsSheet["A" + str(row)].value,"1919")
                if resourceDetailsSheet["A" + str(row)].value == solutionName:
                    solutionMainRole = str(resourceDetailsSheet["E" + str(row)].value).strip()
                    solutionRolesArray = str(resourceDetailsSheet["F" + str(row)].value).split(",") if str(
                        resourceDetailsSheet["E" + str(row)].value).split(",") else []
                    if "teacher" in solutionMainRole.strip().lower():
                        solutionRolesArray.append("TEACHER")
                    solutionStartDate = resourceDetailsSheet["G" + str(row)].value
                    # print(solutionStartDate, "<-------------------solutionStartDate////====")
                    solutionEndDate = resourceDetailsSheet["H" + str(row)].value
                    # print(solutionEndDate, "<---------------------solutionEndDate/////========")
        return [solutionRolesArray, solutionStartDate, solutionEndDate]





    def solutionCreationAndMapping(projectName_for_folder_path, entityToUpload, listOfFoundRoles, accessToken,programFile):
        SolutionFilePath = projectName_for_folder_path + '/solutionDetails/'
        if not os.path.exists(SolutionFilePath):
            os.mkdir(SolutionFilePath)
        with open(projectName_for_folder_path + '/solutionDetails/solutionDetails.csv', 'w',encoding='utf-8') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
            writer.writerows(
                [["solutionExtId", "solutionName", "solutionDescription", "solution_id", "programExternalId", "entityType",
                  "scopeEntityType", "entityNames", "roles", "duplicateTemplateExtId", "duplicateTemplate_id"]])

        projectInternalfile = open(projectName_for_folder_path + '/projectUpload/projectInternal.csv', mode='r',encoding='utf-8')
        projectInternalfile = csv.DictReader(projectInternalfile)
        for projectInternal in projectInternalfile:
            projectExternalId = projectInternal["externalId"]
            project_id = projectInternal["_SYSTEM_ID"]
            project_name = projectInternal["title"]
            project_description = projectInternal["description"]
            if projectInternal["entityType"]:
                projectEntityType = projectInternal["entityType"]
            else:
                projectEntityType = "school"
            solutionExternalId = projectExternalId + "-PROJECT-SOLUTION"

            urlCreateProjectSolutionApi = internal_kong_ip + projectsolutioncreationapi
            headerCreateSolutionApi = {
                'Content-Type': content_type,
                'Authorization': authorization,
                'X-authenticated-user-token': accessToken,
                'X-Channel-id': x_channel_id
            }
            sol_payload = {
                "createdFor": orgIds,
                "rootOrganisations": orgIds,
                "programExternalId": programExternalId,
                "entityType": projectEntityType,
                "externalId": solutionExternalId,
                "name": project_name,
                "description": project_description
            }
            responseCreateSolutionApi = requests.post(url=urlCreateProjectSolutionApi,headers=headerCreateSolutionApi, data=json.dumps(sol_payload))

            messageArr = ["Project Solution Created.","URL : " + str(urlCreateProjectSolutionApi),"Status Code : " + str(responseCreateSolutionApi.status_code),"Response : " + str(responseCreateSolutionApi.text)]
            if responseCreateSolutionApi.status_code == 200:
                responseCreateSolutionApi = responseCreateSolutionApi.json()
                solutionId = responseCreateSolutionApi['result']['_id']
                messageArr.append("Solution Generated : " + str(solutionId))
                Helpers.createAPILog(projectName_for_folder_path, messageArr)
                print("ProjectSolutionCreationApi Success")
                duplicateTemplateExtId = projectExternalId + '_IMPORTED'
                queryparamsMapProjectSolutionApi = projectExternalId + '?solutionId=' + solutionExternalId
                urlMapProjectSolutionApi = internal_kong_ip + mapsolutiontoproject
                headerMapSolutionProject = {
                    'Content-Type': content_type,
                    'Authorization': authorization,
                    'X-authenticated-user-token': accessToken,
                    'X-Channel-id': x_channel_id
                }
                payloadMapSolutionProject = {
                    "externalId": duplicateTemplateExtId,
                    "rating": 5
                }
                responseMapProjectSolutionApi = requests.post(
                    url=urlMapProjectSolutionApi + queryparamsMapProjectSolutionApi,
                    headers=headerMapSolutionProject, data=json.dumps(payloadMapSolutionProject))

                messageArr = ["Successfully mapped the project to Solution",
                              "URL : " + str(urlMapProjectSolutionApi + queryparamsMapProjectSolutionApi),
                              "Status Code : " + str(responseMapProjectSolutionApi.status_code),
                              "Response : " + str(responseMapProjectSolutionApi.text)]
                if responseMapProjectSolutionApi.status_code == 200:
                    responseMapProjectSolutionApi = responseMapProjectSolutionApi.json()
                    duplicateTemplateId = responseMapProjectSolutionApi['result']['_id']
                    messageArr.append("duplicate TemplateId successfully created: " + str(duplicateTemplateId))
                    Helpers.createAPILog(projectName_for_folder_path, messageArr)
                    print("MapSolutionToProjectApi Sucsess")
                    with open(projectName_for_folder_path + '/solutionDetails/solutionDetails.csv', 'a',encoding='utf-8') as file:
                        writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                        writer.writerows([[solutionExternalId, project_name, project_description, solutionId,
                                           programExternalId, projectEntityType,
                                           scopeEntityType, entityToUpload, listOfFoundRoles, duplicateTemplateExtId,
                                           duplicateTemplateId]])
                    solutionDetails = Helpers.fetchSolutionDetailsFromProgramSheet(projectName_for_folder_path, programFile,
                                                                           solutionId, accessToken)
                    scopeEntities = entitiesPGMID
                    scopeRoles = solutionDetails[0]
                    bodySolutionUpdate = {
                        "scope": {"entityType": scopeEntityType, "entities": scopeEntities, "roles": scopeRoles}}
                    Helpers.solutionUpdate(projectName_for_folder_path, accessToken, solutionId, bodySolutionUpdate)

                    userDetails = Helpers.fetchUserDetails(accessToken, projectAuthor)
                    matchedShikshalokamLoginId = userDetails[0]
                    projectCreator = userDetails[2]

                    bodySolutionUpdate = {
                        "creator": projectCreator, "author": matchedShikshalokamLoginId}
                    Helpers.solutionUpdate(projectName_for_folder_path, accessToken, solutionId, bodySolutionUpdate)
                    # Below script will convert date DD-MM-YYYY TO YYYY-MM-DD 00:00:00 to match the code syntax

                    if solutionDetails[1]:
                        startDateArr = str(solutionDetails[1]).split("-")
                        bodySolutionUpdate = {
                            "startDate": startDateArr[2] + "-" + startDateArr[1] + "-" + startDateArr[0] + " 00:00:00"}
                        Helpers.solutionUpdate(projectName_for_folder_path, accessToken, solutionId, bodySolutionUpdate)
                    if solutionDetails[2]:
                        endDateArr = str(solutionDetails[2]).split("-")
                        bodySolutionUpdate = {
                            "endDate": endDateArr[2] + "-" + endDateArr[1] + "-" + endDateArr[0] + " 23:59:59"}
                        Helpers.solutionUpdate(projectName_for_folder_path, accessToken, solutionId, bodySolutionUpdate)
                else:
                    print("Map project to solution api failed.")
                return [solutionExternalId, solutionId]
            else:
                print("Project solution creation api failed.")
                sys.exit()



    def prepareProgramSuccessSheet(MainFilePath, solutionName_for_folder_path, programFile, solutionExternalId, solutionId,accessToken):
        urlFetchSolutionApi = internal_kong_ip + fetchsolutiondoc + solutionId
        headerFetchSolutionApi = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            'internal-access-token': internal_access_token
        }
        payloadFetchSolutionApi = {}

        responseFetchSolutionApi = requests.post(url=urlFetchSolutionApi, headers=headerFetchSolutionApi,
                                                 data=payloadFetchSolutionApi)
        responseFetchSolutionJson = responseFetchSolutionApi.json()
        messageArr = ["Solution Fetch Link.",
                      "solution name : " + responseFetchSolutionJson["result"]["name"],
                      "solution ExternalId : " + responseFetchSolutionJson["result"]["externalId"]]
        messageArr.append("Upload status code : " + str(responseFetchSolutionApi.status_code))
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)

        if responseFetchSolutionApi.status_code == 200:
            print('Fetch solution Api Success')
            solutionName = responseFetchSolutionJson["result"]["name"]
        urlFetchSolutionLinkApi = internal_kong_ip + fetchlink + solutionId
        headerFetchSolutionLinkApi = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            'internal-access-token': internal_access_token
        }
        payloadFetchSolutionLinkApi = {}

        responseFetchSolutionLinkApi = requests.post(url=urlFetchSolutionLinkApi, headers=headerFetchSolutionLinkApi,
                                                     data=payloadFetchSolutionLinkApi)
        messageArr = ["Solution Fetch Link.","solution id : " + solutionId,"solution ExternalId : " + solutionExternalId]
        messageArr.append("Upload status code : " + str(responseFetchSolutionLinkApi.status_code))
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)

        if responseFetchSolutionLinkApi.status_code == 200:
            print('Fetch solution Link Api Success')
            responseProjectUploadJson = responseFetchSolutionLinkApi.json()
            solutionLink = responseProjectUploadJson["result"]
        #     messageArr.append("Response : " + str(responseFetchSolutionLinkApi.text))
        #     # createAPILog(solutionName_for_folder_path, messageArr)

        #     if os.path.exists(MainFilePath + "/" + str(programFile).replace(".xlsx", "") + '-SuccessSheet.xlsx'):
        #         xfile = openpyxl.load_workbook(
        #             MainFilePath + "/" + str(programFile).replace(".xlsx", "") + '-SuccessSheet.xlsx')
        #     else:
        #         xfile = openpyxl.load_workbook(programFile)

        #     resourceDetailsSheet = xfile.get_sheet_by_name('Resource Details')

        #     greenFill = PatternFill(start_color='0000FF00',
        #                             end_color='0000FF00',
        #                             fill_type='solid')
        #     rowCountRD = resourceDetailsSheet.max_row
        #     columnCountRD = resourceDetailsSheet.max_column
        #     for row in range(3, rowCountRD + 1):
        #         if str(resourceDetailsSheet["B" + str(row)].value).rstrip().lstrip().lower() == "course":
        #             resourceDetailsSheet["D1"] = ""
        #             resourceDetailsSheet["E1"] = ""
        #             resourceDetailsSheet['I2'] = "External id of the resource"
        #             resourceDetailsSheet['J2'] = "link to access the resource/Response"
        #             resourceDetailsSheet['I2'].fill = greenFill
        #             resourceDetailsSheet['J2'].fill = greenFill
        #             resourceDetailsSheet['I' + str(row)] = solutionExternalId
        #             resourceDetailsSheet['J' + str(row)] = "The course has been successfully mapped to the program"
        #             resourceDetailsSheet['I' + str(row)].fill = greenFill
        #             resourceDetailsSheet['J' + str(row)].fill = greenFill
        #         elif str(resourceDetailsSheet["A" + str(row)].value).strip() == solutionName:
        #             resourceDetailsSheet["D1"] = ""
        #             resourceDetailsSheet["E1"] = ""
        #             resourceDetailsSheet['I2'] = "External id of the resource"
        #             resourceDetailsSheet['J2'] = "link to access the resource/Response"
        #             resourceDetailsSheet['I2'].fill = greenFill
        #             resourceDetailsSheet['J2'].fill = greenFill
        #             resourceDetailsSheet['I' + str(row)] = solutionExternalId
        #             resourceDetailsSheet['J' + str(row)] = solutionLink
        #             resourceDetailsSheet['I' + str(row)].fill = greenFill
        #             resourceDetailsSheet['J' + str(row)].fill = greenFill

        #     programFile = str(programFile).replace(".xlsx", "")
        #     xfile.save(MainFilePath + "/" + programFile + '-SuccessSheet.xlsx')
        #     print("Program success sheet is created")

        else:
            print("Fetch solution link API Failed")
            messageArr.append("Response : " + str(responseFetchSolutionLinkApi.text))
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            sys.exit()
        return solutionLink

# fetch org Ids 
    def fetchOrgId(accessToken, parentFolder, OrgName):
        url = host + fetchorgdetails
        headers = {'Content-Type': 'application/json',
                   'Authorization': authorization,
                   'x-authenticated-user-token': accessToken}
        orgIds = []
        organisations = str(OrgName).split(",")
        for org in organisations:
            orgBody = {"id": "",
                       "ts": "",
                       "params": {
                           "msgid": "",
                           "resmsgid": "",
                           "status": "success"
                       },
                       "request": {
                           "filters": {
                               "orgName": str(org).strip()
                           }
                       }}

            responseOrgSearch = requests.request("POST", url, headers=headers, data=json.dumps(orgBody))
            if responseOrgSearch.status_code == 200:
                responseOrgSearch = responseOrgSearch.json()
                if responseOrgSearch['result']['response']['content']:
                    orgId = responseOrgSearch['result']['response']['content'][0]['id']
                    orgIds.append(orgId)
                else:
                    print("Email is not present in KB")
            else:
               print(responseOrgSearch.text)
                
        return orgIds

    def solutionUpdate(solutionName_for_folder_path, accessToken, solutionId, bodySolutionUpdate):
        solutionUpdateApi = internal_kong_ip + solutionupdateapi + str(solutionId)
        # print("solutionUpdateApi:",solutionUpdateApi)
        headerUpdateSolutionApi = {
            'Content-Type': 'application/json',
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            "internal-access-token": internal_access_token
            }
        responseUpdateSolutionApi = requests.post(url=solutionUpdateApi, headers=headerUpdateSolutionApi,data=json.dumps(bodySolutionUpdate))
        
        if responseUpdateSolutionApi.status_code == 200:
            print("Solution Update Success.")
            return True
        else:
            print("Solution Update Failed.")
            return False

    # def createSurveySolution(parentFolder, wbSurvey, accessToken):
    #     # print(accessToken)
    #     print(wbSurvey,"wbSurveywbSurvey")
    
    #     sheetNames1 = wbSurvey.sheet_names()
    #     for sheetEnv in sheetNames1:
    #         if sheetEnv.strip().lower() == 'details':
    #             surveySolutionCreationReqBody = {}
    #             detailsEnvSheet = wbSurvey.sheet_by_name(sheetEnv)
    #             keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
    #                    range(detailsEnvSheet.ncols)]

    #             for row_index_env in range(2, detailsEnvSheet.nrows):
    #                 dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
    #                               for
    #                               col_index_env in range(detailsEnvSheet.ncols)}
    #                 # print(dictDetailsEnv,"dictDetailsEnv")
    #                 surveySolutionCreationReqBody['name'] = dictDetailsEnv['solution_name'].encode('utf-8').decode('utf-8')
    #                 surveySolutionCreationReqBody["description"] = "survey Solution"
    #                 surveySolutionExternalId = str(uuid.uuid1())
    #                 surveySolutionCreationReqBody["externalId"] = surveySolutionExternalId
    #                 # if dictDetailsEnv['creator_username'].encode('utf-8').decode('utf-8') == "":
    #                 #     exceptionHandlingFlag = True
    #                 #     print('survey_creator_username column should not be empty in the details sheet')
    #                 #     sys.exit()
    #                 # else:
    #                 #     surveySolutionCreationReqBody['creator'] = dictDetailsEnv['Name_of_the_creator'].encode('utf-8').decode('utf-8')

                    
    #                 userDetails = Helpers.fetchUserDetails(accessToken, dictDetailsEnv['creator_username'])
    #                 # print(userDetails)
    #                 surveySolutionCreationReqBody['author'] = userDetails[0]
    #                 # print("surveySolutionCreationReqBody",surveySolutionCreationReqBody)

    #                 # Below script will convert date DD-MM-YYYY TO YYYY-MM-DD 00:00:00 to match the code syntax 

    #                 if dictDetailsEnv["start_date"]:
    #                     if type(dictDetailsEnv["start_date"]) == str:
    #                         startDateArr = None
    #                         startDateArr = (dictDetailsEnv["start_date"]).split("-")
    #                         surveySolutionCreationReqBody["startDate"] = startDateArr[2] + "-" + startDateArr[1] + "-" + \
    #                                                                  startDateArr[0] + " 00:00:00"
    #                     elif type(dictDetailsEnv["start_date"]) == float:
    #                         surveySolutionCreationReqBody["startDate"] = (
    #                         xlrd.xldate.xldate_as_datetime(dictDetailsEnv["start_date"],
    #                                                        wbSurvey.datemode)).strftime("%Y/%m/%d")
    #                     else:
    #                         surveySolutionCreationReqBody["startDate"] = ""
    #                     if dictDetailsEnv["end_date"]:
    #                         if type(dictDetailsEnv["end_date"]) == str:
    #                             print("enter 1")

    #                             endDateArr = None
    #                             endDateArr = (dictDetailsEnv["end_date"]).split("-")
    #                             surveySolutionCreationReqBody["endDate"] = endDateArr[2] + "-" + endDateArr[1] + "-" + \
    #                                                                    endDateArr[0] + " 23:59:59"
    #                         elif type(dictDetailsEnv["end_date"]) == float:
    #                             print("enter 2")
    #                             surveySolutionCreationReqBody["endDate"] = (
    #                                 xlrd.xldate.xldate_as_datetime(dictDetailsEnv["end_date"],
    #                                                            wbSurvey.datemode)).strftime("%Y/%m/%d")
    #                         else:
    #                             print("enter 3")
    #                             surveySolutionCreationReqBody["endDate"] = ""
    #                         enDt = surveySolutionCreationReqBody["endDate"]
                        
    #                         urlCreateSolutionApi =internal_kong_ip_survey+ surveysolutioncreationapiurl
    #                         print(urlCreateSolutionApi)
    #                         headerCreateSolutionApi = {
    #                         'Content-Type': 'application/json',
    #                         'Authorization': authorization,
    #                         'X-authenticated-user-token': accessToken,
    #                         'X-Channel-id': x_channel_id,
    #                         'appName': appname
    #                     }
    #                         # print(surveySolutionCreationReqBody)
    #                         print(headerCreateSolutionApi)
    #                         # sys.exit()
    #                         responseCreateSolutionApi = requests.post(url=urlCreateSolutionApi,
    #                                                               headers=headerCreateSolutionApi,
    #                                                               data=json.dumps(surveySolutionCreationReqBody))
    #                         print(responseCreateSolutionApi.text)
    #                         responseInText = responseCreateSolutionApi.text
                        
    #                         if responseCreateSolutionApi.status_code == 200:
    #                             responseCreateSolutionApi = responseCreateSolutionApi.json()
    #                             urlSearchSolution = internal_kong_ip_core + fetchsolutiondetails + "survey&page=1&limit=10&search=" + str(surveySolutionExternalId)
    #                             print(urlSearchSolution)
    #                             responseSearchSolution = requests.request("POST", urlSearchSolution,
    #                                                                   headers=headerCreateSolutionApi)
                            
    #                             if responseSearchSolution.status_code == 200:
    #                                 responseSearchSolutionApi = responseSearchSolution.json()
    #                                 surveySolutionExternalId = None
    #                                 surveySolutionExternalId = responseSearchSolutionApi['result']['data'][0]['externalId']
    #                             else:
    #                                 print("Solution fetch API failed")
    #                                 print("URL : " + urlSearchSolution)

    #                             solutionId = None
    #                             solutionId = responseCreateSolutionApi["result"]["solutionId"]
    #                             bodySolutionUpdate = {"creator": userDetails[2]}

    #                             return [solutionId, surveySolutionExternalId]
                            
    #                         else:
    #                             print("somethinghere i found")

    def schedule_deletion(returnPathStr):
        def delete_file():
            try:
                time.sleep(15)
                if os.path.exists(returnPathStr):
                    if os.path.isfile(returnPathStr):
                        os.remove(returnPathStr)
                        print(f"File {returnPathStr} deleted successfully.")

                    elif os.path.isdir(returnPathStr):
                        shutil.rmtree(returnPathStr)
                        print(f"Directory {returnPathStr} deleted successfully.")
                else:
                    print(f"File {returnPathStr} not found.")
            except Exception as e:
                print(f"Error deleting file: {e}")

        threading.Thread(target=delete_file, daemon=True).start()

    def terminatingMessage(msg):
        print(msg)
        sys.exit()

    def checkEntityOfSolution(projectName_for_folder_path, solutionNameOrId, accessToken):
        searchSolutionurl = internal_kong_ip + fetchsolutiondetails + "observation&page=1&limit=100&search=" + solutionNameOrId

        searchSolutionpayload = {}
        searchSolutionheaders = {
            'X-authenticated-user-token': accessToken,
            'internal-access-token': internal_access_token,
            'Authorization': authorization
        }

        searchSolutionresponse = requests.request("GET", searchSolutionurl, headers=searchSolutionheaders,
                                                data=searchSolutionpayload)
        messageArr = ["Solution found",
                    "URL : " + str(searchSolutionurl),
                    "Status Code : " + str(searchSolutionresponse.status_code),
                    "Response : " + str(searchSolutionresponse.text)]

        if searchSolutionresponse.status_code == 200:
            searchSolutionjson = searchSolutionresponse.json()
            
            for listOfSoulution in range(0, len(searchSolutionjson["result"]["data"])):
                solution_id = searchSolutionjson["result"]["data"][listOfSoulution]["_id"]
                messageArr.append("solution found : " + str(solution_id))
                Helpers.createAPILog(projectName_for_folder_path, messageArr)
                print("searchSolutionApi Success")
                solutionDetailsurl = internal_kong_ip + fetchsolutiondoc + solution_id

                solutionDetailspayload = {}
                solutionDetailsheaders = {
                    'X-authenticated-user-token': accessToken,
                    'internal-access-token': internal_access_token,
                    'Authorization': authorization
                }

                solutionDetailsresponse = requests.request("GET", solutionDetailsurl, headers=solutionDetailsheaders,
                                                        data=solutionDetailspayload)

                messageArr = ["Task solution Entity Type found",
                            "URL : " + str(solutionDetailsurl),
                            "Status Code : " + str(solutionDetailsresponse.status_code),
                            "Response : " + str(solutionDetailsresponse.text)]

                if solutionDetailsresponse.status_code == 200:
                    solutionDetailsjson = solutionDetailsresponse.json()
                    if solutionDetailsjson["result"]["isReusable"] == False:
                        solutionEntityType = solutionDetailsjson["result"]["entityType"]
                        solutionExternalId = solutionDetailsjson["result"]["externalId"]
                        messageArr.append("Task solution Entity Type found : " + str(solutionEntityType))
                        Helpers.createAPILog(projectName_for_folder_path, messageArr)
                        print("FetchSolutionDocApi Success")
                        break
                else:
                    
                    messageArr = ["Solution found",
                        "URL : " + str(searchSolutionurl),
                        "Status Code : " + str(searchSolutionresponse.status_code),
                        "Response : " + str(searchSolutionresponse.text)]
                    Helpers.createAPILog(projectName_for_folder_path, messageArr)
                    Helpers.terminatingMessage("FetchSolutionDocApi is failed")

        else:
            Helpers.terminatingMessage("search solution api is failed")
        return [solutionEntityType, solutionExternalId]

    def check_sequence(arr):
        for i in range(1, len(arr)):
            if arr[i] != arr[i - 1] + 1:
                return False
        return True
    
    def createAPILog(solutionName_for_folder_path, messageArr):
        file_exists = solutionName_for_folder_path + '/apiHitLogs/apiLogs.txt'
        # check if the file existis or not and create a file 
        if not path.exists(file_exists):
            API_log = open(file_exists, "w", encoding='utf-8')
            API_log.write("===============================================================================")
            API_log.write("\n")
            API_log.write("ENVIRONMENT : " + str(environment))
            API_log.write("\n")
            API_log.write("===============================================================================")
            API_log.write("\n")
            API_log.close()

        API_log = open(file_exists, "a", encoding='utf-8')
        API_log.write("\n")
        for msg in messageArr:
            API_log.write(msg)
            API_log.write("\n")
        API_log.close()

    def apicheckslog(solutionName_for_folder_path, messageArr):
        file_exists = solutionName_for_folder_path + '/apiHitLogs/apiLogs.csv'
        # global fileheader
        fileheader = ["Resource","Process","Status","Remark"]

        if not path.exists(file_exists):
            with open(file_exists, 'w', newline='',encoding='utf-8') as file:
                writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                writer.writerows([fileheader])
        with open(file_exists, 'a', newline='',encoding='utf-8') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
            writer.writerows([messageArr])
    
    def validateSheets(filePathAddObs, accessToken, parentFolder):
        global criteriaLevelsReport, scopeRoles, criteriaLevels, scopeEntityType , ccRootOrgName , ccRootOrgId
        wbObservation1 = xlrd.open_workbook(filePathAddObs, on_demand=True)
        sheetNames1 = wbObservation1.sheet_names()
        ecmIds = list()
        criteriaLevels = list()
        criteriaExternalIds = list()
        rubrics_sheet_names = ['Instructions', 'details', 'framework', 'ECMs or Domains', 'questions','Criteria_Rubric-Scoring', 'Domain(theme)_rubric_scoring']
        rubrics_sheet_IMP_names = ['Instructions', 'details', 'framework', 'ECMs or Domains', 'questions','Criteria_Rubric-Scoring', 'Domain(theme)_rubric_scoring', 'Imp mapping']
        observation_sheet_names = ['Instructions', 'details', 'criteria', 'questions']
        survey_sheet_names = ['Instructions', 'details', 'questions']
        project_sheet_names = ['Instructions', 'Project upload', 'Tasks upload','Certificate details']

        # 1-with rubrics , 2 - with out rubrics , 3 - survey , 4 - Project 5 - With rubric and IMP
        typeofSolutin = 0

        global environment, observationId, solutionName, pointBasedValue, entityType, allow_multiple_submissions, programName, userEntity, roles, isProgramnamePresent, solutionLanguage, keyWords, entityTypeId, solutionDescription, creator, dikshaLoginId
        if (len(rubrics_sheet_names) == len(sheetNames1)) and ((set(rubrics_sheet_names) == set(sheetNames1))):
            print("--->Observation with rubrics file detected.<---")
            typeofSolutin = 1
        elif (len(observation_sheet_names) == len(sheetNames1)) and ((set(observation_sheet_names) == set(sheetNames1))):
            print("--->Observation without rubrics file detected.<---")
            typeofSolutin = 2
        elif (len(survey_sheet_names) == len(sheetNames1)) and ((set(survey_sheet_names) == set(sheetNames1))):
            print("--->Survey file detected.<---")
            typeofSolutin = 3
        elif (len(project_sheet_names) == len(sheetNames1)) and ((set(project_sheet_names) == set(sheetNames1))):
            print("--->Project file detected.<---")
            typeofSolutin = 4
        elif (len(rubrics_sheet_IMP_names) == len(sheetNames1)) and ((set(rubrics_sheet_IMP_names) == set(sheetNames1))):
            print("--->Observation with rubrics and IMP file detected.<---")
            typeofSolutin = 5
        else:
            typeofSolutin = 0
            print(typeofSolutin)
            Helpers.terminatingMessage("Please check the Input sheet.")
        
        if typeofSolutin == 1 or typeofSolutin == 5:
            for sheetEnv in sheetNames1:
                questionsequenceArr =[]
                if sheetEnv == "Instructions":
                    pass
                else:
                    if sheetEnv.strip().lower() == 'details':
                        print("--->Checking details sheet...")
                        detailsCols = ["observation_solution_name", "observation_solution_description", "Diksha_loginId","Name_of_the_creator", "language", "allow_multiple_submissions", "keywords","scoring_system", "entity_type"]
                        detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                        keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                                range(detailsEnvSheet.ncols)]
                        for row_index_env in range(2, detailsEnvSheet.nrows):
                            dictDetailsEnv = {
                                keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                col_index_env in range(detailsEnvSheet.ncols)}
                            if set(detailsCols) == set(dictDetailsEnv.keys()):
                                solutionName = dictDetailsEnv['observation_solution_name'].encode('utf-8').decode('utf-8') if dictDetailsEnv['observation_solution_name'] else Helpers.terminatingMessage("\"observation_solution_name\" must not be Empty in \"details\" sheet")
                                dikshaLoginId = dictDetailsEnv['Diksha_loginId'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Diksha_loginId'] else Helpers.terminatingMessage("\"Diksha_loginId\" must not be Empty in \"details\" sheet")
                                ccUserDetails = Helpers.fetchUserDetails( accessToken, dikshaLoginId)
                                if not "CONTENT_CREATOR" in ccUserDetails[3]:
                                    Helpers.terminatingMessage("---> "+dikshaLoginId +" is not a CONTENT_CREATOR in Diksha " + environment)
                                ccRootOrgName = ccUserDetails[4]
                                ccRootOrgId = ccUserDetails[5]
                                solutionDescription = dictDetailsEnv['observation_solution_description'].encode('utf-8').decode('utf-8')
                                pointBasedValue = str(dictDetailsEnv['scoring_system']).encode('utf-8').decode('utf-8') if dictDetailsEnv['scoring_system'] else Helpers.terminatingMessage("\"scoring_system\" must not be Empty in \"details\" sheet")
                                # print(pointBasedValue,"<---------------------------------------pointBasedValue=======")
                                entityType = dictDetailsEnv['entity_type'].encode('utf-8').decode('utf-8') if dictDetailsEnv['entity_type'] else Helpers.terminatingMessage("\"entity_type\" must not be Empty in \"details\" sheet")

                                solutionLanguage = dictDetailsEnv['language'].split(",") if dictDetailsEnv['language'] else [""]
                                keyWords = dictDetailsEnv['keywords'].encode('utf-8').decode('utf-8')
                                creator = dictDetailsEnv['Name_of_the_creator'].encode('utf-8').decode('utf-8')  if dictDetailsEnv['Name_of_the_creator'] else Helpers.terminatingMessage("\"Name_of_the_creator\" must not be Empty in \"details\" sheet")
                                allow_multiple_submissions = dictDetailsEnv['allow_multiple_submissions']
                                if allow_multiple_submissions == 1 or allow_multiple_submissions == 'TRUE':
                                    allow_multiple_submissions = True
                                else:
                                    allow_multiple_submissions = False

                                scopeEntityType = scopeEntityType

                                isProgramnamePresent = False
                                if programName == "":
                                    isProgramnamePresent = False
                                else:
                                    isProgramnamePresent = True
                                    Helpers.getProgramInfo(accessToken, parentFolder, programName)
                            else:
                                Helpers.terminatingMessage("--->Columns Mismatch in Details Sheet.")
                    if sheetEnv.strip().lower() == 'framework':
                        frameworkCols = ["Domain ID", "Domain Name", "Criteria ID", "criteria_name", "L1 description","L2 description", "L3 description"]
                        print("--->Checking frameworks sheet...")
                        detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                        keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                                range(detailsEnvSheet.ncols)]
                        listOfThemeCriteria = list()
                        for row_index_env in range(1, detailsEnvSheet.nrows):
                            dictDetailsEnv = {
                                keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                col_index_env in range(detailsEnvSheet.ncols)}
                            countLevelUp = 1
                            for eachColNameCheck in keysEnv:
                                if "L" + str(countLevelUp) + " description" == eachColNameCheck:
                                    countLevelUp += 1
                            for i in range(1, countLevelUp):
                                if not i in criteriaLevels:
                                    criteriaLevels.append(i)

                            if dictDetailsEnv['Criteria ID'].encode('utf-8').decode('utf-8'):
                                if not [dictDetailsEnv['Domain ID'], dictDetailsEnv['Criteria ID']] in listOfThemeCriteria:
                                    listOfThemeCriteria.append([dictDetailsEnv['Domain ID'], dictDetailsEnv['Criteria ID']])
                                else:
                                    Helpers.terminatingMessage("Theme , criteria combo repeating in framework sheet.")
                            if not dictDetailsEnv['Domain ID']:
                                Helpers.terminatingMessage("Domain ID cannot be empty in framework sheet.")
                            if not dictDetailsEnv['Domain Name']:
                                Helpers.terminatingMessage("Theme cannot be empty in framework sheet.")

                            if dictDetailsEnv['Criteria ID']:
                                criteriaExternalIds.append(dictDetailsEnv['Criteria ID'].lower())
                    if sheetEnv.strip().lower() == 'ecms or domains':
                        print("--->Checking ECMs sheet...")
                        global ecmToSection
                        detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                        keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                                range(detailsEnvSheet.ncols)]
                        for row_index_env in range(2, detailsEnvSheet.nrows):
                            dictDetailsEnv = {
                                keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                col_index_env in range(detailsEnvSheet.ncols)}
                            if dictDetailsEnv['ECM Id/Domian ID'].lower() not in ecmIds:
                                ecmIds.append(dictDetailsEnv['ECM Id/Domian ID'].lower())
                            if not dictDetailsEnv['ECM Id/Domian ID']:
                                Helpers.terminatingMessage("ECM Id/Domian ID cannot be empty in ecm\'s sheet.")
                            if not dictDetailsEnv['section_id']:
                                Helpers.terminatingMessage("section_id cannot be empty in ecm\'s sheet.")
                            if not dictDetailsEnv['section_name']:
                                Helpers.terminatingMessage("section_name cannot be empty in ecm\'s sheet.")
                            if not dictDetailsEnv['ECM Name/Domain Name']:
                                Helpers.terminatingMessage("ECM Name/Domain Name cannot be empty in ecm\'s sheet.")
                            ecmToSection[dictDetailsEnv['section_id']] = dictDetailsEnv['ECM Id/Domian ID']
                    if sheetEnv.strip().lower() == 'questions':
                        print("--->Checking questions sheet...")
                        quesExtIds = list()
                        detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                        keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                                range(detailsEnvSheet.ncols)]
                        global numberOfResponses
                        numberOfResponses = 0
                        for qKeys in keysEnv:
                            countRespo = re.search(r"response\(R[0-9]|[1-9][0-9]|100\)$", qKeys)
                            if countRespo and not "_hint" in qKeys and "response" in qKeys:
                                numberOfResponses += 1

                        for n in range(1, numberOfResponses + 1):
                            if not "Score for R" + str(n) in keysEnv or not "response(R" + str(n) + ")_hint" in keysEnv:
                                Helpers.terminatingMessage("Mandatory Key: " + "Score for R" + str(n) + " or " + "response(R" + str(
                                    n) + ")_hint is missing")
                        for row_index_env in range(2, detailsEnvSheet.nrows):
                            dictDetailsEnv = {
                                keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                col_index_env in range(detailsEnvSheet.ncols)}
                            quesExtIds.append(dictDetailsEnv['question_id'].encode('utf-8').decode('utf-8').lower())

                            if not dictDetailsEnv['criteria_id']:
                                Helpers.terminatingMessage("criteria_id cannot be empty in questions sheet.")
                            if not dictDetailsEnv['criteria_id'].lower() in criteriaExternalIds:
                                Helpers.terminatingMessage("Criteria ID : " + dictDetailsEnv['criteria_id'] + " in question sheet not present in criteria sheet.")
                            question_sequence = dictDetailsEnv['question_sequence'] if dictDetailsEnv['question_sequence'] else Helpers.terminatingMessage("\"question_sequence\" must not be Empty in \"questions\" sheet")

                            questionsequenceArr.append(question_sequence)
                            question_sequence_arr = questionsequenceArr

                            if not dictDetailsEnv['question_primary_language']:
                                Helpers.terminatingMessage("question_primary_language cannot be empty in questions sheet.")
                            if not dictDetailsEnv['question_response_type']:
                                Helpers.terminatingMessage("question_response_type cannot be empty in questions sheet.")
                            if not dictDetailsEnv['question_id']:
                                Helpers.terminatingMessage("question_id cannot be empty in questions sheet.")
                            if not dictDetailsEnv['criteria_id']:
                                Helpers.terminatingMessage("criteria_id : " + str(
                                    dictDetailsEnv['criteria_id']) + "  cannot be empty in questions sheet.")
                            if not dictDetailsEnv['criteria_id'].lower() in criteriaExternalIds:
                                Helpers.terminatingMessage("criteria_id : " + str(dictDetailsEnv['criteria_id']) + " in questions sheet is not matching the criteria upload.")
                        if not len(question_sequence_arr) == len(set(question_sequence_arr)):
                                # print(question_sequence_arr)
                                # print(set(question_sequence_arr),"------------================")
                                Helpers.terminatingMessage("\"question_sequence\" must be Unique in \"questions\" sheet")
                        if not len(quesExtIds) == len(set(quesExtIds)):
                            Helpers.terminatingMessage("Duplicate question_id detected in questions sheet.")
                        if not Helpers.check_sequence(question_sequence_arr): Helpers.terminatingMessage("\"question_sequence\" must be in sequence in \"questions\" sheet")
                    if typeofSolutin == 5:
                        if sheetEnv.strip().lower() == 'imp mapping':
                            print("--->Checking Imp mapping sheet...")
                            global countImps
                            countImps = 1
                            detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                            keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                                    range(detailsEnvSheet.ncols)]
                            for row_index_env in range(2, detailsEnvSheet.nrows):
                                dictDetailsEnv = {
                                    keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                    col_index_env in range(detailsEnvSheet.ncols)}
                            for eachCols in dictDetailsEnv.keys():
                                if eachCols.strip() == "L" + str(countImps) + "-improvement-projects":
                                    countImps += 1
                            countImps = countImps - 1

                    if not pointBasedValue.lower() == "null":
                        if sheetEnv.strip().lower() == 'Criteria_Rubric-Scoring':
                            print("--->Checking Criteria Rubrics sheet")
                            cR_extIds = list()
                            detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                            keysEnv = [detailsEnvSheet.cell(0, col_index_env).value for col_index_env in
                                    range(detailsEnvSheet.ncols)]
                            listOfCRs = ["criteriaId", "weightage"]
                            for cl in criteriaLevels:
                                listOfCRs.append("L" + str(cl))
                            for keyys in keysEnv:
                                if not keyys in listOfCRs:
                                    print("--->" + keyys + " : unwanted column detected...")
                                    print("==>PS :  unwanted column will be ignored while uploading...")
                            for row_index_env in range(1, detailsEnvSheet.nrows):
                                dictDetailsEnv = {
                                    keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                    col_index_env in range(detailsEnvSheet.ncols)}
                                cR_extIds.append(dictDetailsEnv['criteriaId'].lower())
                                for cl in criteriaLevels:
                                    if not dictDetailsEnv["L" + str(cl)]:
                                        Helpers.terminatingMessage("L" + str(cl) + " must not be empty in criteria_rubric.")
                                if dictDetailsEnv['criteriaId']:
                                    Helpers.terminatingMessage("criteriaId must be empty in criteria_rubric sheet.")
                                if not dictDetailsEnv['weightage']:
                                    Helpers.terminatingMessage("weightage cannot be empty in criteria_rubric sheet.")
                            if not len(cR_extIds) == len(set(cR_extIds)):
                                Helpers.terminatingMessage("Duplicate externalId detected in criteria_rubric sheet.")
                        if sheetEnv.strip().lower() == 'Domain(theme)_rubric_scoring':
                            print("--->Checking Theme Rubrics sheet")
                            detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                            keysEnv = [detailsEnvSheet.cell(0, col_index_env).value for col_index_env in
                                    range(detailsEnvSheet.ncols)]
                            for row_index_env in range(1, detailsEnvSheet.nrows):
                                dictDetailsEnv = {
                                    keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                    col_index_env in range(detailsEnvSheet.ncols)}
                                if not dictDetailsEnv['domain_Id']:
                                    Helpers.terminatingMessage("domain_Id cannot be empty in theme_rubric sheet.")
                                if not dictDetailsEnv['domain_name']:
                                    Helpers.terminatingMessage("domain_name cannot be empty in theme_rubric sheet.")
                                if not dictDetailsEnv['weightage']:
                                    Helpers.terminatingMessage("weightage cannot be empty in theme_rubric sheet.")

        if typeofSolutin == 2:
            questionsequenceArr =[]
            # Point based value set as null by default for observation without rubrics
            pointBasedValue = "null"
            criteria_id_arr = []
            detailsColNames = ['observation_solution_name', 'observation_solution_description', 'Diksha_loginId','language', 'keywords', 'entity_type', "scope_entity"]
            criteriaColNames = ['criteria_id', 'criteria_name']
            questionsColNames = ["criteria_id","question_sequence","question_id","instance_parent_question_id","parent_question_id","show_when_parent_question_value_is","parent_question_value","page","question_number","question_primary_language","question_secondory_language","question_tip","question_hint","instance_identifier","question_response_type","date_auto_capture","response_required","min_number_value","max_number_value","file_upload","show_remarks","response(R1)","response(R1)_hint","response(R2)","response(R2)_hint","response(R3)","response(R3)_hint","response(R4)","response(R4)_hint","response(R5)","response(R5)_hint","response(R6)","response(R6)_hint","response(R7)","response(R7)_hint","response(R8)","response(R8)_hint","response(R9)","response(R9)_hint","response(R10)","response(R10)_hint","response(R11)","response(R11)_hint","response(R12)","response(R12)_hint","response(R13)","response(R13)_hint","response(R14)","response(R14)_hint","response(R15)","response(R15)_hint","response(R16)","response(R16)_hint","response(R17)","response(R17)_hint","response(R18)","response(R18)_hint","response(R19)","response(R19)_hint","response(R20)","response(R20)_hint","question_weightage","section_header"]
            for sheetColCheck in sheetNames1:
                if sheetColCheck.strip().lower() == 'details':
                    detailsColCheck = wbObservation1.sheet_by_name(sheetColCheck)
                    keysColCheckDetai = [detailsColCheck.cell(0, col_index_check).value for col_index_check in
                                        range(detailsColCheck.ncols)]
                    if len(keysColCheckDetai) != len(detailsColNames):
                        Helpers.terminatingMessage('Columns is missing in details sheet')
                if sheetColCheck.strip().lower() == 'criteria':
                    criteriaColCheck = wbObservation1.sheet_by_name(sheetColCheck)
                    keysColCheckCrit = [criteriaColCheck.cell(0, col_index_check1).value for col_index_check1 in
                                        range(criteriaColCheck.ncols)]
                    if len(keysColCheckCrit) != len(criteriaColNames):
                        Helpers.terminatingMessage('Columns is missing in criteria sheet')
                if sheetColCheck.strip().lower() == 'questions':
                    questionsColCheck = wbObservation1.sheet_by_name(sheetColCheck)
                    keysColCheckQues = [questionsColCheck.cell(0, col_index_check2).value for col_index_check2 in
                                        range(questionsColCheck.ncols)]
                    if len(keysColCheckQues) != len(questionsColNames):
                        Helpers.terminatingMessage('Columns is missing in questions sheet')
            for sheetEnv in sheetNames1:
                if sheetEnv == "Instructions":
                    pass
                else:
                    if sheetEnv.strip().lower() == 'details':
                        print("--->Checking details sheet...")
                        detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                        keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                                range(detailsEnvSheet.ncols)]
                        for row_index_env in range(2, detailsEnvSheet.nrows):
                            dictDetailsEnv = {
                                keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                col_index_env in range(detailsEnvSheet.ncols)}
                            solutionName = dictDetailsEnv['observation_solution_name'].encode('utf-8').decode('utf-8') if dictDetailsEnv['observation_solution_name'] else Helpers.terminatingMessage("\"observation_solution_name\" must not be Empty in \"details\" sheet")
                            solutionDescription = dictDetailsEnv['observation_solution_description'].encode('utf-8').decode('utf-8') if dictDetailsEnv['observation_solution_description'] else Helpers.terminatingMessage("\"observation_solution_description\" must not be Empty in \"details\" sheet")
                            dikshaLoginId = dictDetailsEnv['Diksha_loginId'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Diksha_loginId'] else Helpers.terminatingMessage("\"Diksha_loginId\" must not be Empty in \"details\" sheet")
                            creator = dictDetailsEnv['Name_of_the_creator'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Name_of_the_creator'] else Helpers.terminatingMessage("\"Name_of_the_creator\" must not be Empty in \"details\" sheet")
                            ccUserDetails = Helpers.fetchUserDetails(accessToken, dikshaLoginId)
                            if not "CONTENT_CREATOR" in ccUserDetails[3]:
                                Helpers.terminatingMessage("---> "+dikshaLoginId +" is not a CONTENT_CREATOR in Diksha " + environment)
                            ccRootOrgName = ccUserDetails[4]
                            ccRootOrgId = ccUserDetails[5]
                                
                            entityType = dictDetailsEnv['entity_type'].encode('utf-8').decode('utf-8') if dictDetailsEnv['entity_type'] else Helpers.terminatingMessage("\"entity_type\" must not be Empty in \"details\" sheet")
                            solutionLanguage = dictDetailsEnv['language'].encode('utf-8').decode('utf-8').split(",") if dictDetailsEnv['language'] else [""]
                            Helpers.getProgramInfo(accessToken, parentFolder, programNameInp)
                    elif sheetEnv.strip().lower() == 'criteria':
                        print("--->Checking criteria sheet...")
                        detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                        keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                                range(detailsEnvSheet.ncols)]
                        for row_index_env in range(2, detailsEnvSheet.nrows):
                            dictDetailsEnv = {
                                keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                col_index_env in range(detailsEnvSheet.ncols)}
                            criteria_id = dictDetailsEnv['criteria_id'].encode('utf-8').decode('utf-8') if dictDetailsEnv['criteria_id'] else Helpers.terminatingMessage("\"criteria_id\" must not be Empty in \"criteria\" sheet")
                            criteria_name = dictDetailsEnv['criteria_name'].encode('utf-8').decode('utf-8') if dictDetailsEnv['criteria_name'] else Helpers.terminatingMessage("\"criteria_name\" must not be Empty in \"criteria\" sheet")
                            criteria_id_arr.append(criteria_id)
                        if not len(criteria_id_arr) == len(set(criteria_id_arr)):
                            Helpers.terminatingMessage("\"criteria_id\" must be Unique in \"criteria\" sheet")
                    elif sheetEnv.strip().lower() == 'questions':
                        print("--->Checking question sheet...")
                        detailsEnvSheet = wbObservation1.sheet_by_name(sheetEnv)
                        ques_id_arr = list()
                        keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                                range(detailsEnvSheet.ncols)]
                        for row_index_env in range(2, detailsEnvSheet.nrows):
                            dictDetailsEnv = {
                                keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for
                                col_index_env in range(detailsEnvSheet.ncols)}
                            criteria_id = dictDetailsEnv['criteria_id'].encode('utf-8').decode('utf-8') if dictDetailsEnv['criteria_id'] else Helpers.terminatingMessage("\"criteria_id\" must not be Empty in \"questions\" sheet")
                            question_sequence = dictDetailsEnv['question_sequence'] if dictDetailsEnv['question_sequence'] else Helpers.terminatingMessage("\"question_sequence\" must not be Empty in \"questions\" sheet")

                            questionsequenceArr.append(question_sequence)
                            question_sequence_arr = questionsequenceArr

                            if not criteria_id in criteria_id_arr:
                                Helpers.terminatingMessage("\"criteria_id\" in \"Questions\" sheet must be declared in \"criteria\" sheet")
                            page = dictDetailsEnv['page'].encode('utf-8').decode('utf-8') if dictDetailsEnv['page'] else Helpers.terminatingMessage("\"page\" must not be Empty in \"questions\" sheet")
                            question_number = dictDetailsEnv['question_number'] if dictDetailsEnv['question_number'] else Helpers.terminatingMessage("\"question_number\" must not be Empty in \"questions\" sheet")
                            question_primary_language = dictDetailsEnv['question_primary_language'].encode('utf-8').decode('utf-8') if dictDetailsEnv['question_primary_language'] else Helpers.terminatingMessage("\"question_primary_language\" must not be Empty in \"questions\" sheet")
                            
                            response_required = dictDetailsEnv['response_required'] if str(dictDetailsEnv['response_required']) else Helpers.terminatingMessage("\"response_required\" must not be Empty in \"questions\" sheet")

                            question_id = dictDetailsEnv['question_id'] if dictDetailsEnv['question_id'] else Helpers.terminatingMessage("\"question_id\" must not be Empty in \"questions\" sheet")
                            ques_id_arr.append(question_id)
                            parent_question_id = dictDetailsEnv['question_id']
                            if parent_question_id and not parent_question_id in ques_id_arr:
                                Helpers.terminatingMessage("parent_question_id referenced before assigning in questions sheet.")
                            question_response_type = dictDetailsEnv['question_response_type'].encode('utf-8').decode('utf-8') if dictDetailsEnv[
                                'question_response_type'] else Helpers.terminatingMessage(
                                "\"question_response_type\" must not be Empty in \"questions\" sheet")
                        if not len(question_sequence_arr) == len(set(question_sequence_arr)):
                                Helpers.terminatingMessage("\"question_sequence\" must be Unique in \"questions\" sheet")
                        if not Helpers.check_sequence(question_sequence_arr): Helpers.terminatingMessage("\"question_sequence\" must be in sequence in \"questions\" sheet")

        if typeofSolutin == 3:
            print("Validating survey temp....")
            for sheetEnvCheck in sheetNames1:
                if sheetEnvCheck.strip().lower() == 'instructions' or sheetEnvCheck.strip().lower() == 'details' or sheetEnvCheck.strip().lower() == 'questions':
                    pass
                else:
                    Helpers.terminatingMessage('Sheet Names in excel file is wrong , Sheet Names are details,questions')

            detailsColNames = ["survey_solution_name", "survey_solution_description", "Name_of_the_creator","survey_creator_username", "survey_start_date", "survey_end_date"]
            questionsColNames = ["question_sequence", "question_id", "section_header", "instance_parent_question_id",
                                "parent_question_id", "show_when_parent_question_value_is", "parent_question_value",
                                "page", "question_number", "question_language1", "question_language2", "question_tip",
                                "question_hint", "instance_identifier", "question_response_type", "date_auto_capture",
                                "response_required", "min_number_value", "max_number_value", "file_upload", "show_remarks",
                                "response(R1)", "response(R2)", "response(R3)", "response(R4)", "response(R5)",
                                "response(R6)", "response(R7)", "response(R8)", "response(R9)", "response(R10)",
                                "response(R11)", "response(R12)", "response(R13)", "response(R14)", "response(R15)",
                                "response(R16)", "response(R17)", "response(R18)", "response(R19)", "response(R20)",
                                "response(R1)_hint", "response(R2)_hint", "response(R3)_hint", "response(R4)_hint",
                                "response(R5)_hint", "response(R6)_hint", "response(R7)_hint", "response(R8)_hint",
                                "response(R9)_hint", "response(R10)_hint", "response(R11)_hint", "response(R12)_hint",
                                "response(R13)_hint", "response(R14)_hint", "response(R15)_hint", "response(R16)_hint",
                                "response(R17)_hint", "response(R18)_hint", "response(R19)_hint", "response(R20)_hint"]

            for sheetColCheck in sheetNames1:
                # print(sheetColCheck,"sheetColCheck 2717")
                if sheetColCheck.strip().lower() == 'details':
                    detailsColCheck = wbObservation1.sheet_by_name(sheetColCheck)
                    keysColCheckDetai = [detailsColCheck.cell(0, col_index_check).value for col_index_check in
                                        range(detailsColCheck.ncols)]
                    if len(keysColCheckDetai) != len(detailsColNames):
                        Helpers.terminatingMessage('Some Columns are missing in details sheet')
                if sheetColCheck.strip().lower() == 'questions':
                    questionsColCheck = wbObservation1.sheet_by_name(sheetColCheck)
                    keysColCheckQues = [questionsColCheck.cell(1, col_index_check2).value for col_index_check2 in
                                        range(questionsColCheck.ncols)]
                    # print(keysColCheckQues)
                    if len(keysColCheckQues) != len(questionsColNames):
                        Helpers.terminatingMessage('Some Columns are missing in questions sheet')
                    for row_index_env in range(2, questionsColCheck.nrows):
                        dictDetailsEnv = {
                            keysColCheckQues[col_index_env]: questionsColCheck.cell(row_index_env, col_index_env).value for
                            col_index_env in range(questionsColCheck.ncols)}
                        question_sequenceSUR = dictDetailsEnv['question_sequence'] if dictDetailsEnv[
                            'question_sequence'] else Helpers.terminatingMessage(
                            "\"question_sequence\" must not be Empty in \"details\" sheet")
                        question_idSUR = dictDetailsEnv['question_id'].encode('utf-8').decode('utf-8') if dictDetailsEnv[
                            'question_id'] else Helpers.terminatingMessage("\"question_id\" must not be Empty in \"details\" sheet")
                        pageSUR = dictDetailsEnv['page'] if dictDetailsEnv['page'] else Helpers.terminatingMessage(
                            "\"page\" must not be Empty in \"details\" sheet")
                        question_numberSUR = dictDetailsEnv['question_number'] if dictDetailsEnv[
                            'question_number'] else Helpers.terminatingMessage(
                            "\"question_number\" must not be Empty in \"details\" sheet")
                        question_language1SUR = dictDetailsEnv['question_language1'].encode('utf-8').decode('utf-8') if not dictDetailsEnv['question_language1'] == None else Helpers.terminatingMessage(
                            "\"question_language1\" must not be Empty in \"details\" sheet")
                        question_response_typeSUR = dictDetailsEnv['question_response_type'] if dictDetailsEnv[
                            'question_response_type'] else Helpers.terminatingMessage(
                            "\"question_response_type\" must not be Empty in \"details\" sheet")
                        response_requiredSUR = dictDetailsEnv['response_required'] if dictDetailsEnv[
                            'response_required'] else Helpers.terminatingMessage(
                            "\"response_required\" must not be Empty in \"details\" sheet")
                        
        return typeofSolutin
       
    def mainFunc(MainFilePath, programFile, millisecond, isProgramnamePresent, isCourse,scopeEntityType=scopeEntityType):
        # print(addObservationSolution,"addObservationSolution 2693")
        scopeEntityType = scopeEntityType
        global surveySolutionlink, solutionlink, ObsSolutionLink, ObsRubricSolutionLink
        surveySolutionlink = None
        ObsSolutionLink = None
        solutionlink = None
        ObsRubricSolutionLink = None
        parentFolder = Helpers.createFileStruct(MainFilePath, addObservationSolution)
        # print(parentFolder,"2761")
        accessToken = Helpers.generateAccessToken(parentFolder)
        typeofSolution = Helpers.validateSheets(addObservationSolution, accessToken, parentFolder)
        wbObservation = xlrd.open_workbook(addObservationSolution, on_demand=True)
        # print("wbObservation",wbObservation)
        Helpers.programsFileCheck(programFile, accessToken, parentFolder, MainFilePath)
        wbprogram = xlrd.open_workbook(programFile, on_demand=True)
        programSheetNames = wbprogram.sheet_names()
        wbproject = xlrd.open_workbook(addObservationSolution, on_demand=True)
        projectSheetNames = wbproject.sheet_names()
        # print(projectSheetNames,"<--------------projectSheetNames 2770")
        for programSheets in programSheetNames:
            if programSheets.strip().lower() == 'program details':
                print("Checking program details sheet...")
                programDetailsSheet = wbprogram.sheet_by_name(programSheets)
                keysEnv = [programDetailsSheet.cell(1, col_index_env).value for col_index_env in
                           range(programDetailsSheet.ncols)]
                for row_index_env in range(2, programDetailsSheet.nrows):
                    dictProgramDetails = {
                        keysEnv[col_index_env]: programDetailsSheet.cell(row_index_env, col_index_env).value
                        for col_index_env in range(programDetailsSheet.ncols)}
                    programName = dictProgramDetails['Title of the Program'].encode('utf-8').decode('utf-8')
                    isProgramnamePresent = False
                    if programName == "":
                        isProgramnamePresent = False
                    else:
                        isProgramnamePresent = True
                    scopeEntityType = scopeEntityType
                    userEntity = dictProgramDetails['Targeted state at program level'].encode('utf-8').decode('utf-8').lstrip().rstrip().split(",") if dictProgramDetails['Targeted state at program level'] else Helpers.terminatingMessage("\"scope_entity\" must not be Empty in \"details\" sheet")
        
                    
        for sheets in projectSheetNames:
            if sheets.strip().lower() == 'details'.lower() and typeofSolution in [1, 5]:
                try:
                    ResourceSheet = wbproject.sheet_by_name(sheets)
                    keysEnv = [ResourceSheet.cell(1, col_index_env).value for col_index_env in range(ResourceSheet.ncols)]
                    dictDetailsEnv = {keysEnv[col_index_env]: ResourceSheet.cell(row_index_env, col_index_env).value
                                    for col_index_env in range(ResourceSheet.ncols)}
                    ObsWRResourceName = dictDetailsEnv['observation_solution_name'].encode('utf-8').decode('utf-8')
                except Exception as e:
                    print(f"Error reading 'details' sheet or processing observation solution name: {str(e)}")
                    return None  # Or handle accordingly

                try:
                    def addObsWRFunc(parentFolder, wbObservation, millisecond, accessToken):
                        try:
                            impLedObsFlag = True if typeofSolution == 5 else False
                            Helpers.criteriaUpload(parentFolder, wbObservation, millisecond, accessToken, "framework", impLedObsFlag)
                            print("Criteria Upload success....")
                        except Exception as e:
                            print(f"Error during criteria upload: {str(e)}")
                            return None  # Handle failure

                        try:
                            userDetails = Helpers.fetchUserDetails(accessToken, dikshaLoginId)
                            matchedShikshalokamLoginId = userDetails[0]
                        except Exception as e:
                            print(f"Error fetching user details: {str(e)}")
                            return None

                        try:
                            frameworkExternalId = Helpers.frameWorkUpload(parentFolder, matchedShikshalokamLoginId, wbObservation, millisecond, accessToken)
                            observationExternalId = frameworkExternalId + "-OBSERVATION-TEMPLATE"
                            Helpers.themesUpload(parentFolder, wbObservation, millisecond, accessToken, frameworkExternalId, False)
                            solutionId = Helpers.createSolutionFromFramework(parentFolder, wbObservation, accessToken, frameworkExternalId)
                        except Exception as e:
                            print(f"Error during framework or solution creation: {str(e)}")
                            return None

                        try:
                            # ECM processing
                            ecmsSheet = wbObservation.sheet_by_name('ECMs or Domains')
                            keys = [ecmsSheet.cell(1, col_index).value for col_index in range(ecmsSheet.ncols)]
                            ecm_update = dict()
                            ecm_dict = dict()
                            section = dict()
                            ecmSeqCount = 1
                            for row_index in range(2, ecmsSheet.nrows):
                                dictECMs = {keys[col_index]: ecmsSheet.cell(row_index, col_index).value for col_index in range(ecmsSheet.ncols)}
                                EMC_ID = dictECMs['ECM Id/Domian ID'].encode('utf-8').decode('utf-8').strip() + '_' + str(millisecond)
                                ECM_NAME = dictECMs['ECM Name/Domain Name'].encode('utf-8').decode('utf-8').strip()
                                section.update({dictECMs['section_id']: dictECMs['section_name']})
                                ecm_sections[EMC_ID] = dictECMs['section_id']
                                
                                # Handle boolean conversion safely
                                is_mandatory = dictECMs.get('Is ECM Mandatory?', 'FALSE')
                                is_mandatory = str(is_mandatory).strip().upper() in ['TRUE', '1']

                                ecm_update[EMC_ID] = {
                                    "externalId": EMC_ID, "tip": None, "name": ECM_NAME, "description": None,
                                    "modeOfCollection": "onfield", "canBeNotApplicable": not is_mandatory,
                                    "notApplicable": False, "canBeNotAllowed": not is_mandatory, "remarks": None,
                                    "sequenceNo": ecmSeqCount
                                }
                                ecmSeqCount += 1
                            ecm_dict['evidenceMethods'] = ecm_update
                            Helpers.solutionUpdate(parentFolder, accessToken, solutionId, ecm_dict)
                            Helpers.solutionUpdate(parentFolder, accessToken, solutionId, {"sections": section})
                        except Exception as e:
                            print(f"Error during ECM processing: {str(e)}")
                            return None

                        try:
                            # Continue rest of the process
                            bodySolutionUpdate = {"status": "active", "isDeleted": False, "criteriaLevelReport": criteriaLevelsReport}
                            Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)
                            Helpers.questionUpload(addObservationSolution, parentFolder, frameworkExternalId, millisecond, accessToken,solutionId,typeofSolution)
                            
                            # Handle rubrics
                            if pointBasedValue.lower() != "null":
                                bodySolutionUpdate = {"isRubricDriven": True}
                                Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)
                                Helpers.fetchSolutionCriteria(parentFolder, observationExternalId, accessToken)
                                Helpers.uploadCriteriaRubrics(parentFolder, wbObservation, millisecond, accessToken, frameworkExternalId, True)
                                Helpers.uploadThemeRubrics(parentFolder, wbObservation, accessToken, frameworkExternalId, True)
                            # Handle program information and start/end dates
                            bodySolutionUpdate = {'allowMultipleAssessemts': allow_multiple_submissions, "creator": creator}
                            Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)
                            solutionDetails = Helpers.fetchSolutionDetailsFromProgramSheet(parentFolder, programFile, solutionId, accessToken)
                            if solutionDetails[1]:
                                startDateArr = str(solutionDetails[1]).split("-")
                                bodySolutionUpdate = {"startDate": f"{startDateArr[2]}-{startDateArr[1]}-{startDateArr[0]} 00:00:00"}
                                Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)

                            if solutionDetails[2]:
                                endDateArr = str(solutionDetails[2]).split("-")
                                bodySolutionUpdate = {"endDate": f"{endDateArr[2]}-{endDateArr[1]}-{endDateArr[0]} 23:59:59"}
                                Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)

                            # If program name exists, handle child creation and linking
                            if isProgramnamePresent:
                                childId = Helpers.createChild(parentFolder, observationExternalId, accessToken)
                                if childId[0]:
                                    childSolutionDetails = Helpers.fetchSolutionDetailsFromProgramSheet(parentFolder, programFile, childId[0], accessToken)
                                    bodySolutionUpdate = {"scope": {"entityType": scopeEntityType, "entities": entitiesPGMID, "roles": childSolutionDetails[0]}}
                                    Helpers.solutionUpdate(parentFolder, accessToken, childId[0], bodySolutionUpdate)
                                    if solutionDetails[1]:
                                        startDateArr = str(solutionDetails[1]).split("-")
                                        bodySolutionUpdate = {
                                            "startDate": startDateArr[2] + "-" + startDateArr[1] + "-" + startDateArr[
                                                0] + " 00:00:00"}
                                        Helpers.solutionUpdate(parentFolder, accessToken, childId[0], bodySolutionUpdate)
                                    if solutionDetails[2]:
                                        endDateArr = str(solutionDetails[2]).split("-")
                                        bodySolutionUpdate = {
                                            "endDate": endDateArr[2] + "-" + endDateArr[1] + "-" + endDateArr[0] + " 23:59:59"}
                                        Helpers.solutionUpdate(parentFolder, accessToken, childId[0], bodySolutionUpdate)
                                    ObsRubricSolutionLink = Helpers.prepareProgramSuccessSheet(MainFilePath, parentFolder, programFile, childId[1], childId[0],
                                                            accessToken)
                        except Exception as e:
                            print(f"Error during rubric or program handling: {str(e)}")
                            return None

                        # Final link creation
                        finalObsRubricSolutionLink = {ObsWRResourceName: ObsRubricSolutionLink}
                        return finalObsRubricSolutionLink

                except Exception as e:
                    print(f"Error in 'addObsWRFunc': {str(e)}")
                    return None

                # Call function
                millisecond = int(time.time() * 1000)
                ObsWRSolutionLink = addObsWRFunc(parentFolder, wbObservation, millisecond, accessToken)
                return ObsWRSolutionLink


            elif sheets.strip().lower() == 'details'.lower() and typeofSolution == 2:
                ResourceSheet = wbproject.sheet_by_name(sheets)
                keysEnv = [ResourceSheet.cell(1, col_index_env).value for col_index_env in range(ResourceSheet.ncols)]
                
                # Collect observation solution details
                dictDetailsEnv = {
                    keysEnv[col_index_env]: ResourceSheet.cell(row_index_env, col_index_env).value
                    for col_index_env in range(ResourceSheet.ncols)
                }
                ObsWORResourceName = dictDetailsEnv['observation_solution_name'].encode('utf-8').decode('utf-8')
                
                try:
                    def addObsWORFunc(parentFolder, wbObservation, millisecond, accessToken):
                        print("Create Observation Function called ....")
                        
                        try:
                            # Step 1: Upload criteria
                            Helpers.criteriaUpload(parentFolder, wbObservation, millisecond, accessToken, "criteria", False)
                            print("-------------> criteria upload done")
                            
                            # Step 2: Process user details for Diksha_loginId
                            detailsEnvSheet = wbproject.sheet_by_name(sheets)
                            keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)]
                            
                            for row_index_env in range(2, detailsEnvSheet.nrows):
                                dictDetailsEnv = {
                                    keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                    for col_index_env in range(detailsEnvSheet.ncols)
                                }
                                if 'Diksha_loginId' in dictDetailsEnv:
                                    userDetails = Helpers.fetchUserDetails(accessToken, dictDetailsEnv['Diksha_loginId'])
                                    matchedShikshalokamLoginId = userDetails[0]
                                    print(f"Matched login ID: {matchedShikshalokamLoginId}")
                            
                            # Step 3: Upload framework and themes
                            frameworkExternalId = Helpers.frameWorkUpload(parentFolder, matchedShikshalokamLoginId, wbObservation, millisecond, accessToken)
                            observationExternalId = frameworkExternalId + "-OBSERVATION-TEMPLATE"
                            Helpers.themesUpload(parentFolder, wbObservation, millisecond, accessToken, frameworkExternalId, True)
                            
                            # Step 4: Create solution from framework
                            solutionId = Helpers.createSolutionFromFramework(parentFolder, wbObservation, accessToken, frameworkExternalId)
                            
                            # Step 5: Update solution with sections
                            sectionsObj = {"sections": {'S1': 'Observation Question'}}
                            Helpers.solutionUpdate(parentFolder, accessToken, solutionId, sectionsObj)
                            
                            # Step 6: Upload criteria and evidence methods
                            ecmObj = {
                                "evidenceMethods": {
                                    'OB': {
                                        'externalId': 'OB', 'tip': None, 'name': 'Observation', 'description': None,
                                        'modeOfCollection': 'onfield', 'canBeNotApplicable': False,
                                        'notApplicable': False, 'canBeNotAllowed': False, 'remarks': None
                                    }
                                }
                            }
                            Helpers.solutionUpdate(parentFolder, accessToken, solutionId, ecmObj)
                            Helpers.questionUpload(addObservationSolution, parentFolder, frameworkExternalId, millisecond, accessToken, solutionId, typeofSolution)
                            Helpers.fetchSolutionCriteria(parentFolder, observationExternalId, accessToken)
                            
                            # Handle point-based value and rubrics
                            if pointBasedValue.lower() != "null":
                                Helpers.uploadCriteriaRubrics(parentFolder, wbObservation, millisecond, accessToken, frameworkExternalId, False)
                                Helpers.uploadThemeRubrics(parentFolder, wbObservation, accessToken, frameworkExternalId, False)
                            
                            # Step 7: Activate and update solution status
                            bodySolutionUpdate = {"status": "active", "isDeleted": False, "allowMultipleAssessemts": True, "creator": creator}
                            Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)
                            
                            # Step 8: Update solution dates (start and end)
                            solutionDetails = Helpers.fetchSolutionDetailsFromProgramSheet(parentFolder, programFile, solutionId, accessToken)
                            if solutionDetails[1]:
                                startDateArr = str(solutionDetails[1]).split("-")
                                bodySolutionUpdate = {
                                    "startDate": f"{startDateArr[2]}-{startDateArr[1]}-{startDateArr[0]} 00:00:00"
                                }
                                Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)
                            if solutionDetails[2]:
                                endDateArr = str(solutionDetails[2]).split("-")
                                bodySolutionUpdate = {
                                    "endDate": f"{endDateArr[2]}-{endDateArr[1]}-{endDateArr[0]} 23:59:59"
                                }
                                Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)

                            # Step 9: Handle program name
                            if isProgramnamePresent:
                                childId = Helpers.createChild(parentFolder, observationExternalId, accessToken)
                                if childId[0]:
                                    solutionDetails = Helpers.fetchSolutionDetailsFromProgramSheet(parentFolder, programFile, childId[0], accessToken)
                                    scopeEntities = entitiesPGMID
                                    scopeRoles = solutionDetails[0]
                                    bodySolutionUpdate = {
                                        "scope": {"entityType": scopeEntityType, "entities": scopeEntities, "roles": scopeRoles}
                                    }
                                    Helpers.solutionUpdate(parentFolder, accessToken, childId[0], bodySolutionUpdate)
                                    if solutionDetails[1]:
                                        startDateArr = str(solutionDetails[1]).split("-")
                                        bodySolutionUpdate = {
                                            "startDate": f"{startDateArr[2]}-{startDateArr[1]}-{startDateArr[0]} 00:00:00"
                                        }
                                        Helpers.solutionUpdate(parentFolder, accessToken, childId[0], bodySolutionUpdate)
                                    if solutionDetails[2]:
                                        endDateArr = str(solutionDetails[2]).split("-")
                                        bodySolutionUpdate = {
                                            "endDate": f"{endDateArr[2]}-{endDateArr[1]}-{endDateArr[0]} 23:59:59"
                                        }
                                        Helpers.solutionUpdate(parentFolder, accessToken, childId[0], bodySolutionUpdate)
                                    ObsSolutionLink = Helpers.prepareProgramSuccessSheet(
                                        MainFilePath, parentFolder, programFile, childId[1], childId[0], accessToken
                                    )
                                    print(ObsSolutionLink)
                                else:
                                    print("Failed to create child observation.")
                            
                            finalObsSolutionLink = {ObsWORResourceName: ObsSolutionLink}
                            return finalObsSolutionLink

                        except Exception as e:
                            print(f"Error during observation creation: {str(e)}")
                            raise RuntimeError("Observation creation failed due to an unexpected error.")
                    
                    millisecond = int(time.time() * 1000)
                    ObsWORSolutionLink = addObsWORFunc(parentFolder, wbObservation, millisecond, accessToken)
                    return ObsWORSolutionLink
                
                except Exception as e:
                    print(f"Error occurred: {str(e)}")
                    raise RuntimeError("The process failed due to an unexpected error.")


            elif sheets.strip().lower() == 'Project upload'.lower() and typeofSolution == 4:
                print("Checking project upload sheet...")
                projectsheet = wbproject.sheet_by_name(sheets)
                keysEnv = [projectsheet.cell(1, col_index_env).value for col_index_env in range(projectsheet.ncols)]
                
                for row_index_env in range(1, projectsheet.nrows):
                    projectDetails = {keysEnv[col_index_env]: projectsheet.cell(row_index_env, col_index_env).value
                                    for col_index_env in range(projectsheet.ncols)}
                    ProjectName = projectDetails["title"].encode('utf-8').decode('utf-8')
                    entityType = "school"

                try:
                    def addProjectFunc(filePathAddProject, projectName_for_folder_path, millisAddObs):
                        print('Add Project Function Called')

                        # Create project folder if it doesn't exist
                        if not path.exists(projectName_for_folder_path):
                            os.mkdir(projectName_for_folder_path)

                        # Create a user input folder if it doesn't exist
                        if not path.exists(projectName_for_folder_path + "/user_input_file"):
                            os.mkdir(projectName_for_folder_path + "/user_input_file")
                        
                        # Copy files to the folder
                        shutil.copy(filePathAddProject, projectName_for_folder_path + "/user_input_file")
                        shutil.copy(programFile, projectName_for_folder_path + "/user_input_file")

                        # Log and add project details
                        wbproject = xlrd.open_workbook(filePathAddProject, on_demand=True)
                        projectsheetforcertificate = wbproject.sheet_names()
                        
                        # Process the project upload sheet
                        for prosheet in projectsheetforcertificate:
                            if prosheet.strip().lower() == 'Project upload'.lower():
                                detailsEnvSheet = wbproject.sheet_by_name(prosheet)
                                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)]
                                
                                for row_index_env in range(2, detailsEnvSheet.nrows):
                                    dictDetailsEnv = {
                                        keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                        for col_index_env in range(detailsEnvSheet.ncols)
                                    }

                                    # Handle projects without a certificate
                                    if str(dictDetailsEnv['has certificate']).lower() == 'no':
                                        Helpers.prepareProjectAndTasksSheets(addObservationSolution, projectName_for_folder_path, accessToken)
                                        Helpers.projectUpload(addObservationSolution, projectName_for_folder_path, accessToken)
                                        Helpers.taskUpload(addObservationSolution, projectName_for_folder_path, accessToken)
                                        
                                        ProjectSolutionResp = Helpers.solutionCreationAndMapping(
                                            projectName_for_folder_path, entityToUpload, listOfFoundRoles, accessToken, programFile)
                                        ProjectSolutionExternalId = ProjectSolutionResp[0]
                                        ProjectSolutionId = ProjectSolutionResp[1]
                                        
                                        solutionlink = Helpers.prepareProgramSuccessSheet(
                                            MainFilePath, projectName_for_folder_path, programFile,
                                            ProjectSolutionExternalId, ProjectSolutionId, accessToken)
                                        print(solutionlink)

                                    # Handle projects with a certificate
                                    elif str(dictDetailsEnv['has certificate']).lower() == 'yes':
                                        print("----> Certificate required for project <----")
                                        baseTemplate_id = Helpers.fetchCertificateBaseTemplate(
                                            filePathAddProject, accessToken, projectName_for_folder_path)
                                        Helpers.downloadlogosign(filePathAddProject, projectName_for_folder_path)
                                        Helpers.editsvg(accessToken, filePathAddProject, projectName_for_folder_path, baseTemplate_id)
                                        Helpers.prepareProjectAndTasksSheets(addObservationSolution, projectName_for_folder_path, accessToken)
                                        Helpers.projectUpload(addObservationSolution, projectName_for_folder_path, accessToken)
                                        Helpers.taskUpload(addObservationSolution, projectName_for_folder_path, accessToken)
                                        
                                        ProjectSolutionResp = Helpers.solutionCreationAndMapping(
                                            projectName_for_folder_path, entityToUpload, listOfFoundRoles, accessToken, programFile)
                                        ProjectSolutionExternalId = ProjectSolutionResp[0]
                                        ProjectSolutionId = ProjectSolutionResp[1]

                                        # Handle certificate template
                                        certificatetemplateid = Helpers.prepareaddingcertificatetemp(
                                            filePathAddProject, projectName_for_folder_path, accessToken, ProjectSolutionId, programID, baseTemplate_id)

                                        solutionlink = Helpers.prepareProgramSuccessSheet(
                                            MainFilePath, projectName_for_folder_path, programFile,
                                            ProjectSolutionExternalId, ProjectSolutionId, accessToken)
                                        
                                    finalprojectsolutionlink = {ProjectName: solutionlink}
                                    return finalprojectsolutionlink

                    # Calculate the current time in milliseconds
                    millisecond = int(time.time() * 1000)
                    projectSolutionLink = addProjectFunc(addObservationSolution, parentFolder, millisecond)
                    return projectSolutionLink

                except Exception as e:
                    print(f"Error occurred during project creation: {str(e)}")
                    raise RuntimeError("The project creation failed due to an unexpected error")

    
            elif sheets.strip().lower() == 'details'.lower() and typeofSolution == 3:
                try:
                    ResourceSheet = wbproject.sheet_by_name(sheets)
                    keysEnv = [ResourceSheet.cell(1, col_index_env).value for col_index_env in range(ResourceSheet.ncols)]
                    dictDetailsEnv = {keysEnv[col_index_env]: ResourceSheet.cell(row_index_env, col_index_env).value for col_index_env in range(ResourceSheet.ncols)}
                    SurveyResourceName = dictDetailsEnv['survey_solution_name'].encode('utf-8').decode('utf-8')

                    def addsurveyFunc(parentFolder, wbObservation, millisecond, accessToken):
                        try:
                            # Validate program file and survey sheets
                            Helpers.programsFileCheck(programFile, accessToken, parentFolder, MainFilePath)
                            wbprogram = xlrd.open_workbook(programFile, on_demand=True)
                            programSheetNames = wbprogram.sheet_names()

                            wbSurvey = xlrd.open_workbook(addObservationSolution, on_demand=True)
                            surevySheetNames = wbSurvey.sheet_names()

                            # Create survey solution
                            surveyResp = Helpers.createSurveySolution(parentFolder, wbSurvey, accessToken)
                            surTempExtID = surveyResp[1]
                            surTempSolID = surveyResp[0]

                            # Update solution status
                            bodySolutionUpdate = {"status": "active", "isDeleted": False}
                            Helpers.solutionUpdate(parentFolder, accessToken, surveyResp[0], bodySolutionUpdate)

                            # Upload survey questions
                            Helpers.uploadSurveyQuestions(MainFilePath, parentFolder, wbSurvey, addObservationSolution, accessToken, surTempExtID, surTempSolID, millisecond, programFile)

                            finalsurveySolutionlink = {SurveyResourceName: surveySolutionlink}
                            return finalsurveySolutionlink

                        except KeyError as e:
                            print(f"KeyError: {str(e)} - Possible missing column or incorrect key in sheet.")
                            return None
                        except xlrd.XLRDError as e:
                            print(f"XLRDError: {str(e)} - Issue with reading the Excel file.")
                            return None
                        except Exception as e:
                            print(f"An error occurred: {str(e)}")
                            return None

                    millisecond = int(time.time() * 1000)
                    surveySollink = addsurveyFunc(parentFolder, wbObservation, millisecond, accessToken)

                    if surveySollink:
                        return surveySollink
                    else:
                        print("Survey creation failed due to an error.")

                except KeyError as e:
                    print(f"KeyError: {str(e)} - Check if 'survey_solution_name' exists in the sheet.")
                except xlrd.XLRDError as e:
                    print(f"XLRDError: {str(e)} - Unable to load the Excel sheet: {sheets}")
                except Exception as e:
                    print(f"An unexpected error occurred: {str(e)}")



        # Helpers.SolutionFileCheck(addObservationSolution, accessToken, parentFolder, MainFilePath)
        # print(wbObservation,"wbObservation")
        # surveyResp = Helpers.createSurveySolution(parentFolder, wbObservation, accessToken)
        # surTempExtID = surveyResp[1]
        # # surveyChildId = surveyResp[0]
        # bodySolutionUpdate = {"status": "active", "isDeleted": False}
        # Helpers.solutionUpdate(parentFolder, accessToken, surveyResp[0], bodySolutionUpdate)
        # surveyChildId = Helpers.uploadSurveyQuestions(parentFolder, wbObservation, addObservationSolution, accessToken, surTempExtID,
        #                         surveyResp[0], millisecond)
        
        # local = os.getcwd()
        # sucessSheetName = Helpers.preparesolutionUploadSheet(MainFilePath,parentFolder,surveyChildId)
        # print("surveychild id",surveyChildId)
        # clickheretodownload = Helpers.uploadSuccessSheetToBucket(surveyChildId,sucessSheetName,accessToken)
        # Helpers.schedule_deletion(MainFilePath)
        # return [surveyChildId,local+'/'+sucessSheetName,clickheretodownload]
    
    
    def loadSurveyFile(programFile):
        MainFilePath = Helpers.createFileStructForProgram(programFile)
        global downloaded_file
        global addObservationSolution
        # if downloaded_file is None:
        downloaded_file = []
        print(downloaded_file, "downloaded_file 3044")

        wbPgm = xlrd.open_workbook(programFile, on_demand=True)
        sheetNames = wbPgm.sheet_names()
        pgmSheets = ["Instructions", "Program Details", "Resource Details", "Program Manager Details"]
        print(sheetNames)
        print(pgmSheets)

        solutionDict = {}
        programName = ""  # Initialize the programName variable

        if len(sheetNames) == len(pgmSheets) and sheetNames == pgmSheets:
            print("--->Program Template detected.<---")

            for sheetEnv in sheetNames:
                if sheetEnv.strip().lower() == 'program details':
                    print("Checking program details sheet...")
                    programDetailsSheet = wbPgm.sheet_by_name(sheetEnv)
                    keysEnv = [programDetailsSheet.cell(1, col_index_env).value for col_index_env in
                            range(programDetailsSheet.ncols)]
                    for row_index_env in range(2, programDetailsSheet.nrows):
                        dictProgramDetails = {
                            keysEnv[col_index_env]: programDetailsSheet.cell(row_index_env, col_index_env).value
                            for col_index_env in range(programDetailsSheet.ncols)
                        }
                        # Extracting the program name
                        programName = dictProgramDetails['Title of the Program'].encode('utf-8').decode('utf-8')
                        isProgramnamePresent = bool(programName)  # Check if program name exists
                        
                        # Example handling of userEntity
                        userEntity = dictProgramDetails['Targeted state at program level'].encode('utf-8').decode('utf-8').lstrip().rstrip().split(",")

                if sheetEnv.strip().lower() == 'resource details':
                    print("--->Checking Resource Details sheet...")
                    messageArr = []
                    messageArr.append("--->Checking Resource Details sheet...")
                    detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                    keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)]
                    
                    for row_index_env in range(2, detailsEnvSheet.nrows):
                        millisecond = int(time.time() * 1000)
                        dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                        for col_index_env in range(detailsEnvSheet.ncols)}
                        resourceNamePGM = dictDetailsEnv['Name of resources in program'].encode('utf-8').decode('utf-8')
                        resourceTypePGM = dictDetailsEnv['Type of resources'].encode('utf-8').decode('utf-8')
                        resourceLinkOrExtPGM = dictDetailsEnv['Resource Link']
                        
                        if str(dictDetailsEnv['Type of resources']).lower().strip() == "course":
                            isCourse = False
                        else:
                            isCourse = False
                            resourceStatus = dictDetailsEnv['Resource Status']
                            if resourceStatus.strip() == "New Upload":
                                print("--->Resource Name : " + str(resourceNamePGM))
                                resourceLinkOrExtPGM = str(resourceLinkOrExtPGM).split('/')[5]
                                file_url = 'https://docs.google.com/spreadsheets/d/' + resourceLinkOrExtPGM + '/export?format=xlsx'
                                if not os.path.isdir('InputFiles'):
                                    os.mkdir('InputFiles')
                                dest_file = 'InputFiles'
                                download_file = wget.download(file_url, dest_file)
                                downloaded_file.append(download_file)

            print("--->Solution input file successfully downloaded: " + str(downloaded_file))
            for addObservationSolution in downloaded_file:
                print(f"Processing file: {addObservationSolution}")
                solutionSL = Helpers.mainFunc(MainFilePath, programFile, millisecond, isProgramnamePresent, isCourse)
                for resourceName, solutionLink in solutionSL.items():
                    solutionDict[resourceName] = solutionLink
            downloaded_file = None

        # Combine solutionDict and programName into a single dictionary for returning
        result = {
            "solutionDict": solutionDict,
            "programName": programName  # Ensure programName is extracted from the 'Program Details' sheet
        }

        # print(f"Type of solutionDict: {type(solutionDict)}")
        # print(f"Program Name: {programName}")

        return json.dumps(result)

    
    # function to upload criteria   
    def criteriaUpload(solutionName_for_folder_path, wbObservation, millisAddObs, accessToken, tabName, projectDrivenFlag):
        criteriaColNames = ["criteriaId", "criteria_name"]
        criteriaSheet = wbObservation.sheet_by_name(tabName)
        keys = [criteriaSheet.cell(1, col_index).value for col_index in range(criteriaSheet.ncols)]
        criteriaUploadFieldnames = ['criteriaID', 'criteriaName']
        dictCriteriaToCsv = dict()
        criteriaLevelsFromFramework = dict()
        global criteriaLevelsCount
        if tabName == "framework":
            fetchLevelsFromFramework = wbObservation.sheet_by_name('framework')
            if projectDrivenFlag:
                criteriaImpDict = {}
                impsToCriteria = wbObservation.sheet_by_name('Imp mapping')
                keysFromImpSheet = [impsToCriteria.cell(1, col_index).value for col_index in range(impsToCriteria.ncols)]
                for row_indexImp in range(2, impsToCriteria.nrows):
                    dictImp = {keysFromImpSheet[col_index]: impsToCriteria.cell(row_indexImp, col_index).value for col_index in range(impsToCriteria.ncols)}
                    criteriaImpDict[dictImp['criteriaId'].strip()] = {}
                    for levls in range(1, countImps + 1):
                        criteriaImpDict[dictImp['criteriaId'].strip()].update({'L' + str(levls) + '-improvement-projects': dictImp['L' + str(levls) + '-improvement-projects'].strip()})

            keysFromFrameWork = [fetchLevelsFromFramework.cell(1, col_index).value for col_index in
                                range(fetchLevelsFromFramework.ncols)]
            levelCount = 1

            for eachHeaders in keysFromFrameWork:
                if eachHeaders == "L" + str(levelCount) + " description":
                    levelCount += 1
            levelCount = levelCount - 1

            for row_indexFrameWork in range(2, fetchLevelsFromFramework.nrows):
                dictFramework = {
                    keysFromFrameWork[col_index]: fetchLevelsFromFramework.cell(row_indexFrameWork, col_index).value for
                    col_index in range(fetchLevelsFromFramework.ncols)}
                criteriaLevelsFromFramework[dictFramework["Criteria ID"]] = {}

                for levlsNo in range(1, levelCount + 1):
                    criteriaLevelsFromFramework[dictFramework["Criteria ID"]].update(
                        {"L" + str(levlsNo): dictFramework["L" + str(levlsNo) + " description"]})
                    if not "L" + str(levlsNo) in criteriaColNames:
                        criteriaColNames.append("L" + str(levlsNo))

            for row_index in range(2, criteriaSheet.nrows):
                dictCriteria = {keys[col_index]: criteriaSheet.cell(row_index, col_index).value for col_index in
                                range(criteriaSheet.ncols)}
                dictCriteriaToCsv = {}

                dictCriteriaToCsv['criteriaID'] = dictCriteria['Criteria ID'].strip() + '_' + str(millisAddObs)
                criteriaLookUp[dictCriteriaToCsv['criteriaID'].strip()] = dictCriteria['Criteria Name'].encode('utf-8').decode('utf-8')
                dictCriteriaToCsv['criteriaName'] = dictCriteria['Criteria Name'].encode('utf-8').decode('utf-8')
                criteriaName = dictCriteria['Criteria Name'].encode('utf-8').decode('utf-8')
                dictCriteriaToCsv['type'] = 'auto'
                for levlsNo in range(1, levelCount + 1):
                    dictCriteriaToCsv['L' + str(levlsNo)] = dictCriteria["L" + str(levlsNo) + " description"]
                if projectDrivenFlag:
                    for eachImps in criteriaImpDict[dictCriteria['Criteria ID'].strip()]:
                        dictCriteriaToCsv[eachImps] = criteriaImpDict[dictCriteria['Criteria ID'].strip()][eachImps]

                if not 'type' in criteriaUploadFieldnames:
                    criteriaUploadFieldnames.append('type')
                for eachCols in criteriaColNames:
                    if not eachCols in ['criteria_id', 'criteria_name', 'type', "criteriaId"]:
                        if not eachCols in criteriaUploadFieldnames:
                            criteriaUploadFieldnames.append(eachCols)
                if projectDrivenFlag:
                    for levls in range(1, countImps + 1):
                        if not (str('L' + str(levls) + '-improvement-projects') in criteriaUploadFieldnames):
                            criteriaUploadFieldnames.append('L' + str(levls) + '-improvement-projects')
                criteriaFilePath = solutionName_for_folder_path + '/criteriaUpload/'
                file_exists = os.path.isfile(solutionName_for_folder_path + '/criteriaUpload/uploadSheet.csv')
                criteriaLevelsCount = levelCount
                if not os.path.exists(criteriaFilePath):
                    os.mkdir(criteriaFilePath)
                with open(solutionName_for_folder_path + '/criteriaUpload/uploadSheet.csv', 'a',encoding='utf-8') as criteriaUploadFile:
                    writerCriteriaUpload = csv.DictWriter(criteriaUploadFile, fieldnames=list(criteriaUploadFieldnames),
                                                        lineterminator='\n')
                    if not file_exists:
                        writerCriteriaUpload.writeheader()
                    writerCriteriaUpload.writerow(dictCriteriaToCsv)
                    
        elif tabName == "criteria":
            criteriaSheet = wbObservation.sheet_by_name(tabName)
            keys = [criteriaSheet.cell(1, col_index).value for col_index in range(criteriaSheet.ncols)]
            for row_index in range(2, criteriaSheet.nrows):
                dictCriteria = {keys[col_index]: criteriaSheet.cell(row_index, col_index).value for col_index in
                                range(criteriaSheet.ncols)}
                dictCriteria['criteriaID'] = dictCriteria['criteria_id'].encode('utf-8').decode('utf-8').strip() + '_' + str(millisAddObs)
                criteriaLookUp[dictCriteria['criteriaID']] = dictCriteria['criteria_name'].encode('utf-8').decode('utf-8')
                del dictCriteria['criteria_id']
                dictCriteria['criteriaName'] = dictCriteria['criteria_name'].encode('utf-8').decode('utf-8')
                criteriaName = dictCriteria['criteria_name']
                del dictCriteria['criteria_name']
                dictCriteria['L1'] = 'NA'
                dictCriteria['type'] = 'auto'
                criteriaFilePath = solutionName_for_folder_path + '/criteriaUpload/'
                file_exists = os.path.isfile(solutionName_for_folder_path + '/criteriaUpload/uploadSheet.csv')
                if not os.path.exists(criteriaFilePath):
                    os.mkdir(criteriaFilePath)
                criteriaUploadFieldnames = []
                criteriaUploadFieldnames = ['criteriaID', 'criteriaName', 'L1', 'type']
                with open(solutionName_for_folder_path + '/criteriaUpload/uploadSheet.csv', 'a',encoding='utf-8') as criteriaUploadFile:
                    writerCriteriaUpload = csv.DictWriter(criteriaUploadFile, fieldnames=criteriaUploadFieldnames,
                                                        lineterminator='\r')
                    if not file_exists:
                        writerCriteriaUpload.writeheader()
                    writerCriteriaUpload.writerow(dictCriteria)

        urlCriteriaUploadApi = internal_kong_ip + criteriauploadapiurl
        headerCriteriaUploadApi = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id
        }
        filesCriteria = {
            'criteria': open(solutionName_for_folder_path + '/criteriaUpload/uploadSheet.csv', 'rb')
        }

        responseCriteriaUploadApi = requests.post(url=urlCriteriaUploadApi, headers=headerCriteriaUploadApi,
                                                files=filesCriteria)
        messageArr = ["Criteria Upload Sheet Prepared.",
                    "File path : " + solutionName_for_folder_path + '/criteriaUpload/uploadSheet.csv']
        messageArr.append("Upload status code : " + str(responseCriteriaUploadApi.status_code))
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)

        if responseCriteriaUploadApi.status_code == 200:
            print('CriteriaUploadApi Success')
            with open(solutionName_for_folder_path + '/criteriaUpload/uploadInternalIdsSheet.csv', 'w+',encoding='utf-8') as criteriaRes:
                criteriaRes.write(responseCriteriaUploadApi.text)
        else:

            messageArr.append("Response : " + str(responseCriteriaUploadApi.text))
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print("Criteria Upload failed.")
            sys.exit()
    
    def frameWorkUpload(solutionName_for_folder_path, matchedShikshalokamLoginId, wbObservation, millisAddObs, accessToken):
        global criteriaLevelsReport
        dateTime = datetime.now()
        frameworkDocInsertObj = {}
        try:
            observationSheet = wbObservation.sheet_by_name("details")  # Reading the "details" sheet
        except xlrd.biffh.XLRDError:
            print("Error: 'details' sheet not found in wbObservation.")
            return None
        
        headers = [observationSheet.cell(1, col_index).value for col_index in range(observationSheet.ncols)]
        values = [observationSheet.cell(2, col_index).value for col_index in range(observationSheet.ncols)]        
        observationData = dict(zip(headers, values))
        solutionName = observationData.get('observation_solution_name', 'Default Solution Name')  # Replace 'Solution Name' with the actual header
        solutionDescription = observationData.get('observation_solution_description', 'Default Description')  # Replace 'Solution Description' with the actual header
        solutionKeywords = observationData.get('keywords', 'Default keywords')
        solutionEntityType = observationData.get('entity_type', 'Default entity_type')
        # Generating a unique External ID for the framework
        frameworkExternalId = str(uuid.uuid1())
        frameworkDocInsertObj['externalId'] = frameworkExternalId

        # Assigning the fetched name and description
        frameworkDocInsertObj['name'] = solutionName
        frameworkDocInsertObj['description'] = solutionDescription
        frameworkDocInsertObj['parentId'] = None
        frameworkDocInsertObj['resourceType'] = ['Observations Framework']
        frameworkDocInsertObj['language'] = solutionLanguage
        frameworkDocInsertObj['levelToScoreMapping'] = dict()
        frameworkDocInsertObj['keywords'] = solutionKeywords
        keywords = frameworkDocInsertObj['keywords']

        # if keyWords and (keyWords != 'Framework' or keyWords != 'Frameworks' or keyWords != 'Observation' or keyWords != 'Observations'):
        #     keywordsFinalArr = ['Framework', 'Observation']
        #     keywordsArr = keyWords.encode('utf-8').decode('utf-8').split(',')
        #     for keyw in keywordsArr:
        #         keywordsFinalArr.append(keyw)
        #     frameworkDocInsertObj['keywords'] = keywordsFinalArr
        #     print(keywordsFinalArr,"<---------------------------------keywordsFinalArr")
        # else:
        #     frameworkDocInsertObj['keywords'] = ['Framework', 'Observation']
        frameworkDocInsertObj['concepts'] = []
        frameworkDocInsertObj['createdFor'] = [ccRootOrgId]  # createdForArr
        frameworkDocInsertObj['rootOrg'] = [ccRootOrgId]  # rootOrgArr
        
        criteriaFrameworkArr = []
        with open(solutionName_for_folder_path + '/criteriaUpload/uploadInternalIdsSheet.csv', 'r',encoding='utf-8') as criteriaInternalFile:
            criteriaInternalReader = csv.DictReader(criteriaInternalFile)
            criteriaWeightage = 100 / (len(list(criteriaInternalReader)))
            criteriaInternalFile.seek(0, 0)
            next(criteriaInternalReader, None)
            for crit in criteriaInternalReader:
                dictCritInter = {}
                dictCritInter = dict(crit)
                criteriaFrameworkObj = {
                    'criteriaId': str(ObjectId(dictCritInter['Criteria Internal Id'])),
                    'weightage': criteriaWeightage
                }
                criteriaFrameworkArr.append(criteriaFrameworkObj)
        criteriaInternalFile.close()
        frameworkDocInsertObj['themes'] = [{
            'type': 'theme',
            'label': 'theme',
            'name': 'Observation Theme',
            'externalId': 'OB',
            'weightage': 100,
            'criteria': criteriaFrameworkArr
        }]
        # print(pointBasedValue,"<-------------------point based value in framwork")
        if not pointBasedValue.lower() == "null":
            frameworkDocInsertObj['flattenedThemes'] = {
                "type": "theme",
                "label": "theme",
                "name": "Observation Theme",
                "externalId": "OB",
                "weightage": 1,
                "criteria": criteriaFrameworkArr,
                "rubric": {
                    "expressionVariables": {
                        "SCORE": "OB.sumOfPointsOfAllChildren()"
                    },
                    "levels": {
                        "L1": {
                            "expression": "(0<=SCORE<=100000)"
                        }
                    }
                },
                "hierarchyLevel": 0,
                "hierarchyTrack": []
            }
            frameworkDocInsertObj['scoringSystem'] = pointBasedValue
            frameworkDocInsertObj['isRubricDriven'] = True
            criteriaLevelsReport = True
            frameworkDocInsertObj['themes'] = [{
                'type': 'theme',
                'label': 'theme',
                'name': 'Observation Theme',
                'externalId': 'OB',
                'weightage': 100,
                'criteria': criteriaFrameworkArr,
                "rubric": {
                    "expressionVariables": {
                        "SCORE": "OB.sumOfPointsOfAllChildren()"
                    },
                    "levels": {
                        "L1": {
                            "expression": "(0<=SCORE<=100000)"
                        }
                    }
                }
            }]
            for levs in range(1, criteriaLevelsCount + 1):
                levelToScore = {"L" + str(levs): {'points': levs * 10, 'label': 'Level ' + str(levs)}}
                frameworkDocInsertObj['levelToScoreMapping'].update(levelToScore)
            frameworkDocInsertObj['noOfRatingLevels'] = criteriaLevelsCount
            
        else:
            frameworkDocInsertObj['scoringSystem'] = None
            frameworkDocInsertObj['isRubricDriven'] = False
        
        fetchentitytypeid = Helpers.fetchEntityId(solutionName_for_folder_path, accessToken,
                                                      entitiesPGM.lstrip().rstrip().split(","), scopeEntityType)
    
        frameworkDocInsertObj['entityTypeId'] = fetchentitytypeid
        frameworkDocInsertObj['entityType'] = solutionEntityType
        frameworkDocInsertObj['type'] = 'observation'
        frameworkDocInsertObj['subType'] = solutionEntityType
        frameworkDocInsertObj['status'] = "active"
        frameworkDocInsertObj['updatedBy'] = 'INITIALIZE'
        frameworkDocInsertObj['createdBy'] = 'INITIALIZE'
        frameworkDocInsertObj['createdAt'] = str(dateTime)
        frameworkDocInsertObj['updatedAt'] = str(dateTime)
        frameworkDocInsertObj['author'] = matchedShikshalokamLoginId
        frameworkDocInsertObj['isTempObTest'] = 'observationAutomation'

        # Adding Credits and license into Frameworks
        frameworkDocInsertObj['creator'] = str(creator)
        frameworkDocInsertObj['license'] = {}
        frameworkDocInsertObj['license']['author'] = str(creator)
        frameworkDocInsertObj['license']['creator'] = str(creator)
        frameworkDocInsertObj['license']['copyright'] = str(ccRootOrgName)
        frameworkDocInsertObj['license']['copyrightYear'] = int(dateTime.strftime("%Y"))
        frameworkDocInsertObj['license']['contentType'] = "Observation"
        frameworkDocInsertObj['license']['organisation'] = [ccRootOrgName]
        frameworkDocInsertObj['license']['orgDetails'] = {}
        frameworkDocInsertObj['license']['orgDetails']['email'] = None
        frameworkDocInsertObj['license']['orgDetails']['orgName'] = ccRootOrgName
        frameworkDocInsertObj['license']['licenseDetails'] = {}
        frameworkDocInsertObj['license']['licenseDetails']['name'] = "CC BY 4.0"
        frameworkDocInsertObj['license']['licenseDetails']['url'] = "https://creativecommons.org/licenses/by/4.0/legalcode"
        frameworkDocInsertObj['license']['licenseDetails']['description'] = "For details see below:"
    
        urlCreateFrameworkApi = internal_kong_ip + frameworkcreationapi
        frameworkFilePath = solutionName_for_folder_path + '/framework/'
        file_exists_framework = os.path.isfile(solutionName_for_folder_path + '/framework/uploadFile.json')
        if not os.path.exists(frameworkFilePath):
            os.mkdir(frameworkFilePath)

        with open(frameworkFilePath + "uploadFile.json", "w",encoding='utf-8') as outfile:
            json.dump(frameworkDocInsertObj, outfile)
        headerFrameworkUploadApi = {'Authorization': authorization,
                                    'X-authenticated-user-token': accessToken,
                                    'X-Channel-id': x_channel_id}
        filesFramework = {'framework': open(solutionName_for_folder_path + '/framework/uploadFile.json', 'rb')}

        responseFrameworkUploadApi = requests.post(url=urlCreateFrameworkApi, headers=headerFrameworkUploadApi,
                                                files=filesFramework)
        messageArr = ["Framwork json file created.",
                    "File loc : " + solutionName_for_folder_path + '/framework/uploadFile.json',
                    "Framework upload API called,", "Status code : " + str(responseFrameworkUploadApi.status_code)]
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        if responseFrameworkUploadApi.status_code == 200:
            print('Framework upload Success')
            return frameworkExternalId

        else:
            messageArr = ["Framwork upload Failed.", "Response : " + responseFrameworkUploadApi.text]
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print('Framework upload api failed in ' + environment,
                'status_code response from api is ' + str(responseFrameworkUploadApi.status_code))
            sys.exit()


    def themesUpload(solutionName_for_folder_path, wbObservation, millisAddObs, accessToken, frameworkExternalId,obsWORubWS):
        global dictCritLookUp
        with open(solutionName_for_folder_path + '/criteriaUpload/uploadInternalIdsSheet.csv', 'r',encoding='utf-8') as criteriaInternalFile:
            criteriaInternalReader = csv.DictReader(criteriaInternalFile)
            for crit in criteriaInternalReader:
                dictCritLookUp[crit['Criteria External Id']] = crit['Criteria Internal Id']
        if obsWORubWS:
            print("Themes Observation without rubrics with scores")
            themeUploadFieldnames = ["theme", "aoi", "indicators", "criteriaInternalId"]
            themesUploadCsv = dict()
            for dictCritLookUpKey, dictCritLookUpValue in dictCritLookUp.items():
                themesUploadCsv['theme'] = "Observation Theme" + "###" + "OB" + "###40"
                themesUploadCsv['aoi'] = ""
                themesUploadCsv['indicators'] = ""
                themesUploadCsv['criteriaInternalId'] = dictCritLookUpValue + "###40"
                themeFilePath = solutionName_for_folder_path + '/themeUpload/'
                file_exists = os.path.isfile(solutionName_for_folder_path + '/themeUpload/uploadSheet.csv')

                if not os.path.exists(themeFilePath):
                    os.mkdir(themeFilePath)
                with open(solutionName_for_folder_path + '/themeUpload/uploadSheet.csv', 'a',encoding='utf-8') as themeUploadFile:
                    writerthemeUpload = csv.DictWriter(themeUploadFile, fieldnames=list(themeUploadFieldnames),
                                                    lineterminator='\n')
                    if not file_exists:
                        writerthemeUpload.writeheader()
                    writerthemeUpload.writerow(themesUploadCsv)

        else:
            frameWorkSheet = wbObservation.sheet_by_name('framework')
            keys = [frameWorkSheet.cell(1, col_index).value for col_index in range(frameWorkSheet.ncols)]
            themeUploadFieldnames = ["theme", "aoi", "indicators", "criteriaInternalId"]
            themesUploadCsv = dict()
            for row_index in range(2, frameWorkSheet.nrows):
                dictCriteria = {keys[col_index]: frameWorkSheet.cell(row_index, col_index).value for col_index in
                                range(frameWorkSheet.ncols)}
                themesUploadCsv['theme'] = dictCriteria['Domain Name'].encode('utf-8').decode('utf-8') + "###" + dictCriteria['Domain ID'] + "###40"
                themesUploadCsv['aoi'] = ""
                themesUploadCsv['indicators'] = ""
                themesUploadCsv['criteriaInternalId'] = dictCritLookUp[dictCriteria['Criteria ID'].strip() + '_' + str(
                    millisAddObs)] + "###40"  # if dictCriteria['Criteria ID'] else  ""
                themeFilePath = solutionName_for_folder_path + '/themeUpload/'
                file_exists = os.path.isfile(solutionName_for_folder_path + '/themeUpload/uploadSheet.csv')

                if not os.path.exists(themeFilePath):
                    os.mkdir(themeFilePath)
                with open(solutionName_for_folder_path + '/themeUpload/uploadSheet.csv', 'a',encoding='utf-8') as themeUploadFile:
                    writerthemeUpload = csv.DictWriter(themeUploadFile, fieldnames=list(themeUploadFieldnames),
                                                    lineterminator='\n')
                    if not file_exists:
                        writerthemeUpload.writeheader()
                    writerthemeUpload.writerow(themesUploadCsv)

        urlThemesUploadApi = internal_kong_ip + themeuploadapiurl + frameworkExternalId
        headerThemesUploadApi = {'Authorization': authorization,
                                'X-authenticated-user-token': accessToken,
                                'X-Channel-id': x_channel_id}
        filesThemes = {'themes': open(solutionName_for_folder_path + '/themeUpload/uploadSheet.csv', 'rb')}
        responseThemeUploadApi = requests.post(url=urlThemesUploadApi, headers=headerThemesUploadApi, files=filesThemes)
        messageArr = ["Themes upload sheet prepared.",
                    "File path : " + solutionName_for_folder_path + '/themeUpload/uploadSheet.csv',
                    "Theme upload to framework API called.", "URL : " + urlThemesUploadApi,
                    "Status code : " + str(responseThemeUploadApi.status_code)]
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        if responseThemeUploadApi.status_code == 200:
            print('Theme UploadApi Success')
            with open(solutionName_for_folder_path + '/themeUpload/uploadInternalIdsSheet.csv', 'w+',encoding='utf-8') as criteriaRes:
                criteriaRes.write(responseThemeUploadApi.text)
        else:
            messageArr = ["Themes upload failed.", "Response : " + str(responseThemeUploadApi.text)]
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print("Theme upload failed.")
            sys.exit()

    def createSolutionFromFramework(solutionName_for_folder_path, wbObservation, accessToken, frameworkExternalId):
        urlCreateSolutionApi = internal_kong_ip + solutioncreationapiurl
        headerCreateSolutionApi = {
            'Content-Type': content_type,
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id
        }
        try:
            observationSheet = wbObservation.sheet_by_name("details")  # Reading the "details" sheet
        except xlrd.biffh.XLRDError:
            print("Error: 'details' sheet not found in wbObservation.")
            return None
        
        headers = [observationSheet.cell(1, col_index).value for col_index in range(observationSheet.ncols)]
        values = [observationSheet.cell(2, col_index).value for col_index in range(observationSheet.ncols)]      
        observationData = dict(zip(headers, values))
        entityType = observationData.get('entity_type', 'Default entity_type') 
        queryparamsCreateSolutionApi = '?frameworkId=' + str(frameworkExternalId) + '&entityType=' + entityType
        responseCreateSolutionApi = requests.post(url=urlCreateSolutionApi + queryparamsCreateSolutionApi,
                                                headers=headerCreateSolutionApi)

        messageArr = ["Solution Created from Framework.",
                    "URL : " + str(urlCreateSolutionApi + queryparamsCreateSolutionApi),
                    "Status Code : " + str(responseCreateSolutionApi.status_code),
                    "Response : " + str(responseCreateSolutionApi.text)]
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        messageArr = []
        if responseCreateSolutionApi.status_code == 200:
            responseCreateSolutionApi = responseCreateSolutionApi.json()
            solutionId = responseCreateSolutionApi['result']['templateId']
            messageArr.append("Parent Solution Generated : " + str(solutionId))
            print("Parent Solution Generated : " + str(solutionId))
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        else:
            messageArr.append("Solution from framework api failed.")
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print("Solution from framework api failed.")
            sys.exit()
        return solutionId
    
    def questionUpload(filePathAddObs, solutionName_for_folder_path, frameworkExternalId, millisAddObs, accessToken,
                   solutionId, typeofSolution):
        wbObservation = xlrd.open_workbook(filePathAddObs, on_demand=True)
        excelBook = open_workbook(filePathAddObs)
        sheetNam = excelBook.sheet_names()
        shCnt = 0
        countColSeq = 0
        questShee = wbObservation.sheet_by_name('questions')
        Qukeys = [questShee.cell(1, col_index).value for col_index in range(questShee.ncols)]
        countColSeq = Qukeys.index('question_sequence')
        questionsResponseDict = dict()

        for i in sheetNam:
            if i.strip().lower() == 'questions':
                sheetNam1 = excelBook.sheets()[shCnt]
            shCnt = shCnt + 1
        dataSort = [sheetNam1.row_values(i) for i in range(sheetNam1.nrows)]
        labels = dataSort[1]
        dataSort = dataSort[2:]
        dataSort.sort(key=lambda x: int(x[countColSeq]))
        openWorkBookSort = xlrd.open_workbook(filePathAddObs)
        openWorkBookSort1 = xl_copy(openWorkBookSort)
        sheet1 = openWorkBookSort1.add_sheet('questions_sequence_sorted')
        print("Question Sorted.")
        for idx, label in enumerate(labels):
            sheet1.write(0, idx, label)

        for idx_r, row in enumerate(dataSort):
            for idx_c, value in enumerate(row):
                sheet1.write(idx_r + 1, idx_c, value)

        openWorkBookSort1.save(filePathAddObs)
        wbObservation = xlrd.open_workbook(filePathAddObs, on_demand=True)
        questionsSheet = wbObservation.sheet_by_name('questions_sequence_sorted')
        keys2 = [questionsSheet.cell(0, col_index2).value for col_index2 in range(questionsSheet.ncols)]
        questionsList = list()
        for row_index2 in range(1, questionsSheet.nrows):
            d2 = {keys2[col_index2]: questionsSheet.cell(row_index2, col_index2).value for col_index2 in
                range(questionsSheet.ncols)}
            questionsList.append(d2)
        questionSeqByEcmDict = dict()
        questionSeqByEcmSectionDict = dict()
        questionSeqByEcmArr = []
        quesSeqCnt = 1.0
        questionUploadFieldnames = []
        questionUploadExceptSliderFieldnames = []
        questionUploadSliderFieldNames = []
        if typeofSolution == 1:
            for ques00 in questionsList:
                questionSeqByEcmDict[ecmToSection[ques00['section_id']] + "_" + str(millisAddObs)] = {
                    ecm_sections[ecmToSection[ques00['section_id']] + "_" + str(millisAddObs)]: []}
        elif typeofSolution == 2:
            questionSeqByEcmDict["OB"] = {
                "S1": []
            }

        for ques1 in questionsList:
            if not pointBasedValue.lower() == "null":
                questionUploadExceptSliderFieldnames = ['solutionId', 'criteriaExternalId', 'name', 'evidenceMethod',
                                                        'section', 'instanceParentQuestionId', 'hasAParentQuestion',
                                                        'parentQuestionOperator', 'parentQuestionValue', 'parentQuestionId',
                                                        'externalId', 'question0', 'question1', 'tip', 'hint',
                                                        'instanceIdentifier', 'responseType', 'dateFormat', 'autoCapture',
                                                        'validation', 'validationIsNumber', 'validationRegex',
                                                        'validationMax', 'validationMin', 'file', 'fileIsRequired',
                                                        'fileUploadType', 'allowAudioRecording', 'minFileCount',
                                                        'maxFileCount', 'caption', 'questionGroup', 'modeOfCollection',
                                                        'accessibility', 'showRemarks', 'rubricLevel', 'isAGeneralQuestion',
                                                        'R1', 'R1-hint', 'R2', 'R2-hint', 'R3', 'R3-hint', 'R4', 'R4-hint',
                                                        'R5', 'R5-hint', 'R6', 'R6-hint', 'R7', 'R7-hint', 'R8', 'R8-hint',
                                                        'R9', 'R9-hint', 'R10', 'R10-hint', 'R11', 'R11-hint', 'R12',
                                                        'R12-hint', 'R13', 'R13-hint', 'R14', 'R14-hint', 'R15', 'R15-hint',
                                                        'R16', 'R16-hint', 'R17', 'R17-hint', 'R18', 'R18-hint', 'R19',
                                                        'R19-hint', 'R20', 'R20-hint', 'R1-score', 'R2-score', 'R3-score',
                                                        'R4-score', 'R5-score', 'R6-score', 'R7-score', 'R8-score',
                                                        'R9-score', 'R10-score', 'R11-score', 'R12-score', 'R13-score',
                                                        'R14-score', 'R15-score', 'R16-score', 'R17-score', 'R18-score',
                                                        'R19-score', 'R20-score', 'weightage', 'sectionHeader', 'page',
                                                        'questionNumber', '_arrayFields', 'prefillFromEntityProfile',
                                                        'isEditable', 'entityFieldName']
                if ques1['question_response_type'].strip().lower() == 'slider' and ques1['slider_value_with_score'].strip():
                    noOfSliderColumn = ques1['slider_value_with_score'].strip().split(',')
                    possibleSliderColumn = (int(ques1['max_number_value']) + 1) - (int(ques1['min_number_value']))
                    sliderCnt = int(ques1['min_number_value'])
                    if len(noOfSliderColumn) == possibleSliderColumn:
                        for sliderIndex, sliCn in enumerate(noOfSliderColumn):
                            questionUploadSliderFieldNames.append('slider-value-' + str(sliderIndex + 1))
                            questionUploadSliderFieldNames.append('slider-value-' + str(sliderIndex + 1) + '-score')
            else:
                questionUploadFieldnames = ['solutionId', 'criteriaExternalId', 'name', 'evidenceMethod', 'section',
                                            'instanceParentQuestionId', 'hasAParentQuestion', 'parentQuestionOperator',
                                            'parentQuestionValue', 'parentQuestionId', 'externalId', 'question0',
                                            'question1', 'tip', 'hint', 'instanceIdentifier', 'responseType', 'dateFormat',
                                            'autoCapture', 'validation', 'validationIsNumber', 'validationRegex',
                                            'validationMax', 'validationMin', 'file', 'fileIsRequired', 'fileUploadType',
                                            'allowAudioRecording', 'minFileCount', 'maxFileCount', 'caption',
                                            'questionGroup', 'modeOfCollection', 'accessibility', 'showRemarks',
                                            'rubricLevel', 'isAGeneralQuestion', 'R1', 'R1-hint', 'R2', 'R2-hint', 'R3',
                                            'R3-hint', 'R4', 'R4-hint', 'R5', 'R5-hint', 'R6', 'R6-hint', 'R7', 'R7-hint',
                                            'R8', 'R8-hint', 'R9', 'R9-hint', 'R10', 'R10-hint', 'R11', 'R11-hint', 'R12',
                                            'R12-hint', 'R13', 'R13-hint', 'R14', 'R14-hint', 'R15', 'R15-hint', 'R16',
                                            'R16-hint', 'R17', 'R17-hint', 'R18', 'R18-hint', 'R19', 'R19-hint', 'R20',
                                            'R20-hint', 'sectionHeader', 'page', 'questionNumber', '_arrayFields',
                                            'prefillFromEntityProfile', 'isEditable', 'entityFieldName']
        if len(questionUploadExceptSliderFieldnames) > 0:
            if len(questionUploadSliderFieldNames) > 0:
                questionUploadFieldnames = questionUploadExceptSliderFieldnames + questionUploadSliderFieldNames
            else:
                questionUploadFieldnames = questionUploadExceptSliderFieldnames
        for ques in questionsList:
            questionFilePath = solutionName_for_folder_path + '/questionUpload/'
            file_exists_ques = os.path.isfile(solutionName_for_folder_path + '/questionUpload/uploadSheet.csv')
            if not os.path.exists(questionFilePath):
                os.mkdir(questionFilePath)
            with open(solutionName_for_folder_path + '/questionUpload/uploadSheet.csv', 'a',
                    encoding='utf-8') as questionUploadFile:
                writerQuestionUpload = csv.DictWriter(questionUploadFile, fieldnames=questionUploadFieldnames,
                                                    lineterminator='\n')
                if not file_exists_ques:
                    writerQuestionUpload.writeheader()
                questionFileObj = {}
                observationExternalId = None
                observationExternalId = frameworkExternalId + "-OBSERVATION-TEMPLATE"
                questionFileObj['solutionId'] = observationExternalId
                questionFileObj['criteriaExternalId'] = ques['criteria_id'].strip() + '_' + str(millisAddObs)
                try:
                    questionFileObj['name'] = criteriaLookUp[questionFileObj['criteriaExternalId']]
                except:
                    print("criteria Id error....")
                    print(questionFileObj['criteriaExternalId'] + " not found.")
                    sys.exit()
                if typeofSolution == 1 or typeofSolution == 5:
                    questionFileObj['evidenceMethod'] = ecmToSection[ques['section_id']] + "_" + str(millisAddObs)
                    questionFileObj['section'] = ques['section_id']
                elif typeofSolution == 2:
                    questionFileObj['evidenceMethod'] = "OB"
                    questionFileObj['section'] = "S1"
                questionsResponseDict[ques['question_id'].strip() + '_' + str(millisAddObs)] = {
                    "response(R1)": ques["response(R1)".replace(" ", "")],
                    "response(R2)": ques["response(R2)".replace(" ", "")],
                    "response(R3)": ques["response(R3)".replace(" ", "")],
                    "response(R4)": ques["response(R4)".replace(" ", "")],
                    "response(R5)": ques["response(R5)".replace(" ", "")],
                    "response(R6)": ques["response(R6)".replace(" ", "")],
                    "response(R7)": ques["response(R7)".replace(" ", "")],
                    "response(R8)": ques["response(R8)".replace(" ", "")],
                    "response(R9)": ques["response(R9)".replace(" ", "")],
                    "response(R10)": ques["response(R10)".replace(" ", "")],
                    "response(R11)": ques["response(R11)".replace(" ", "")],
                    "response(R12)": ques["response(R12)".replace(" ", "")],
                    "response(R13)": ques["response(R13)".replace(" ", "")],
                    "response(R14)": ques["response(R14)".replace(" ", "")],
                    "response(R15)": ques["response(R15)".replace(" ", "")],
                    "response(R16)": ques["response(R16)".replace(" ", "")],
                    "response(R17)": ques["response(R17)".replace(" ", "")],
                    "response(R18)": ques["response(R18)".replace(" ", "")],
                    "response(R19)": ques["response(R19)".replace(" ", "")],
                    "response(R20)": ques["response(R20)".replace(" ", "")]}
                hasInstanceParentFlag = False
                if ques['instance_parent_question_id'].encode('utf-8').decode('utf-8'):
                    hasInstanceParentFlag = True
                    questionFileObj['instanceParentQuestionId'] = ques['instance_parent_question_id'].encode('utf-8').decode('utf-8').strip() + '_' + str(
                        millisAddObs)
                    questionFileObj['hasAParentQuestion'] = 'NO'
                else:
                    hasInstanceParentFlag = False
                    questionFileObj['instanceParentQuestionId'] = 'NA'
                notEqualsFlag = False
                if ques['parent_question_id'].encode('utf-8').decode('utf-8').strip():
                    questionFileObj['hasAParentQuestion'] = 'YES'
                    if ques['show_when_parent_question_value_is'].encode('utf-8').decode('utf-8').lower().lstrip().rstrip() == 'or' or ques[
                        'show_when_parent_question_value_is'].encode('utf-8').decode('utf-8').lower().lstrip().rstrip() == '||':
                        notEqualsFlag = False
                        questionFileObj['parentQuestionOperator'] = '||'
                        questionFileObj['parentQuestionValue'] = ques['parent_question_value'].encode('utf-8').decode('utf-8').lstrip().rstrip().replace(
                            " ", "")
                    elif ques['show_when_parent_question_value_is'].lower().lstrip().rstrip() == 'equals':
                        notEqualsFlag = False
                        questionFileObj['parentQuestionOperator'] = "EQUALS"
                        questionFileObj['parentQuestionValue'] = ques['parent_question_value'].encode('utf-8').decode('utf-8').lstrip().rstrip().replace(
                            " ", "")
                    elif ques['show_when_parent_question_value_is'].encode('utf-8').decode('utf-8').lstrip().rstrip() == 'NOT_EQUALS_TO' or ques[
                        'show_when_parent_question_value_is'].encode('utf-8').decode('utf-8').lower().lstrip().rstrip() == 'NOT_EQUALS_TO'.lower():
                        notEqualsFlag = True
                        questionFileObj['parentQuestionOperator'] = "||"
                    else:
                        questionFileObj['parentQuestionOperator'] = ""
                    if type(ques['parent_question_value']) != str:
                        if (ques['parent_question_value'] and ques['parent_question_value'].is_integer() == True):
                            questionFileObj['parentQuestionValue'] = int(ques['parent_question_value'])
                        elif (ques['parent_question_value'] and ques['parent_question_value'].is_integer() == False):
                            questionFileObj['parentQuestionValue'] = ques[
                                'parent_question_value'].encode('utf-8').decode('utf-8').lstrip().rstrip().replace(" ", "")
                    else:
                        questionFileObj['parentQuestionId'] = ques['parent_question_id'].encode('utf-8').decode('utf-8').strip() + '_' + str(millisAddObs)
                        if notEqualsFlag:
                            Qkeys = ques.keys()
                            final_parent_question_value = str()
                            avoidResponses = ques['parent_question_value'].lstrip().rstrip().split(",")
                            for i in Qkeys:
                                searchResponse = re.search("^response\(R[0-9]\)$|^response\(R[0-2][0-9]\)$", i)
                                if searchResponse:
                                    try:
                                        responseCheck = questionsResponseDict[questionFileObj['parentQuestionId']][
                                            searchResponse.string]
                                    except:
                                        print(questionFileObj[
                                                'parentQuestionId'] + " Referenced before intialising in questions sheet.")
                                        print("Please check question sequence...")
                                        print("Aborting...")
                                        messageArr = [questionFileObj[
                                                        'parentQuestionId'] + " Referenced before intialising in questions sheet.",
                                                    "Please check question sequesnce...", ]
                                        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
                                        sys.exit()
                                    if responseCheck:
                                        if not searchResponse.string.replace("response(", "").replace(")",
                                                                                                    "") in avoidResponses:
                                            final_parent_question_value += searchResponse.string.replace("response(",
                                                                                                        "").replace(")",
                                                                                                                    "") + ","
                            questionFileObj['parentQuestionValue'] = final_parent_question_value.encode('utf-8').decode('utf-8').rstrip(",").lstrip(",")
                        else:
                            pass
                else:
                    questionFileObj['parentQuestionOperator'] = None
                    questionFileObj['parentQuestionValue'] = None
                    questionFileObj['parentQuestionId'] = None
                questionFileObj['externalId'] = ques['question_id'].strip() + '_' + str(millisAddObs)
                if typeofSolution == 1:
                    questionSeqByEcmDict[questionFileObj['evidenceMethod']][
                        ecm_sections[questionFileObj['evidenceMethod']]].append(
                        ques['question_id'].strip() + '_' + str(millisAddObs))
                elif typeofSolution == 2:
                    questionSeqByEcmDict["OB"]["S1"].append(ques['question_id'].strip() + '_' + str(millisAddObs))

                questionFileObj['question0'] = ques['question_primary_language'].encode('utf-8').decode('utf-8')
                if not questionFileObj['question0']:
                    questionFileObj['question0'] = None
                if ques['question_secondory_language']:
                    questionFileObj['question1'] = ques['question_secondory_language'].encode('utf-8').decode('utf-8')
                else:
                    questionFileObj['question1'] = None
                if ques['question_tip']:
                    questionFileObj['tip'] = ques['question_tip'].encode('utf-8').decode('utf-8')
                else:
                    questionFileObj['tip'] = None
                if ques['question_hint']:
                    questionFileObj['hint'] = ques['question_hint'].encode('utf-8').decode('utf-8')
                else:
                    questionFileObj['hint'] = None
                if ques['instance_identifier']:
                    questionFileObj['instanceIdentifier'] = ques['instance_identifier'].encode('utf-8').decode('utf-8')
                else:
                    questionFileObj['instanceIdentifier'] = None
                if ques['question_response_type'].strip().lower():
                    questionFileObj['responseType'] = ques['question_response_type'].strip().lower()
                if questionFileObj['responseType'] == "date":
                    questionFileObj['dateFormat'] = "DD-MM-YYYY"
                    if ques['date_auto_capture'] and ques['date_auto_capture'] == 1 or str(
                            ques['date_auto_capture']).lower() == "true":
                        questionFileObj['autoCapture'] = 'TRUE'
                    elif ques['date_auto_capture'] and ques['date_auto_capture'] == 0 or str(
                            ques['date_auto_capture']).lower() == "false":
                        questionFileObj['autoCapture'] = 'FALSE'
                    else:
                        questionFileObj['autoCapture'] = 'FALSE'

                else:
                    questionFileObj['dateFormat'] = ""
                    questionFileObj['autoCapture'] = None
                if ques['response_required']:
                    if ques['response_required'] == 1 or str(ques['response_required']).lower() == "true":
                        questionFileObj['validation'] = 'TRUE'
                    else:
                        questionFileObj['validation'] = 'FALSE'
                else:
                    questionFileObj['validation'] = 'FALSE'
                if ques['question_response_type'].strip().lower() == 'number':
                    questionFileObj['validationIsNumber'] = 'TRUE'
                    questionFileObj['validationRegex'] = 'isNumber'
                    if (ques['max_number_value'] and ques['max_number_value'].is_integer() == True):
                        questionFileObj['validationMax'] = int(ques['max_number_value'])
                    elif (ques['max_number_value'] and ques['max_number_value'].is_integer() == False):
                        questionFileObj['validationMax'] = ques['max_number_value']
                    else:
                        questionFileObj['validationMax'] = 10000
                    if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                        questionFileObj['validationMin'] = int(ques['min_number_value'])
                    elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                        questionFileObj['validationMin'] = ques['min_number_value']
                    else:
                        questionFileObj['validationMin'] = 0
                elif ques['question_response_type'].strip().lower() == 'slider':
                    questionFileObj['validationIsNumber'] = None
                    questionFileObj['validationRegex'] = 'isNumber'
                    if (ques['max_number_value'] and ques['max_number_value'].is_integer() == True):
                        questionFileObj['validationMax'] = int(ques['max_number_value'])
                    elif (ques['max_number_value'] and ques['max_number_value'].is_integer() == False):
                        questionFileObj['validationMax'] = ques['max_number_value']
                    else:
                        questionFileObj['validationMax'] = 5
                    if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                        questionFileObj['validationMin'] = int(ques['min_number_value'])
                    elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                        questionFileObj['validationMin'] = ques['min_number_value']
                    else:
                        questionFileObj['validationMin'] = 0
                else:
                    questionFileObj['validationIsNumber'] = None
                    questionFileObj['validationRegex'] = None
                    questionFileObj['validationMax'] = None
                    questionFileObj['validationMin'] = None
                if ques['file_upload'] == 1 or ques['file_upload'] == "TRUE":
                    questionFileObj['file'] = 'Snapshot'
                    questionFileObj['fileIsRequired'] = 'TRUE'
                    questionFileObj['fileUploadType'] = 'image/jpeg,docx,pdf,ppt'
                    questionFileObj['minFileCount'] = 0
                    questionFileObj['maxFileCount'] = 10
                else:
                    questionFileObj['file'] = 'NA'
                    questionFileObj['fileIsRequired'] = "FALSE"
                    questionFileObj['fileUploadType'] = None
                    questionFileObj['minFileCount'] = None
                    questionFileObj['maxFileCount'] = None
                questionFileObj['allowAudioRecording'] = False
                questionFileObj['caption'] = 'FALSE'
                questionFileObj['questionGroup'] = 'A1'
                questionFileObj['modeOfCollection'] = 'onfield'
                questionFileObj['accessibility'] = 'No'
                if ques['show_remarks'] == 1 or ques['show_remarks'] == "TRUE":
                    questionFileObj['showRemarks'] = 'TRUE'
                else:
                    questionFileObj['showRemarks'] = 'FALSE'
                questionFileObj['rubricLevel'] = None
                questionFileObj['isAGeneralQuestion'] = None
                if not pointBasedValue.lower() == "null":
                    if ques['question_response_type'].strip().lower() == 'radio' or ques[
                        'question_response_type'].strip() == 'multiselect':
                        questionFileObj['R1-score'] = ques['Score for R1']
                        questionFileObj['R2-score'] = ques['Score for R2']
                        questionFileObj['R3-score'] = ques['Score for R3']
                        questionFileObj['R4-score'] = ques['Score for R4']
                        questionFileObj['R5-score'] = ques['Score for R5']
                        questionFileObj['R6-score'] = ques['Score for R6']
                        questionFileObj['R7-score'] = ques['Score for R7']
                        questionFileObj['R8-score'] = ques['Score for R8']
                        questionFileObj['R9-score'] = ques['Score for R9']
                        questionFileObj['R10-score'] = ques['Score for R10']
                        questionFileObj['R11-score'] = ques['Score for R11']
                        questionFileObj['R12-score'] = ques['Score for R12']
                        questionFileObj['R13-score'] = ques['Score for R13']
                        questionFileObj['R14-score'] = ques['Score for R14']
                        questionFileObj['R15-score'] = ques['Score for R15']
                        questionFileObj['R16-score'] = ques['Score for R16']
                        questionFileObj['R17-score'] = ques['Score for R17']
                        questionFileObj['R18-score'] = ques['Score for R18']
                        questionFileObj['R19-score'] = ques['Score for R19']
                        questionFileObj['R20-score'] = ques['Score for R20']
                    if ques['question_response_type'].strip().lower() == 'slider' and ques[
                        'slider_value_with_score'].strip():
                        noOfSliderColumnQuestionVal = ques['slider_value_with_score'].strip().split(',')
                        possibleSliderColumnQuesVal = (int(ques['max_number_value']) + 1) - (int(ques['min_number_value']))
                        if len(noOfSliderColumnQuestionVal) == possibleSliderColumnQuesVal:
                            for sliVal in noOfSliderColumnQuestionVal:
                                sliValArr = []
                                sliValArr = sliVal.split(':')
                                questionFileObj['slider-value-' + str(sliValArr[0])] = sliValArr[0]
                                questionFileObj['slider-value-' + str(sliValArr[0]) + '-score'] = sliValArr[1]
                    if str(ques['question_weightage']):
                        questionFileObj['weightage'] = ques['question_weightage']
                    else:
                        questionFileObj['weightage'] = 0
                if ques['question_response_type'].strip().lower() == 'radio' or ques[
                    'question_response_type'].strip() == 'multiselect':
                    if type(ques['response(R1)']) != str:
                        if (ques['response(R1)'] and ques['response(R1)'].is_integer() == True):
                            questionFileObj['R1'] = int(ques['response(R1)'])
                        elif (ques['response(R1)'] and ques['response(R1)'].is_integer() == False):
                            questionFileObj['R1'] = ques['response(R1)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R1'] = ques['response(R1)']
                    if type(ques['response(R1)_hint']) != str:
                        if (ques['response(R1)_hint'] and ques['response(R1)_hint'].is_integer() == True):
                            questionFileObj['R1-hint'] = int(ques['response(R1)_hint'])
                        elif (ques['response(R1)_hint'] and ques['response(R1)_hint'].is_integer() == False):
                            questionFileObj['R1-hint'] = ques['response(R1)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R1-hint'] = ques['response(R1)_hint'].encode('utf-8').decode('utf-8')
                    if type(ques['response(R2)']) != str:
                        if (ques['response(R2)'] and ques['response(R2)'].is_integer() == True):
                            questionFileObj['R2'] = int(ques['response(R2)'])
                        elif (ques['response(R2)'] and ques['response(R2)'].is_integer() == False):
                            questionFileObj['R2'] = ques['response(R2)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R2'] = ques['response(R2)']
                    if type(ques['response(R2)_hint']) != str:
                        if (ques['response(R2)_hint'] and ques['response(R2)_hint'].is_integer() == True):
                            questionFileObj['R2-hint'] = int(ques['response(R2)_hint'])
                        elif (ques['response(R2)_hint'] and ques['response(R2)_hint'].is_integer() == False):
                            questionFileObj['R2-hint'] = ques['response(R2)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R2-hint'] = ques['response(R2)_hint']
                    if type(ques['response(R3)']) != str:
                        if (ques['response(R3)'] and ques['response(R3)'].is_integer() == True):
                            questionFileObj['R3'] = int(ques['response(R3)'])
                        elif (ques['response(R3)'] and ques['response(R3)'].is_integer() == False):
                            questionFileObj['R3'] = ques['response(R3)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R3'] = ques['response(R3)']
                    if type(ques['response(R3)_hint']) != str:
                        if (ques['response(R3)_hint'] and ques['response(R3)_hint'].is_integer() == True):
                            questionFileObj['R3-hint'] = int(ques['response(R3)_hint'])
                        elif (ques['response(R3)_hint'] and ques['response(R3)_hint'].is_integer() == False):
                            questionFileObj['R3-hint'] = ques['response(R3)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R3-hint'] = ques['response(R3)_hint']
                    if type(ques['response(R4)']) != str:
                        if (ques['response(R4)'] and ques['response(R4)'].is_integer() == True):
                            questionFileObj['R4'] = int(ques['response(R4)'])
                        elif (ques['response(R4)'] and ques['response(R4)'].is_integer() == False):
                            questionFileObj['R4'] = ques['response(R4)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R4'] = ques['response(R4)']
                    if type(ques['response(R4)_hint']) != str:
                        if (ques['response(R4)_hint'] and ques['response(R4)_hint'].is_integer() == True):
                            questionFileObj['R4-hint'] = int(ques['response(R4)_hint'])
                        elif (ques['response(R4)_hint'] and ques['response(R4)_hint'].is_integer() == False):
                            questionFileObj['R4-hint'] = ques['response(R4)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R4-hint'] = ques['response(R4)_hint']
                    if type(ques['response(R5)']) != str:
                        if (ques['response(R5)'] and ques['response(R5)'].is_integer() == True):
                            questionFileObj['R5'] = int(ques['response(R5)'])
                        elif (ques['response(R5)'] and ques['response(R5)'].is_integer() == False):
                            questionFileObj['R5'] = ques['response(R5)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R5'] = ques['response(R5)']
                    if type(ques['response(R5)_hint']) != str:
                        if (ques['response(R5)_hint'] and ques['response(R5)_hint'].is_integer() == True):
                            questionFileObj['R5-hint'] = int(ques['response(R5)_hint'])
                        elif (ques['response(R5)_hint'] and ques['response(R5)_hint'].is_integer() == False):
                            questionFileObj['R5-hint'] = ques['response(R5)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R5-hint'] = ques['response(R5)_hint']
                    if type(ques['response(R6)']) != str:
                        if (ques['response(R6)'] and ques['response(R6)'].is_integer() == True):
                            questionFileObj['R6'] = int(ques['response(R6)'])
                        elif (ques['response(R6)'] and ques['response(R6)'].is_integer() == False):
                            questionFileObj['R6'] = ques['response(R6)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R6'] = ques['response(R6)']
                    if type(ques['response(R6)_hint']) != str:
                        if (ques['response(R6)_hint'] and ques['response(R6)_hint'].is_integer() == True):
                            questionFileObj['R6-hint'] = int(ques['response(R6)_hint'])
                        elif (ques['response(R6)_hint'] and ques['response(R6)_hint'].is_integer() == False):
                            questionFileObj['R6-hint'] = ques['response(R6)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R6-hint'] = ques['response(R6)_hint']
                    if type(ques['response(R7)']) != str:
                        if (ques['response(R7)'] and ques['response(R7)'].is_integer() == True):
                            questionFileObj['R7'] = int(ques['response(R7)'])
                        elif (ques['response(R7)'] and ques['response(R7)'].is_integer() == False):
                            questionFileObj['R7'] = ques['response(R7)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R7'] = ques['response(R7)']
                    if type(ques['response(R7)_hint']) != str:
                        if (ques['response(R7)_hint'] and ques['response(R7)_hint'].is_integer() == True):
                            questionFileObj['R7-hint'] = int(ques['response(R7)_hint'])
                        elif (ques['response(R7)_hint'] and ques['response(R7)_hint'].is_integer() == False):
                            questionFileObj['R7-hint'] = ques['response(R7)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R7-hint'] = ques['response(R7)_hint']
                    if type(ques['response(R8)']) != str:
                        if (ques['response(R8)'] and ques['response(R8)'].is_integer() == True):
                            questionFileObj['R8'] = int(ques['response(R8)'])
                        elif (ques['response(R8)'] and ques['response(R8)'].is_integer() == False):
                            questionFileObj['R8'] = ques['response(R8)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R8'] = ques['response(R8)']
                    if type(ques['response(R8)_hint']) != str:
                        if (ques['response(R8)_hint'] and ques['response(R8)_hint'].is_integer() == True):
                            questionFileObj['R8-hint'] = int(ques['response(R8)_hint'])
                        elif (ques['response(R8)_hint'] and ques['response(R8)_hint'].is_integer() == False):
                            questionFileObj['R8-hint'] = ques['response(R8)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R8-hint'] = ques['response(R8)_hint']
                    if type(ques['response(R9)']) != str:
                        if (ques['response(R9)'] and ques['response(R9)'].is_integer() == True):
                            questionFileObj['R9'] = int(ques['response(R9)'])
                        elif (ques['response(R9)'] and ques['response(R9)'].is_integer() == False):
                            questionFileObj['R9'] = ques['response(R9)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R9'] = ques['response(R9)']
                    if type(ques['response(R9)_hint']) != str:
                        if (ques['response(R9)_hint'] and ques['response(R9)_hint'].is_integer() == True):
                            questionFileObj['R9-hint'] = int(ques['response(R9)_hint'])
                        elif (ques['response(R9)_hint'] and ques['response(R9)_hint'].is_integer() == False):
                            questionFileObj['R9-hint'] = ques['response(R9)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R9-hint'] = ques['response(R9)_hint']
                    if type(ques['response(R10)']) != str:
                        if (ques['response(R10)'] and ques['response(R10)'].is_integer() == True):
                            questionFileObj['R10'] = int(ques['response(R10)'])
                        elif (ques['response(R10)'] and ques['response(R10)'].is_integer() == False):
                            questionFileObj['R10'] = ques['response(R10)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R10'] = ques['response(R10)']
                    if type(ques['response(R10)_hint']) != str:
                        if (ques['response(R10)_hint'] and ques['response(R10)_hint'].is_integer() == True):
                            questionFileObj['R10-hint'] = int(ques['response(R10)_hint'])
                        elif (ques['response(R10)_hint'] and ques['response(R10)_hint'].is_integer() == False):
                            questionFileObj['R10-hint'] = ques['response(R10)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R10-hint'] = ques['response(R10)_hint']
                    if type(ques['response(R11)']) != str:
                        if (ques['response(R11)'] and ques['response(R11)'].is_integer() == True):
                            questionFileObj['R11'] = int(ques['response(R11)'])
                        elif (ques['response(R11)'] and ques['response(R11)'].is_integer() == False):
                            questionFileObj['R11'] = ques['response(R11)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R11'] = ques['response(R11)']
                    if type(ques['response(R11)_hint']) != str:
                        if (ques['response(R11)_hint'] and ques['response(R11)_hint'].is_integer() == True):
                            questionFileObj['R11-hint'] = int(ques['response(R11)_hint'])
                        elif (ques['response(R11)_hint'] and ques['response(R11)_hint'].is_integer() == False):
                            questionFileObj['R11-hint'] = ques['response(R11)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R11-hint'] = ques['response(R11)_hint']
                    if type(ques['response(R12)']) != str:
                        if (ques['response(R12)'] and ques['response(R12)'].is_integer() == True):
                            questionFileObj['R12'] = int(ques['response(R12)'])
                        elif (ques['response(R12)'] and ques['response(R12)'].is_integer() == False):
                            questionFileObj['R12'] = ques['response(R12)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R12'] = ques['response(R12)']
                    if type(ques['response(R12)_hint']) != str:
                        if (ques['response(R12)_hint'] and ques['response(R12)_hint'].is_integer() == True):
                            questionFileObj['R12-hint'] = int(ques['response(R12)_hint'])
                        elif (ques['response(R12)_hint'] and ques['response(R12)_hint'].is_integer() == False):
                            questionFileObj['R12-hint'] = ques['response(R12)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R12-hint'] = ques['response(R12)_hint']
                    if type(ques['response(R13)']) != str:
                        if (ques['response(R13)'] and ques['response(R13)'].is_integer() == True):
                            questionFileObj['R13'] = int(ques['response(R13)'])
                        elif (ques['response(R13)'] and ques['response(R13)'].is_integer() == False):
                            questionFileObj['R13'] = ques['response(R13)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R13'] = ques['response(R13)']
                    if type(ques['response(R13)_hint']) != str:
                        if (ques['response(R13)_hint'] and ques['response(R13)_hint'].is_integer() == True):
                            questionFileObj['R13-hint'] = int(ques['response(R13)_hint'])
                        elif (ques['response(R13)_hint'] and ques['response(R13)_hint'].is_integer() == False):
                            questionFileObj['R13-hint'] = ques['response(R13)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R13-hint'] = ques['response(R13)_hint']
                    if type(ques['response(R14)']) != str:
                        if (ques['response(R14)'] and ques['response(R14)'].is_integer() == True):
                            questionFileObj['R14'] = int(ques['response(R14)'])
                        elif (ques['response(R14)'] and ques['response(R14)'].is_integer() == False):
                            questionFileObj['R14'] = ques['response(R14)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R14'] = ques['response(R14)']
                    if type(ques['response(R14)_hint']) != str:
                        if (ques['response(R14)_hint'] and ques['response(R14)_hint'].is_integer() == True):
                            questionFileObj['R14-hint'] = int(ques['response(R14)_hint'])
                        elif (ques['response(R14)_hint'] and ques['response(R14)_hint'].is_integer() == False):
                            questionFileObj['R14-hint'] = ques['response(R14)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R14-hint'] = ques['response(R14)_hint']
                    if type(ques['response(R15)']) != str:
                        if (ques['response(R15)'] and ques['response(R15)'].is_integer() == True):
                            questionFileObj['R15'] = int(ques['response(R15)'])
                        elif (ques['response(R15)'] and ques['response(R15)'].is_integer() == False):
                            questionFileObj['R15'] = ques['response(R15)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R15'] = ques['response(R15)']
                    if type(ques['response(R15)_hint']) != str:
                        if (ques['response(R15)_hint'] and ques['response(R15)_hint'].is_integer() == True):
                            questionFileObj['R15-hint'] = int(ques['response(R15)_hint'])
                        elif (ques['response(R15)_hint'] and ques['response(R15)_hint'].is_integer() == False):
                            questionFileObj['R15-hint'] = ques['response(R15)_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R15-hint'] = ques['response(R15)_hint']
                    if type(ques['response(R16)']) != str:
                        if (ques['response(R16)'] and ques['response(R16)'].is_integer() == True):
                            questionFileObj['R16'] = int(ques['response(R16)'])
                        elif (ques['response(R16)'] and ques['response(R16)'].is_integer() == False):
                            questionFileObj['R16'] = ques['response(R16)'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['R16'] = ques['response(R16)']
                    if type(ques['response(R16)_hint']) != str:
                        if (ques['response(R16)_hint'] and ques['response(R16)_hint'].is_integer() == True):
                            questionFileObj['R16-hint'] = int(ques['response(R16)_hint'])
                        elif (ques['response(R16)_hint'] and ques['response(R16)_hint'].is_integer() == False):
                            questionFileObj['R16-hint'] = ques['response(R16)_hint']
                    else:
                        questionFileObj['R16-hint'] = ques['response(R16)_hint']
                    if type(ques['response(R17)']) != str:
                        if (ques['response(R17)'] and ques['response(R17)'].is_integer() == True):
                            questionFileObj['R17'] = int(ques['response(R17)'])
                        elif (ques['response(R17)'] and ques['response(R17)'].is_integer() == False):
                            questionFileObj['R17'] = ques['response(R17)']
                    else:
                        questionFileObj['R17'] = ques['response(R17)']
                    if type(ques['response(R17)_hint']) != str:
                        if (ques['response(R17)_hint'] and ques['response(R17)_hint'].is_integer() == True):
                            questionFileObj['R17-hint'] = int(ques['response(R17)_hint'])
                        elif (ques['response(R17)_hint'] and ques['response(R17)_hint'].is_integer() == False):
                            questionFileObj['R17-hint'] = ques['response(R17)_hint']
                    else:
                        questionFileObj['R17-hint'] = ques['response(R17)_hint']
                    if type(ques['response(R18)']) != str:
                        if (ques['response(R18)'] and ques['response(R18)'].is_integer() == True):
                            questionFileObj['R18'] = int(ques['response(R18)'])
                        elif (ques['response(R18)'] and ques['response(R18)'].is_integer() == False):
                            questionFileObj['R18'] = ques['response(R18)']
                    else:
                        questionFileObj['R18'] = ques['response(R18)']
                    if type(ques['response(R18)_hint']) != str:
                        if (ques['response(R18)_hint'] and ques['response(R18)_hint'].is_integer() == True):
                            questionFileObj['R18-hint'] = int(ques['response(R18)_hint'])
                        elif (ques['response(R18)_hint'] and ques['response(R18)_hint'].is_integer() == False):
                            questionFileObj['R18-hint'] = ques['response(R18)_hint']
                    else:
                        questionFileObj['R18-hint'] = ques['response(R18)_hint']
                    if type(ques['response(R19)']) != str:
                        if (ques['response(R19)'] and ques['response(R19)'].is_integer() == True):
                            questionFileObj['R19'] = int(ques['response(R19)'])
                        elif (ques['response(R19)'] and ques['response(R19)'].is_integer() == False):
                            questionFileObj['R19'] = ques['response(R19)']
                    else:
                        questionFileObj['R19'] = ques['response(R19)']
                    if type(ques['response(R19)_hint']) != str:
                        if (ques['response(R19)_hint'] and ques['response(R19)_hint'].is_integer() == True):
                            questionFileObj['R19-hint'] = int(ques['response(R19)_hint'])
                        elif (ques['response(R19)_hint'] and ques['response(R19)_hint'].is_integer() == False):
                            questionFileObj['R19-hint'] = ques['response(R19)_hint']
                    else:
                        questionFileObj['R19-hint'] = ques['response(R19)_hint']
                    if type(ques['response(R20)']) != str:
                        if (ques['response(R20)'] and ques['response(R20)'].is_integer() == True):
                            questionFileObj['R20'] = int(ques['response(R20)'])
                        elif (ques['response(R20)'] and ques['response(R20)'].is_integer() == False):
                            questionFileObj['R20'] = ques['response(R20)']
                    else:
                        questionFileObj['R20'] = ques['response(R20)']
                    if type(ques['response(R20)_hint']) != str:
                        if (ques['response(R20)_hint'] and ques['response(R20)_hint'].is_integer() == True):
                            questionFileObj['R20-hint'] = int(ques['response(R20)_hint'])
                        elif (ques['response(R20)_hint'] and ques['response(R20)_hint'].is_integer() == False):
                            questionFileObj['R20-hint'] = ques['response(R20)_hint']
                    else:
                        questionFileObj['R20-hint'] = ques['response(R20)_hint']
                else:
                    questionFileObj['R1'] = None
                    questionFileObj['R1-hint'] = None
                    questionFileObj['R2'] = None
                    questionFileObj['R2-hint'] = None
                    questionFileObj['R3'] = None
                    questionFileObj['R3-hint'] = None
                    questionFileObj['R4'] = None
                    questionFileObj['R4-hint'] = None
                    questionFileObj['R5'] = None
                    questionFileObj['R5-hint'] = None
                    questionFileObj['R6'] = None
                    questionFileObj['R6-hint'] = None
                    questionFileObj['R7'] = None
                    questionFileObj['R7-hint'] = None
                    questionFileObj['R8'] = None
                    questionFileObj['R8-hint'] = None
                    questionFileObj['R9'] = None
                    questionFileObj['R9-hint'] = None
                    questionFileObj['R10'] = None
                    questionFileObj['R10-hint'] = None
                    questionFileObj['R11'] = None
                    questionFileObj['R11-hint'] = None
                    questionFileObj['R12'] = None
                    questionFileObj['R12-hint'] = None
                    questionFileObj['R13'] = None
                    questionFileObj['R13-hint'] = None
                    questionFileObj['R14'] = None
                    questionFileObj['R14-hint'] = None
                    questionFileObj['R15'] = None
                    questionFileObj['R15-hint'] = None
                    questionFileObj['R16'] = None
                    questionFileObj['R16-hint'] = None
                    questionFileObj['R17'] = None
                    questionFileObj['R17-hint'] = None
                    questionFileObj['R18'] = None
                    questionFileObj['R18-hint'] = None
                    questionFileObj['R19'] = None
                    questionFileObj['R19-hint'] = None
                    questionFileObj['R20'] = None
                    questionFileObj['R20-hint'] = None
                    questionFileObj['_arrayFields'] = None
                if ques['section_header']:
                    questionFileObj['sectionHeader'] = ques['section_header'].encode('utf-8').decode('utf-8')
                else:
                    questionFileObj['sectionHeader'] = None
                questionFileObj['page'] = ques['page']
                if type(ques['question_number']) != str:
                    if ques['question_number'] and ques['question_number'].is_integer() == True:
                        questionFileObj['questionNumber'] = int(ques['question_number'])
                    elif ques['question_number']:
                        questionFileObj['questionNumber'] = ques['question_number']
                else:
                    questionFileObj['questionNumber'] = ques['question_number']
                questionFileObj['prefillFromEntityProfile'] = None
                questionFileObj['isEditable'] = 'TRUE'
                questionFileObj['entityFieldName'] = None
                questionFileObj['_arrayFields'] = 'parentQuestionValue'
                writerQuestionUpload.writerow(questionFileObj)
        bodySolutionUpdate = {"questionSequenceByEcm": questionSeqByEcmDict}
        Helpers.solutionUpdate(solutionName_for_folder_path, accessToken, solutionId, bodySolutionUpdate)

        urlQuestionsUploadApi = internal_kong_ip + questionuploadapiurl
        headerQuestionUploadApi = {'Authorization': authorization,
                                'X-authenticated-user-token': accessToken,
                                'X-Channel-id': x_channel_id}
        filesQuestion = {
            'questions': open(solutionName_for_folder_path + '/questionUpload/uploadSheet.csv', 'rb')
        }
        responseQuestionUploadApi = requests.post(url=urlQuestionsUploadApi, headers=headerQuestionUploadApi,
                                                files=filesQuestion)
        messageArr = ["Question Upload sheet prepared.",
                    "File loc : " + solutionName_for_folder_path + '/questionUpload/uploadSheet.csv',
                    "Question upload API called.", "Status code : " + str(responseQuestionUploadApi.status_code)]
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        if responseQuestionUploadApi.status_code == 200:
            print('QuestionUploadApi Success')
            with open(solutionName_for_folder_path + '/questionUpload/uploadInternalIdsSheet.csv','w+',
                    encoding='utf-8') as questionRes:
                questionRes.write(responseQuestionUploadApi.text)
        else:
            messageArr = ["Question Upload Failed.", "Response : " + str(responseQuestionUploadApi.text)]
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print("Question Upload failed.")
            sys.exit()

    def uploadCriteriaRubrics(solutionName_for_folder_path, wbObservation, millisAddObs, accessToken, frameworkExternalId,
                          withRubricsFlag):
        if withRubricsFlag:
            criteriaRubricSheet = wbObservation.sheet_by_name('Criteria_Rubric-Scoring')
            dictSolCritLookUp = dict()
            filePath = os.path.join(solutionName_for_folder_path + "/solutionCriteriaFetch/", "solutionCriteriaDetails.csv")
            with open(filePath, 'r',encoding='utf-8') as criteriaInternalFile:
                criteriaInternalReader = csv.DictReader(criteriaInternalFile)
                for crit in criteriaInternalReader:
                    dictSolCritLookUp[crit['criteriaID']] = [crit['criteriaInternalId'], crit['criteriaName']]
    
        else:
            criteriaRubricSheet = wbObservation.sheet_by_name('criteria')
            dictSolCritLookUp = dict()
            filePath = os.path.join(solutionName_for_folder_path + "/solutionCriteriaFetch/", "solutionCriteriaDetails.csv")
            with open(filePath, 'r',encoding='utf-8') as criteriaInternalFile:
                criteriaInternalReader = csv.DictReader(criteriaInternalFile)
                for crit in criteriaInternalReader:
                    dictSolCritLookUp[crit['criteriaID']] = [crit['criteriaInternalId'], crit['criteriaName']]

        keys = [criteriaRubricSheet.cell(1, col_index).value for col_index in range(criteriaRubricSheet.ncols)]
        criteriaRubricUploadFieldnames = ["externalId", "name", "criteriaId", "weightage", "expressionVariables"]

        if withRubricsFlag:
            for cl in criteriaLevels:
                criteriaRubricUploadFieldnames.append("L" + str(cl))
        else:
            criteriaRubricUploadFieldnames.append("L1")
        criteriaRubricUpload = dict()
        criteriaRubricsFilePath = solutionName_for_folder_path + '/criteriaRubrics/'
        file_exists_ques = os.path.isfile(solutionName_for_folder_path + '/criteriaRubrics/uploadSheet.csv')
        if not os.path.exists(criteriaRubricsFilePath):
            os.mkdir(criteriaRubricsFilePath)
        if withRubricsFlag:
            for row_index in range(2, criteriaRubricSheet.nrows):
                file_exists_ques = os.path.isfile(solutionName_for_folder_path + '/criteriaRubrics/uploadSheet.csv')
                with open(solutionName_for_folder_path + '/criteriaRubrics/uploadSheet.csv', 'a',
                        encoding='utf-8') as questionUploadFile:
                    writerQuestionUpload = csv.DictWriter(questionUploadFile, fieldnames=criteriaRubricUploadFieldnames,
                                                        lineterminator='\n')
                    if not file_exists_ques:
                        writerQuestionUpload.writeheader()
                    dictCriteriaRubric = {keys[col_index]: criteriaRubricSheet.cell(row_index, col_index).value for
                                        col_index in range(criteriaRubricSheet.ncols)}
                    criteriaRubricUpload['externalId'] = dictCriteriaRubric['criteriaId'] + "_" + str(millisAddObs)
                    print(criteriaRubricUpload['externalId'])
                    criteriaRubricUpload['name'] = dictSolCritLookUp[criteriaRubricUpload['externalId']][1]
                    criteriaRubricUpload['criteriaId'] = dictSolCritLookUp[criteriaRubricUpload['externalId']][0]
                    if dictCriteriaRubric['weightage']:
                        criteriaRubricUpload['weightage'] = dictCriteriaRubric['weightage']
                    else:
                        criteriaRubricUpload['weightage'] = 0
                    criteriaRubricUpload['expressionVariables'] = "SCORE=" + criteriaRubricUpload[
                        'criteriaId'] + ".scoreOfAllQuestionInCriteria()"
                    for cl in criteriaLevels:
                        criteriaRubricUpload['L' + str(cl)] = dictCriteriaRubric['L' + str(cl) + " SCORE"]
                    writerQuestionUpload.writerow(criteriaRubricUpload)
        else:
            for criteriaIds, criteriaDetails in dictSolCritLookUp.items():
                file_exists_ques = os.path.isfile(solutionName_for_folder_path + '/criteriaRubrics/uploadSheet.csv')
                with open(solutionName_for_folder_path + '/criteriaRubrics/uploadSheet.csv', 'a',
                        encoding='utf-8') as questionUploadFile:
                    writerQuestionUpload = csv.DictWriter(questionUploadFile, fieldnames=criteriaRubricUploadFieldnames,
                                                        lineterminator='\n')
                    if not file_exists_ques:
                        writerQuestionUpload.writeheader()
                    criteriaRubricUpload['externalId'] = criteriaIds
                    criteriaRubricUpload['name'] = criteriaDetails[1]
                    criteriaRubricUpload['weightage'] = 1
                    criteriaRubricUpload['criteriaId'] = criteriaDetails[0]
                    criteriaRubricUpload['expressionVariables'] = 'SCORE=' + str(
                        criteriaDetails[0]) + '.scoreOfAllQuestionInCriteria()'
                    criteriaRubricUpload['L1'] = '0<=SCORE<=100000'
                    writerQuestionUpload.writerow(criteriaRubricUpload)

        urlCriteriaRubricUploadApi = internal_kong_ip + criteriarubricuploadapiurl + frameworkExternalId + "-OBSERVATION-TEMPLATE"
        headerCriteriaRubricUploadApi = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id
        }
        filesCriteriaRubric = {
            'criteria': open(solutionName_for_folder_path + '/criteriaRubrics/uploadSheet.csv', 'rb')
        }
        responseCriteriaRubricUploadApi = requests.post(url=urlCriteriaRubricUploadApi,
                                                        headers=headerCriteriaRubricUploadApi, files=filesCriteriaRubric)
        messageArr = ["Criteria Rubric upload sheet prepared.",
                    "File Loc : " + solutionName_for_folder_path + '/criteriaRubrics/uploadSheet.csv',
                    "Status Code : " + str(responseCriteriaRubricUploadApi.status_code)]
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)
        if responseCriteriaRubricUploadApi.status_code == 200:
            with open(solutionName_for_folder_path + '/criteriaRubrics/uploadInternalIdsSheet.csv',
                    'w+',encoding='utf-8') as criteriaRubricRes:
                criteriaRubricRes.write(responseCriteriaRubricUploadApi.text)
        else:
            messageArr = ["Criteria Rubric upload Failed.", "Response : " + str(responseCriteriaRubricUploadApi.text)]
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print("Criteria Rubric upload Failed.")
            sys.exit()

    def fetchSolutionCriteria(solutionName_for_folder_path, observationId, accessToken):
        url = internal_kong_ip + ferchsolutioncriteria + observationId

        headers = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'internal-access-token': internal_access_token
        }

        response = requests.request("POST", url, headers=headers)
        messageArr = ["Criteria solution fetch API called.", "Status Code  : " + str(response.status_code), "URL : " + url]
        Helpers.createAPILog(solutionName_for_folder_path, messageArr)

        os.mkdir(solutionName_for_folder_path + "/solutionCriteriaFetch/")
        if response.status_code == 200:
            print("Solution criteria fetched.")
            with open(solutionName_for_folder_path + "/solutionCriteriaFetch/solutionCriteriaDetails.csv",
                    'w+',encoding='utf-8') as solutionCriteriaFetch:
                solutionCriteriaFetch.write(response.text)
        else:
            messageArr = ["Criteria solution fetch API failed.", "Response  : " + str(response.text)]
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print("Solution criteria fetch failed. Status Code : " + str(response.status_code))
            sys.exit()


    def uploadThemeRubrics(solutionName_for_folder_path, wbObservation, accessToken, frameworkExternalId, withRubricsFlag):
        themeRubricUploadFieldnames = ["externalId", "name", "weightage"]
        themeRubricsFilePath = os.path.join(solutionName_for_folder_path, "themeRubrics/")
        if not os.path.exists(themeRubricsFilePath):
            os.mkdir(themeRubricsFilePath)
        themeRubricUpload = dict()
        if withRubricsFlag:
            themeRubricSheet = wbObservation.sheet_by_name('Domain(theme)_rubric_scoring')
            keys = [themeRubricSheet.cell(1, col_index).value for col_index in range(themeRubricSheet.ncols)]
            themeRubricUploadFieldnames = ["externalId", "name", "weightage"]
            if withRubricsFlag:
                for cl in criteriaLevels:
                    themeRubricUploadFieldnames.append("L" + str(cl))
            else:
                themeRubricUploadFieldnames.append("L1")

            for row_index in range(2, themeRubricSheet.nrows):
                file_exists_ques = os.path.isfile(solutionName_for_folder_path + '/themeRubrics/uploadSheet.csv')
                with open(solutionName_for_folder_path + '/themeRubrics/uploadSheet.csv', 'a',
                        encoding='utf-8') as themeRubricsUploadFile:
                    writerThemeRubricsUpload = csv.DictWriter(themeRubricsUploadFile,
                                                            fieldnames=themeRubricUploadFieldnames, lineterminator='\n')
                    if not file_exists_ques:
                        writerThemeRubricsUpload.writeheader()

                    dictThemeRubric = {keys[col_index]: themeRubricSheet.cell(row_index, col_index).value for col_index in
                                    range(themeRubricSheet.ncols)}
                    themeRubricUpload['externalId'] = dictThemeRubric['domain_Id']
                    themeRubricUpload['name'] = dictThemeRubric['domain_name'].encode('utf-8').decode('utf-8')
                    if dictThemeRubric['weightage']:
                        themeRubricUpload['weightage'] = dictThemeRubric['weightage']
                    else:
                        themeRubricUpload['weightage'] = 0
                    if withRubricsFlag:
                        for cl in criteriaLevels:
                            themeRubricUpload['L' + str(cl)] = dictThemeRubric['L' + str(cl)]
                    else:
                        themeRubricUpload['L1'] = '0<=SCORE<=100000'
                    writerThemeRubricsUpload.writerow(themeRubricUpload)
        else:
            themeRubricUploadFieldnames.append("L1")
            file_exists_ques = os.path.isfile(solutionName_for_folder_path + '/themeRubrics/uploadSheet.csv')
            with open(solutionName_for_folder_path + '/themeRubrics/uploadSheet.csv', 'a',
                    encoding='utf-8') as themeRubricsUploadFile:
                writerThemeRubricsUpload = csv.DictWriter(themeRubricsUploadFile, fieldnames=themeRubricUploadFieldnames,
                                                        lineterminator='\n')
                if not file_exists_ques:
                    writerThemeRubricsUpload.writeheader()
                themeRubricUpload['externalId'] = "OB"
                themeRubricUpload['name'] = "Observation Theme"
                themeRubricUpload['weightage'] = 1
                themeRubricUpload['L1'] = '0<=SCORE<=100000'
                writerThemeRubricsUpload.writerow(themeRubricUpload)
        urlThemeRubricUploadApi = internal_kong_ip + themerubricuploadapiurl + frameworkExternalId + "-OBSERVATION-TEMPLATE"
        headerThemeRubricUploadApi = {
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id
        }
        filesThemeRubric = {
            'themes': open(solutionName_for_folder_path + '/themeRubrics/uploadSheet.csv', 'rb')
        }
        responseThemeRubricUploadApi = requests.post(url=urlThemeRubricUploadApi, headers=headerThemeRubricUploadApi,
                                                    files=filesThemeRubric)
        if responseThemeRubricUploadApi.status_code == 200:
            print('ThemeRubricUploadApi Success')
            with open(solutionName_for_folder_path + '/themeRubrics/uploadInternalIdsSheet.csv', 'w+',encoding='utf-8') as themeRubricRes:
                themeRubricRes.write(responseThemeRubricUploadApi.text)
        else:
            messageArr = ['theme rubric upload api failed in ' + environment,
                        ' status_code response from api is ' + str(responseThemeRubricUploadApi.status_code),
                        "Response : " + str(responseThemeRubricUploadApi.text)]
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            print('theme rubric upload api failed in ' + environment + ' status_code response from api is ' + str(responseThemeRubricUploadApi.status_code))
            sys.exit()

    def prepareSuccessSheet(solutionName_for_folder_path, filePathAddObs, observationExternalId, millisAddObs):
        updateSuccessWorkBook = xlrd.open_workbook(filePathAddObs, on_demand=True)
        updateWbNumberOfSheets = updateSuccessWorkBook.nsheets
        updateWbSheetNames = updateSuccessWorkBook.sheet_names()
        updateCriteriaSheet = updateSuccessWorkBook.sheet_by_name('Criteria_Rubric-Scoring')
        updateQuestionsSheet = updateSuccessWorkBook.sheet_by_name('questions')
        updateDetailsSheet = updateSuccessWorkBook.sheet_by_name('details')
        copyOfUpdateWb = copy(updateSuccessWorkBook)
        updateQuestionsSheetCopy = copyOfUpdateWb.get_sheet('questions')
        for each in range(updateWbNumberOfSheets):
            eachUpdateWorkSheet = copyOfUpdateWb.get_sheet(each)
            if (eachUpdateWorkSheet.name).strip() == 'Criteria_Rubric-Scoring':
                for row_idx_crit in range(1, updateCriteriaSheet.nrows):
                    for col_idx_crit in range(0, updateCriteriaSheet.ncols):
                        if col_idx_crit == 0:
                            eachUpdateWorkSheet.write(row_idx_crit, col_idx_crit,
                                                    updateCriteriaSheet.cell(row_idx_crit, col_idx_crit).value.replace(
                                                        '\n', '').strip() + '_' + str(millisAddObs))
            if (eachUpdateWorkSheet.name).strip().lower() == 'questions':
                for row_idx_ques in range(1, updateQuestionsSheet.nrows):
                    for col_idx_ques in range(0, updateQuestionsSheet.ncols):
                        if col_idx_ques == 2 or col_idx_ques == 0:
                            eachUpdateWorkSheet.write(row_idx_ques, col_idx_ques,
                                                    updateQuestionsSheet.cell(row_idx_ques, col_idx_ques).value.replace(
                                                        '\n', '').strip() + '_' + str(millisAddObs))
                for row_0 in range(0, updateQuestionsSheet.nrows):
                    if row_0 == 0:
                        eachUpdateWorkSheet.write(row_0, updateQuestionsSheet.ncols, 'question_operations')
                    else:
                        eachUpdateWorkSheet.write(row_0, updateQuestionsSheet.ncols, None)
            if (eachUpdateWorkSheet.name).strip().lower() == 'details':
                eachUpdateWorkSheet.write(1, 1, observationExternalId)
                for row_details_0 in range(0, updateDetailsSheet.nrows):
                    if row_details_0 == 0:
                        eachUpdateWorkSheet.write(row_details_0, updateDetailsSheet.ncols, 'solution_name_update')
                    else:
                        eachUpdateWorkSheet.write(row_details_0, updateDetailsSheet.ncols, None)
        copyOfUpdateWb.save(solutionName_for_folder_path.replace('.xlsx', '') + '_styles.xlsx')
        workbook = open_workbook(solutionName_for_folder_path.replace('.xlsx', '') + '_styles.xlsx')
        # Process each sheet
        for sheet in workbook.sheets():
            # Make a copy of the master worksheet
            new_workbook = copy(workbook)
            # for each time we copy the master workbook, remove all sheets except
            #  for the curren sheet (as defined by sheet.name)
            new_workbook._Workbook__worksheets = [worksheet for worksheet in new_workbook._Workbook__worksheets if
                                                worksheet.name != 'questions_sequence_sorted']
            # Save the new_workbook based on sheet.name
            new_workbook.save(solutionName_for_folder_path.replace('.xlsx', '') + '_styles.xlsx'.format(sheet.name))
        workbookXlsxWriter = xlsxwriter.Workbook(solutionName_for_folder_path.replace('.xlsx', '') + '_Success.xlsx')
        updateSuccessWorkBookReopen = xlrd.open_workbook(solutionName_for_folder_path.replace('.xlsx', '') + '_styles.xlsx',
                                                        on_demand=True)
        updateWbNumberOfSheetsReopen = updateSuccessWorkBookReopen.nsheets
        updateWbSheetNamesReopen = updateSuccessWorkBookReopen.sheet_names()
        updateQuestionsSheetReopen = updateSuccessWorkBookReopen.sheet_by_name('questions')
        updateDetailsSheetReopen = updateSuccessWorkBookReopen.sheet_by_name('details')
        cellFormat = workbookXlsxWriter.add_format()
        cellFormat.set_bg_color('00FF00')
        unlockCell = workbookXlsxWriter.add_format({'locked': False})
        for ele in updateWbSheetNamesReopen:
            if ele == 'details' or ele == 'questions' or ele == 'questions_sequence_sorted':
                updateWbSheetNamesReopen.remove(ele)
        for suSh in updateWbSheetNamesReopen:
            worksheetXlsxWriter = workbookXlsxWriter.add_worksheet(suSh)
            eachSheetByName = updateSuccessWorkBookReopen.sheet_by_name(suSh)
            for row_indx_sheets in range(eachSheetByName.nrows):
                for col_indx_sheets in range(eachSheetByName.ncols):
                    worksheetXlsxWriter.write(row_indx_sheets, col_indx_sheets,
                                            eachSheetByName.cell(row_indx_sheets, col_indx_sheets).value)
        questionsWorkSheetSuccess = workbookXlsxWriter.add_worksheet('questions')
        for row_idx_ques_succ in range(updateQuestionsSheetReopen.nrows):
            for col_idx_ques_succ in range(updateQuestionsSheetReopen.ncols):
                if col_idx_ques_succ == 0 or col_idx_ques_succ == 2:
                    questionsWorkSheetSuccess.protect()
                    questionsWorkSheetSuccess.write(row_idx_ques_succ, col_idx_ques_succ,
                                                    updateQuestionsSheetReopen.cell(row_idx_ques_succ,
                                                                                    col_idx_ques_succ).value, cellFormat)
                else:
                    questionsWorkSheetSuccess.write(row_idx_ques_succ, col_idx_ques_succ,
                                                    updateQuestionsSheetReopen.cell(row_idx_ques_succ,
                                                                                    col_idx_ques_succ).value, unlockCell)
                if updateQuestionsSheetReopen.ncols - 1 == col_idx_ques_succ:
                    questionsWorkSheetSuccess.data_validation(1, updateQuestionsSheetReopen.ncols - 1,
                                                            updateQuestionsSheetReopen.nrows,
                                                            updateQuestionsSheetReopen.ncols - 1,
                                                            {'validate': 'list', 'source': ['ADD', 'UPDATE', 'DELETE']})
        questionsWorkSheetSuccess.write_comment(0, 0,
                                                'criteria_id column is locked can\'t be edited , as it will be useful in updating the observations')
        questionsWorkSheetSuccess.write_comment(0, 2,
                                                'question_id column is locked can\'t be edited , as it will be useful in updating the observations')
        questionsWorkSheetSuccess.write_comment(0, updateQuestionsSheetReopen.ncols - 1,
                                                'question_operation column can be used in updating the questions , select either one of the options to update else leave blank and send the template to genie with update observation template command')
        detailsWorkSheetSuccess = workbookXlsxWriter.add_worksheet('details')
        for row_idx_deta_succ in range(updateDetailsSheetReopen.nrows):
            for col_idx_deta_succ in range(updateDetailsSheetReopen.ncols):
                if col_idx_deta_succ == 1:
                    detailsWorkSheetSuccess.protect()
                    detailsWorkSheetSuccess.write(row_idx_deta_succ, col_idx_deta_succ,
                                                updateDetailsSheetReopen.cell(row_idx_deta_succ, col_idx_deta_succ).value,
                                                cellFormat)
                else:
                    detailsWorkSheetSuccess.write(row_idx_deta_succ, col_idx_deta_succ,
                                                updateDetailsSheetReopen.cell(row_idx_deta_succ, col_idx_deta_succ).value,
                                                unlockCell)
                if updateDetailsSheetReopen.ncols - 1 == col_idx_deta_succ:
                    detailsWorkSheetSuccess.data_validation(1, updateDetailsSheetReopen.ncols - 1,
                                                            updateDetailsSheetReopen.nrows,
                                                            updateDetailsSheetReopen.ncols - 1,
                                                            {'validate': 'list', 'source': ['TRUE', 'FALSE']})
        detailsWorkSheetSuccess.write_comment(0, 1,
                                            'observation_id column is locked can\'t be edited , as it will be useful in updating the observations')
        detailsWorkSheetSuccess.write_comment(0, updateDetailsSheetReopen.ncols - 1,
                                            'solution_name_update column can be used in updating the solution_name , select either TRUE or FALSE and send the template to genie with update observation template command')
        sheet_names = ['Instructions', 'details', 'Criteria upload', 'Criteria_Rubric-Scoring',
                    'Domain(theme)_rubric_scoring', 'questions', 'framework', 'ECMs or Domains']
        workbookXlsxWriter.worksheets_objs.sort(key=lambda x: sheet_names.index(x.name))
        workbookXlsxWriter.close()
        print("Success sheet prepared.")


    def createChild(solutionName_for_folder_path, observationExternalId, accessToken):
        print("it has entered the creation api")
        childObservationExternalId = str(observationExternalId + "_CHILD")
        urlSol_prog_mapping = internal_kong_ip + solutiontoprogrammappingapiurl + "?solutionId=" + observationExternalId + "&entityType=" + entityType
        
        payloadSol_prog_mapping = {
            "externalId": childObservationExternalId,
            "name": solutionName.lstrip().rstrip(),
            "description": solutionDescription.lstrip().rstrip(),
            "programExternalId": programExternalId
        }
        headersSol_prog_mapping = {'Authorization': authorization,
                                'X-authenticated-user-token': accessToken,
                                'Content-Type': content_type}
        responseSol_prog_mapping = requests.request("POST", urlSol_prog_mapping, headers=headersSol_prog_mapping,
                                                    data=json.dumps(payloadSol_prog_mapping))
        messageArr = ["Create child API called.", "URL : " + urlSol_prog_mapping,
                    "Status code : " + str(responseSol_prog_mapping.status_code),
                    "Response : " + responseSol_prog_mapping.text, "body : " + str(payloadSol_prog_mapping)]
        if responseSol_prog_mapping.status_code == 200:
            print("Solution mapped to program : " + programName)
            print("Child solution : " + childObservationExternalId)

            responseSol_prog_mapping = responseSol_prog_mapping.json()
            child_id = responseSol_prog_mapping['result']['_id']
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            return [child_id, childObservationExternalId]
        else:
            print("Unable to create child solution")

            messageArr.append("Unable to create child solution")
            Helpers.createAPILog(solutionName_for_folder_path, messageArr)
            return False
    
    def createSurveySolution(parentFolder, wbSurvey, accessToken):
        print("Create Survey Solution Func Called....")
        # print(wbSurvey,"4732")
        sheetNames1 = wbSurvey.sheet_names()
        # print(sheetNames1,"sheetNames1 4744")
        # wbprogram = xlrd.open_workbook(wbObservation, on_demand=True)
        # programSheetNames = wbprogram.sheet_names()
        # wbObservation1 = xlrd.open_workbook(wbsurvey, on_demand=True)
        # sheetNames1 = wbObservation1.sheet_names()
        # print(sheetNames1,"4733")
        for sheetEnv in sheetNames1:
            if sheetEnv.strip().lower() == 'details':
                surveySolutionCreationReqBody = {}
                detailsEnvSheet = wbSurvey.sheet_by_name(sheetEnv)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                        range(detailsEnvSheet.ncols)]

                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                    for
                                    col_index_env in range(detailsEnvSheet.ncols)}
                    surveySolutionCreationReqBody['name'] = dictDetailsEnv['survey_solution_name'].encode('utf-8').decode('utf-8')
                    surveySolutionCreationReqBody["description"] = dictDetailsEnv['survey_solution_description'].encode('utf-8').decode('utf-8')
                    surveySolutionExternalId = str(uuid.uuid1())
                    surveySolutionCreationReqBody["externalId"] = surveySolutionExternalId
                    if dictDetailsEnv['Name_of_the_creator']== "":
                        exceptionHandlingFlag = True
                        print('survey_creator_username column should not be empty in the details sheet')
                        sys.exit()
                    else:
                        surveySolutionCreationReqBody['creator'] = dictDetailsEnv['Name_of_the_creator']


                    userDetails = Helpers.fetchUserDetails( accessToken, dictDetailsEnv['survey_creator_username'])
                    surveySolutionCreationReqBody['author'] = userDetails[0]
                    if dictDetailsEnv["survey_start_date"]:
                        if type(dictDetailsEnv["survey_start_date"]) == str:
                            startDateArr = None
                            startDateArr = (dictDetailsEnv["survey_start_date"]).split("-")
                            surveySolutionCreationReqBody["startDate"] = startDateArr[2] + "-" + startDateArr[1] + "-" + \
                                                                        startDateArr[0] + " 00:00:00"
                        elif type(dictDetailsEnv["survey_start_date"]) == float:
                            surveySolutionCreationReqBody["startDate"] = (
                                xlrd.xldate.xldate_as_datetime(dictDetailsEnv["survey_start_date"],
                                                            wbSurvey.datemode)).strftime("%Y/%m/%d")
                        else:
                            surveySolutionCreationReqBody["startDate"] = ""
                        if dictDetailsEnv["survey_end_date"]:
                            if type(dictDetailsEnv["survey_end_date"]) == str:
                                endDateArr = None
                                endDateArr = (dictDetailsEnv["survey_end_date"]).split("-")
                                surveySolutionCreationReqBody["endDate"] = endDateArr[2] + "-" + endDateArr[1] + "-" + \
                                                                        endDateArr[0] + "T23:59:59.000Z"
                            elif type(dictDetailsEnv["survey_end_date"]) == float:
                                surveySolutionCreationReqBody["endDate"] = (
                                    xlrd.xldate.xldate_as_datetime(dictDetailsEnv["survey_end_date"],
                                                                wbSurvey.datemode)).strftime("%Y/%m/%d")
                            else:
                                surveySolutionCreationReqBody["endDate"] = ""
                            enDt = surveySolutionCreationReqBody["endDate"]
                            
                            urlCreateSolutionApi = internal_kong_ip+ surveysolutioncreationapiurl
                            headerCreateSolutionApi = {
                                'Content-Type': content_type,
                                'Authorization': authorization,
                                'X-authenticated-user-token': accessToken,
                                'X-Channel-id': x_channel_id,
                                'appName': appname
                            }
                            responseCreateSolutionApi = requests.post(url=urlCreateSolutionApi,
                                                                    headers=headerCreateSolutionApi,
                                                                    data=json.dumps(surveySolutionCreationReqBody))
                            responseInText = responseCreateSolutionApi.text
                            messageArr = ["********* Create Survey Solution *********", "URL : " + urlCreateSolutionApi,
                                        "BODY : " + str(surveySolutionCreationReqBody),
                                        "Status code : " + str(responseCreateSolutionApi.status_code),
                                        "Response : " + responseCreateSolutionApi.text]
                            fileheader = [surveySolutionCreationReqBody['name'].encode('utf-8').decode('utf-8'),'Program Sheet Validation'," "]
                            Helpers.createAPILog(parentFolder, messageArr)
                            Helpers.apicheckslog(parentFolder,fileheader)
                            if responseCreateSolutionApi.status_code == 200:
                                responseCreateSolutionApi = responseCreateSolutionApi.json()
                                urlSearchSolution = internal_kong_ip + fetchsolutiondetails + "survey&page=1&limit=10&search=" + str(surveySolutionExternalId)
                                responseSearchSolution = requests.request("POST", urlSearchSolution,
                                                                        headers=headerCreateSolutionApi)
                                messageArr = ["********* Search Survey Solution *********", "URL : " + urlSearchSolution,
                                            "Status code : " + str(responseSearchSolution.status_code),
                                            "Response : " + responseSearchSolution.text]
                                Helpers.createAPILog(parentFolder, messageArr)
                                Helpers.apicheckslog(parentFolder, messageArr)
                                if responseSearchSolution.status_code == 200:
                                    responseSearchSolutionApi = responseSearchSolution.json()
                                    surveySolutionExternalId = None
                                    surveySolutionExternalId = responseSearchSolutionApi['result']['data'][0]['externalId']
                                else:
                                    print("Solution fetch API failed")
                                    print("URL : " + urlSearchSolution)
                                    Helpers.terminatingMessage("Status Code : " + responseSearchSolution.status_code)

                                solutionId = None
                                solutionId = responseCreateSolutionApi["result"]["solutionId"]
                                bodySolutionUpdate = {"creator": dictDetailsEnv['Name_of_the_creator'].encode('utf-8').decode('utf-8')}
                                Helpers.solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)

                                return [solutionId, surveySolutionExternalId]
                            else:
                                Helpers.terminatingMessage("Survey creation Failed, check logs!")

    # upload survey questions 
    def uploadSurveyQuestions(MainFilePath, parentFolder, wbSurvey, addObservationSolution, accessToken, surTempExtID, surTempSolID, millisecond, programFile):
        print("Upload Survey Questions Func Called....")
        # print(parentFolder,"4854")
        # wbSurvey = xlrd.open_workbook(wbSurvey, on_demand=True)
        # print(f"Type of wbSurvey: {type(wbSurvey)}")
        sheetNam = wbSurvey.sheet_names()
        # print(sheetNam,"4854")
        global surveySolutionlink
        stDt = None
        enDt = None
        shCnt = 0
        for i in sheetNam:
            if i.strip().lower() == 'questions':
                sheetNam1 = wbSurvey.sheets()[shCnt]
            shCnt = shCnt + 1
        dataSort = [sheetNam1.row_values(i) for i in range(sheetNam1.nrows)]
        labels = dataSort[1]
        dataSort = dataSort[2:]
        dataSort.sort(key=lambda x: int(x[0]))
        openWorkBookSort1 = xl_copy(wbSurvey)
        sheet1 = openWorkBookSort1.add_sheet('questions_sequence_sorted')

        for idx, label in enumerate(labels):
            sheet1.write(0, idx, label)

        for idx_r, row in enumerate(dataSort):
            for idx_c, value in enumerate(row):
                sheet1.write(idx_r + 1, idx_c, value)
        newFileName = str(addObservationSolution)
        openWorkBookSort1.save(newFileName)
        openNewFile = xlrd.open_workbook(newFileName, on_demand=True)
        wbSurvey = openNewFile
        sheetNames = wbSurvey.sheet_names()
        # print("reached till here 4881")
        for sheet2 in sheetNames:
            if sheet2.strip().lower() == 'questions_sequence_sorted':
                questionsList = []
                questionsSheet = wbSurvey.sheet_by_name(sheet2.lower())
                keys2 = [questionsSheet.cell(0, col_index2).value for col_index2 in
                        range(questionsSheet.ncols)]
                for row_index2 in range(1, questionsSheet.nrows):
                    d2 = {keys2[col_index2]: questionsSheet.cell(row_index2, col_index2).value
                        for col_index2 in range(questionsSheet.ncols)}
                    questionsList.append(d2)
                questionSeqByEcmArr = []
                quesSeqCnt = 1.0
                questionUploadFieldnames = []
                questionUploadFieldnames = ['solutionId', 'instanceParentQuestionId','hasAParentQuestion', 'parentQuestionOperator','parentQuestionValue', 'parentQuestionId','externalId', 'question0', 'question1', 'tip','hint', 'instanceIdentifier', 'responseType','dateFormat', 'autoCapture', 'validation','validationIsNumber', 'validationRegex','validationMax', 'validationMin', 'file','fileIsRequired', 'fileUploadType','allowAudioRecording', 'minFileCount','maxFileCount', 'caption', 'questionGroup','modeOfCollection', 'accessibility', 'showRemarks','rubricLevel', 'isAGeneralQuestion', 'R1','R1-hint', 'R2', 'R2-hint', 'R3', 'R3-hint', 'R4','R4-hint', 'R5', 'R5-hint', 'R6', 'R6-hint', 'R7','R7-hint', 'R8', 'R8-hint', 'R9', 'R9-hint', 'R10','R10-hint', 'R11', 'R11-hint', 'R12', 'R12-hint','R13', 'R13-hint', 'R14', 'R14-hint', 'R15','R15-hint', 'R16', 'R16-hint', 'R17', 'R17-hint','R18', 'R18-hint', 'R19', 'R19-hint', 'R20','R20-hint', 'sectionHeader', 'page','questionNumber', '_arrayFields']

                for ques in questionsList:

                    questionFilePath = parentFolder + '/questionUpload/'
                    file_exists_ques = os.path.isfile(
                        parentFolder + '/questionUpload/uploadSheet.csv')
                    # print(questionFilePath,"4904")
                    if not os.path.exists(questionFilePath):
                        os.mkdir(questionFilePath)
                    with open(parentFolder + '/questionUpload/uploadSheet.csv', 'a',
                            encoding='utf-8') as questionUploadFile:
                        writerQuestionUpload = csv.DictWriter(questionUploadFile, fieldnames=questionUploadFieldnames, lineterminator='\n')
                        if not file_exists_ques:
                            writerQuestionUpload.writeheader()
                        questionFileObj = {}
                        surveyExternalId = None
                        questionFileObj['solutionId'] = surTempExtID
                        if ques['instance_parent_question_id'].encode('utf-8').decode('utf-8'):
                            questionFileObj['instanceParentQuestionId'] = ques[
                                                                            'instance_parent_question_id'].strip() + '_' + str(
                                millisecond)
                        else:
                            questionFileObj['instanceParentQuestionId'] = 'NA'
                        if ques['parent_question_id'].encode('utf-8').decode('utf-8').strip():
                            questionFileObj['hasAParentQuestion'] = 'YES'
                            if ques['show_when_parent_question_value_is'] == 'or':
                                questionFileObj['parentQuestionOperator'] = '||'
                            else:
                                questionFileObj['parentQuestionOperator'] = ques['show_when_parent_question_value_is']
                            if type(ques['parent_question_value']) != str:
                                if (ques['parent_question_value'] and ques[
                                    'parent_question_value'].is_integer() == True):
                                    questionFileObj['parentQuestionValue'] = int(ques['parent_question_value'])
                                elif (ques['parent_question_value'] and ques[
                                    'parent_question_value'].is_integer() == False):
                                    questionFileObj['parentQuestionValue'] = ques['parent_question_value']
                            else:
                                questionFileObj['parentQuestionValue'] = ques['parent_question_value']
                                questionFileObj['parentQuestionId'] = ques['parent_question_id'].encode('utf-8').decode('utf-8').strip() + '_' + str(
                                    millisecond)
                        else:
                            questionFileObj['hasAParentQuestion'] = 'NO'
                            questionFileObj['parentQuestionOperator'] = None
                            questionFileObj['parentQuestionValue'] = None
                            questionFileObj['parentQuestionId'] = None
                        questionFileObj['externalId'] = ques['question_id'].strip() + '_' + str(millisecond)
                        if quesSeqCnt == ques['question_sequence']:
                            questionSeqByEcmArr.append(ques['question_id'].strip() + '_' + str(millisecond))
                            quesSeqCnt = quesSeqCnt + 1.0
                        if ques['question_language1']:
                            questionFileObj['question0'] = ques['question_language1'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['question0'] = None
                        if ques['question_language2']:
                            questionFileObj['question1'] = ques['question_language2'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['question1'] = None
                        if ques['question_tip']:
                            questionFileObj['tip'] = ques['question_tip'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['tip'] = None
                        if ques['question_hint']:
                            questionFileObj['hint'] = ques['question_hint'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['hint'] = None
                        if ques['instance_identifier']:
                            questionFileObj['instanceIdentifier'] = ques['instance_identifier'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['instanceIdentifier'] = None
                        if ques['question_response_type'].strip().lower():
                            questionFileObj['responseType'] = ques['question_response_type'].strip().lower()
                        if ques['question_response_type'].strip().lower() == 'date':
                            questionFileObj['dateFormat'] = "DD-MM-YYYY"
                        else:
                            questionFileObj['dateFormat'] = None
                        if ques['question_response_type'].strip().lower() == 'date':
                            if ques['date_auto_capture'] and ques['date_auto_capture'] == 1:
                                questionFileObj['autoCapture'] = 'TRUE'
                            elif ques['date_auto_capture'] and ques['date_auto_capture'] == 0:
                                questionFileObj['autoCapture'] = 'false'
                            else:
                                questionFileObj['autoCapture'] = 'false'
                        else:
                            questionFileObj['autoCapture'] = None
                        if ques['response_required']:
                            if ques['response_required'] == 1:
                                questionFileObj['validation'] = 'TRUE'
                            elif ques['response_required'] == 0:
                                questionFileObj['validation'] = 'FALSE'
                        else:
                            questionFileObj['validation'] = 'FALSE'
                        if ques['question_response_type'].strip().lower() == 'number':
                            questionFileObj['validationIsNumber'] = 'TRUE'
                            questionFileObj['validationRegex'] = 'isNumber'
                            if (ques['max_number_value'] and ques['max_number_value'].is_integer() == True):
                                questionFileObj['validationMax'] = int(ques['max_number_value'])
                            elif (ques['max_number_value'] and ques['max_number_value'].is_integer() == False):
                                questionFileObj['validationMax'] = ques['max_number_value']
                            else:
                                questionFileObj['validationMax'] = 10000

                            if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                                questionFileObj['validationMin'] = int(ques['min_number_value'])
                            elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                                questionFileObj['validationMin'] = ques['min_number_value']
                            else:
                                questionFileObj['validationMax'] = 10000

                            if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                                questionFileObj['validationMin'] = int(ques['min_number_value'])
                            elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                                questionFileObj['validationMin'] = ques['min_number_value']
                            else:
                                questionFileObj['validationMin'] = 0

                        elif ques['question_response_type'].strip().lower() == 'slider':
                            questionFileObj['validationIsNumber'] = None
                            questionFileObj['validationRegex'] = 'isNumber'
                            if (ques['max_number_value'] and ques['max_number_value'].is_integer() == True):
                                questionFileObj['validationMax'] = int(ques['max_number_value'])
                            elif (ques['max_number_value'] and ques['max_number_value'].is_integer() == False):
                                questionFileObj['validationMax'] = ques['max_number_value']
                            else:
                                questionFileObj['validationMax'] = 5

                            if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                                questionFileObj['validationMin'] = int(ques['min_number_value'])
                            elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                                questionFileObj['validationMin'] = ques['min_number_value']
                            else:
                                questionFileObj['validationMin'] = 0
                        else:
                            questionFileObj['validationIsNumber'] = None
                            questionFileObj['validationRegex'] = None
                            questionFileObj['validationMax'] = None
                            questionFileObj['validationMin'] = None
                        if ques['file_upload'] == 1:
                            questionFileObj['file'] = 'Snapshot'
                            questionFileObj['fileIsRequired'] = 'TRUE'
                            questionFileObj['fileUploadType'] = 'image/jpeg,docx,pdf,ppt'
                            questionFileObj['minFileCount'] = 0
                            questionFileObj['maxFileCount'] = 10
                        elif ques['file_upload'] == 0:
                            questionFileObj['file'] = 'NA'
                            questionFileObj['fileIsRequired'] = None
                            questionFileObj['fileUploadType'] = None
                            questionFileObj['minFileCount'] = None
                            questionFileObj['maxFileCount'] = None

                        questionFileObj['caption'] = 'FALSE'
                        questionFileObj['questionGroup'] = 'A1'
                        questionFileObj['modeOfCollection'] = 'onfield'
                        questionFileObj['accessibility'] = 'No'
                        if ques['show_remarks'] == 1:
                            questionFileObj['showRemarks'] = 'TRUE'
                        elif ques['show_remarks'] == 0:
                            questionFileObj['showRemarks'] = 'FALSE'
                        questionFileObj['rubricLevel'] = None
                        questionFileObj['isAGeneralQuestion'] = None
                        if ques['question_response_type'].strip().lower() == 'radio' or ques[
                            'question_response_type'].strip() == 'multiselect':
                            for quesIndex in range(1, 21):
                                if type(ques['response(R' + str(quesIndex) + ')']) != str:
                                    if (ques['response(R' + str(quesIndex) + ')'] and ques[
                                        'response(R' + str(quesIndex) + ')'].is_integer() == True):
                                        questionFileObj['R' + str(quesIndex) + ''] = int(
                                            ques['response(R' + str(quesIndex) + ')'])
                                    elif (ques['response(R' + str(quesIndex) + ')'] and ques[
                                        'response(R' + str(quesIndex) + ')'].is_integer() == False):
                                        questionFileObj['R' + str(quesIndex) + ''] = ques[
                                            'response(R' + str(quesIndex) + ')']
                                else:
                                    questionFileObj['R' + str(quesIndex) + ''] = ques[
                                        'response(R' + str(quesIndex) + ')']

                                if type(ques['response(R' + str(quesIndex) + ')_hint']) != str:
                                    if (ques['response(R' + str(quesIndex) + ')_hint'] and ques[
                                        'response(R' + str(quesIndex) + ')_hint'].is_integer() == True):
                                        questionFileObj['R' + str(quesIndex) + '-hint'] = int(
                                            ques['response(R' + str(quesIndex) + ')_hint'])
                                    elif (ques['response(R' + str(quesIndex) + ')_hint'] and ques[
                                        'response(R' + str(quesIndex) + ')_hint'].is_integer() == False):
                                        questionFileObj['R' + str(quesIndex) + '-hint'] = ques[
                                            'response(R' + str(quesIndex) + ')_hint']
                                else:
                                    questionFileObj['R' + str(quesIndex) + '-hint'] = ques[
                                        'response(R' + str(quesIndex) + ')_hint']
                                questionFileObj['_arrayFields'] = 'parentQuestionValue'
                        else:
                            for quesIndex in range(1, 21):
                                questionFileObj['R' + str(quesIndex)] = None
                                questionFileObj['R' + str(quesIndex) + '-hint'] = None
                        if ques['section_header']:
                            questionFileObj['sectionHeader'] = ques['section_header'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['sectionHeader'] = None

                        questionFileObj['page'] = ques['page']
                        if type(ques['question_number']) != str:
                            if ques['question_number'] and ques['question_number'].is_integer() == True:
                                questionFileObj['questionNumber'] = int(ques['question_number'])
                            elif ques['question_number']:
                                questionFileObj['questionNumber'] = ques['question_number']
                            else:
                                questionFileObj['questionNumber'] = ques['question_number']
                        writerQuestionUpload.writerow(questionFileObj)
                        
                urlQuestionsUploadApi = internal_kong_ip + questionuploadapiurl
                headerQuestionUploadApi = {
                    'Authorization': authorization,
                    'X-authenticated-user-token': accessToken,
                    'X-Channel-id': x_channel_id
                }
                filesQuestion = {
                    'questions': open(parentFolder + '/questionUpload/uploadSheet.csv', 'rb')
                }
                responseQuestionUploadApi = requests.post(url=urlQuestionsUploadApi,
                                                        headers=headerQuestionUploadApi, files=filesQuestion)
                if responseQuestionUploadApi.status_code == 200:
                    print('Question upload Success')

                    messageArr = ["********* Question Upload api *********", "URL : " + urlQuestionsUploadApi,
                                "Path : " + str(parentFolder) + str('/questionUpload/uploadSheet.csv'),
                                "Status code : " + str(responseQuestionUploadApi.status_code),
                                "Response : " + responseQuestionUploadApi.text]
                    Helpers.createAPILog(parentFolder, messageArr)
                    messageArr1 = ["Questions","Question upload Success","Passed",str(responseQuestionUploadApi.status_code)]
                    Helpers.apicheckslog(parentFolder,messageArr1)

                    with open(parentFolder + '/questionUpload/uploadInternalIdsSheet.csv', 'w+',encoding='utf-8') as questionRes:
                        questionRes.write(responseQuestionUploadApi.text)
                    urlImportSoluTemplate = internal_kong_ip + importsurveysolutiontemplateurl + str(surTempSolID) + "?appName=manage-learn"
                    headerImportSoluTemplateApi = {
                        'Authorization': authorization,
                        'X-authenticated-user-token': accessToken,
                        'X-Channel-id': x_channel_id
                    }
                    responseImportSoluTemplateApi = requests.get(url=urlImportSoluTemplate,
                                                                headers=headerImportSoluTemplateApi)
                    if responseImportSoluTemplateApi.status_code == 200:
                        print('Creating Child Success')

                        messageArr = ["********* Creating Child api *********", "URL : " + urlImportSoluTemplate,
                                    "Status code : " + str(responseImportSoluTemplateApi.status_code),
                                    "Response : " + responseImportSoluTemplateApi.text]
                        Helpers.createAPILog(parentFolder, messageArr)
                        responseImportSoluTemplateApi = responseImportSoluTemplateApi.json()
                        solutionIdSuc = responseImportSoluTemplateApi["result"]["solutionId"]
                        urlSurveyProgramMapping = internal_kong_ip + importsurveysolutiontoprogramurl + str(solutionIdSuc) + "?programId=" + programExternalId.lstrip().rstrip()
                        headeSurveyProgramMappingApi = {
                            'Authorization': authorization,
                            'X-authenticated-user-token': accessToken,
                            'X-Channel-id': x_channel_id
                        }
                        responseSurveyProgramMappingApi = requests.get(url=urlSurveyProgramMapping,headers=headeSurveyProgramMappingApi)
                        if responseSurveyProgramMappingApi.status_code == 200:
                            print('Program Mapping Success')
                            
                            messageArr = ["********* Program mapping api *********", "URL : " + urlSurveyProgramMapping,
                                        "Status code : " + str(responseSurveyProgramMappingApi.status_code),
                                        "Response : " + responseSurveyProgramMappingApi.text]
                            Helpers.createAPILog(parentFolder, messageArr)
                            surveyLink = None
                            solutionIdSuc = None
                            surveyExternalIdSuc = None
                            surveyLink = responseImportSoluTemplateApi["result"]["link"]
                            solutionIdSuc = responseImportSoluTemplateApi["result"]["solutionId"]
                            solutionExtIdSuc = responseImportSoluTemplateApi["result"]["solutionExternalId"]
                            print("Survey Child Id : " + str(solutionExtIdSuc))
                            solutionDetails = Helpers.fetchSolutionDetailsFromProgramSheet(parentFolder, programFile, solutionIdSuc,
                                                                                accessToken)
                            scopeEntities = entitiesPGMID
                            scopeRoles = solutionDetails[0]
                            surveyScopeBody = {
                                "scope": {"entityType": scopeEntityType, "entities": scopeEntities, "roles": scopeRoles}}
                            Helpers.solutionUpdate(parentFolder, accessToken, solutionIdSuc, surveyScopeBody)
                            surveySolutionlink = Helpers.prepareProgramSuccessSheet(MainFilePath, parentFolder, programFile, solutionExtIdSuc,
                                                    solutionIdSuc, accessToken)
                            
                            print('Survey Successfully Added')
                            print(surveySolutionlink)
                        else:
                            print('Program Mapping Failed')
                            messageArr = ["********* Program mapping api *********", "URL : " + urlSurveyProgramMapping,
                                        "Status code : " + str(responseSurveyProgramMappingApi.status_code),
                                        "Response : " + responseSurveyProgramMappingApi.text]
                            Helpers.createAPILog(parentFolder, messageArr)
                    else:
                        print('Creating Child API Failed')
                        messageArr = ["********* Program mapping api *********", "URL : " + urlImportSoluTemplate,
                                    "Status code : " + str(responseImportSoluTemplateApi.status_code),
                                    "Response : " + responseImportSoluTemplateApi.text]
                        Helpers.createAPILog(parentFolder, messageArr)
                else:
                    print('QuestionUploadApi Failed')
                    messageArr = ["********* Question Upload api *********", "URL : " + urlQuestionsUploadApi,
                                "Path : " + str(parentFolder) + str('/questionUpload/uploadSheet.csv'),
                                "Status code : " + str(responseQuestionUploadApi.status_code),
                                "Response : " + responseQuestionUploadApi.text]
                    Helpers.createAPILog(parentFolder, messageArr)
        return surveySolutionlink